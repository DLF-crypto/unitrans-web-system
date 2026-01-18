import os
import json
from datetime import datetime
from decimal import Decimal
from openpyxl import load_workbook, Workbook
from sqlalchemy import extract

from openpyxl.styles import Border, Side, Alignment, Font

def generate_supplier_invoices(year, month, db, models, invoice_folder):
    """
    根据年份和月份生成供应商对账单。
    """
    Waybill = models['Waybill']
    Supplier = models['Supplier']
    SupplierInvoice = models['SupplierInvoice']
    SupplierQuote = models['SupplierQuote']
    
    # 确定时间范围
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # 加载所有供应商报价
    all_quotes = SupplierQuote.query.all()

    def get_supplier_quote_split(wb):
        """查找并计算供应商报价中的快递费和挂号费（支持阶梯报价和最低计费重量）"""
        if not wb.supplier_id:
            return 0, 0
            
        for q in all_quotes:
            if (q.supplier_id == wb.supplier_id and 
                q.product_id == wb.product_id and
                q.valid_from <= wb.order_time <= q.valid_to):
                
                # 逻辑与 waybill_import_handler 一致
                weight = Decimal(str(wb.weight or 0))
                min_w = Decimal(str(q.min_weight)) if q.min_weight else Decimal(0)
                calc_weight = max(weight, min_w)
                
                price_tiers = json.loads(q.price_tiers) if q.price_tiers else []
                matched_tier = next((t for t in price_tiers if Decimal(str(t['start'])) < calc_weight <= Decimal(str(t['end']))), None)
                
                if matched_tier:
                    express = float(Decimal(str(matched_tier.get('express', 0))) * calc_weight)
                    reg = float(Decimal(str(matched_tier.get('reg', 0))))
                    return express, reg
                else:
                    # 如果没匹配到阶梯，尝试使用原有的固定报价作为兜底（虽然新逻辑下应该都有阶梯）
                    express = float(q.express_fee * calc_weight if q.express_fee else 0)
                    reg = float(q.registration_fee if q.registration_fee else 0)
                    return express, reg
        return 0, 0

    # 统一字体和对齐样式
    base_font = Font(name='黑体', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(border_style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 查询有该月份供应商成本的运单，并按供应商分组
    waybills = Waybill.query.filter(
        Waybill.order_time >= start_date,
        Waybill.order_time < end_date,
        Waybill.supplier_id != None,
        Waybill.supplier_cost > 0
    ).all()

    if not waybills:
        return 0

    # 按 supplier_id 分组
    groups = {}
    for wb in waybills:
        if wb.supplier_id not in groups:
            groups[wb.supplier_id] = []
        groups[wb.supplier_id].append(wb)

    suppliers_dict = {s.id: s for s in Supplier.query.all()}
    count = 0
    month_str = f"{year}{month:02d}"

    for sid, group_wbs in groups.items():
        supplier = suppliers_dict.get(sid)
        if not supplier:
            continue

        # 计算总成本
        total_cost = sum(float(wb.supplier_cost) for wb in group_wbs)
        
        # 文件名
        file_name = f"AP-{supplier.short_name}-{month_str}.xlsx"
        file_path = os.path.join(invoice_folder, file_name)

        # 创建新的 Excel
        wb_new = Workbook()
        ws = wb_new.active
        ws.title = "Details"

        # 表头
        headers = ["订单号", "转单号", "下单时间", "重量", "快递费", "挂号费", "其他收费", "总收费", "备注"]
        for c, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = Font(name='黑体', size=10, bold=True)
            cell.alignment = center_align
            cell.border = full_border

        # 填入数据
        for r, wb in enumerate(group_wbs, 2):
            express, reg = get_supplier_quote_split(wb)
            row_data = [
                wb.order_no,
                wb.transfer_no or "",
                wb.order_time.strftime("%Y-%m-%d %H:%M:%S"),
                float(wb.weight),
                express,
                reg,
                float(wb.other_fee or 0),
                float(wb.supplier_cost),
                wb.remark or ""
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = base_font
                cell.alignment = center_align
                cell.border = full_border
                
                # 数字格式
                if c == 4: # 重量
                    cell.number_format = '0.000'
                elif c in [5, 6, 7, 8]: # 费用
                    cell.number_format = '0.00'

        # 设置列宽
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 18

        wb_new.save(file_path)

        # 覆盖逻辑
        old_invoices = SupplierInvoice.query.filter_by(
            supplier_id=sid,
            year=year,
            month=month
        ).all()
        for old_inv in old_invoices:
            if old_inv.file_name and old_inv.file_name != file_name:
                old_path = os.path.join(invoice_folder, old_inv.file_name)
                if os.path.exists(old_path):
                    try: os.remove(old_path)
                    except: pass
            db.session.delete(old_inv)
        
        db.session.flush()

        # 插入新记录
        new_inv = SupplierInvoice(
            supplier_id=sid,
            year=year,
            month=month,
            amount=total_cost,
            file_name=file_name,
            created_at=datetime.utcnow()
        )
        db.session.add(new_inv)
        count += 1

    db.session.commit()
    return count


def generate_customer_invoices(year, month, db, models, invoice_folder):
    """
    根据年份和月份生成客户账单。
    """
    Waybill = models['Waybill']
    Product = models['Product']
    Customer = models['Customer']
    Invoice = models['Invoice']
    CustomerQuote = models['CustomerQuote']
    
    # 确定时间范围
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # 加载所有报价用于明细拆分
    all_quotes = CustomerQuote.query.all()

    def get_quote_split(wb, quote_type):
        """拆分报价中的快递费和挂号费"""
        cid = None
        if quote_type == "尾程报价":
            cid = wb.last_leg_customer_id
        
        if not cid:
            return 0, 0
            
        # 查找匹配的报价
        for q in all_quotes:
            if (q.customer_id == cid and 
                q.quote_type == quote_type and
                q.valid_from <= wb.order_time <= q.valid_to):
                
                # 检查产品ID维度
                if q.product_ids:
                    p_list = q.product_ids.split(",")
                    if str(wb.product_id) not in p_list:
                        continue
                
                return float(q.express_fee * wb.weight if q.express_fee else 0), float(q.registration_fee if q.registration_fee else 0)
        return 0, 0

    def get_quote_unit_price(wb, fee_label):
        """从报价单中直接获取维护的单价，避免除法产生的长小数"""
        cid = None
        if fee_label == '头程收费':
            cid = wb.first_leg_customer_id
        elif fee_label == '单号收费':
            cid = wb.unit_customer_id
        
        if not cid: return 0
        
        target_type = fee_label.replace('收费', '报价')
        for q in all_quotes:
            if (q.customer_id == cid and q.quote_type == target_type and
                q.valid_from <= wb.order_time <= q.valid_to):
                
                # 检查产品ID维度
                if target_type == '头程报价' and q.product_ids:
                    p_list = q.product_ids.split(",")
                    if str(wb.product_id) not in p_list:
                        continue
                
                if fee_label == '头程收费': return float(q.air_freight or 0)
                if fee_label == '单号收费': return float(q.unit_fee or 0)
        return 0

    # 边框样式
    thin_side = Side(border_style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 统一字体：黑体，10号
    base_font = Font(name='黑体', size=10)
    center_align = Alignment(horizontal='center', vertical='center')

    def safe_set_value(ws, cell_ref, value, is_header=False, apply_style=True, is_detail=False):
        """安全地设置单元格值，跳过只读的合并单元格，并应用字体和对齐"""
        if isinstance(cell_ref, str):
            cell = ws[cell_ref]
        else:
            cell = ws.cell(row=cell_ref[0], column=cell_ref[1])
            
        if type(cell).__name__ != 'MergedCell':
            cell.value = value
            # Summary 页面从第5行起，或 Detail 页面全页面应用黑体10号
            if apply_style and (is_detail or cell.row >= 5):
                is_bold = cell.font.bold if cell.font else False
                cell.font = Font(name='黑体', size=10, bold=is_bold)
                # 表格区域应用居中（Summary第9行起，Detail全页面）
                if is_detail or cell.row >= 9:
                    cell.alignment = center_align
            return cell
        else:
            # 如果是合并单元格，我们需要通过 ws[ref] 的方式找到主单元格来赋值
            # 在 openpyxl 中，对合并单元格的第一个格子赋值是允许的
            # 这里的 cell_ref 如果是 'A5' 这种字符串，或者 (5, 1) 这种坐标
            try:
                if isinstance(cell_ref, str):
                    main_cell = ws[cell_ref]
                else:
                    main_cell = ws.cell(row=cell_ref[0], column=cell_ref[1])
                main_cell.value = value
                if apply_style and (is_detail or main_cell.row >= 5):
                    is_bold = main_cell.font.bold if main_cell.font else False
                    main_cell.font = Font(name='黑体', size=10, bold=is_bold)
                    if is_detail or main_cell.row >= 9:
                        main_cell.alignment = center_align
                return main_cell
            except:
                pass
        return None

    # 收费类别配置映射
    # (Waybill字段名, 收费类别名称, 客户ID字段名)
    CONFIGS = [
        ('unit_fee', '单号收费', 'unit_customer_id'),
        ('first_leg_fee', '头程收费', 'first_leg_customer_id'),
        ('last_leg_fee', '尾程收费', 'last_leg_customer_id'),
        ('differential_fee', '差价收费', 'differential_customer_id'),
    ]

    count = 0
    now_str = datetime.now().strftime("%Y-%m-%d")
    month_str = f"{year}{month:02d}"

    # 为了性能，预先加载所有产品和客户
    products_dict = {p.id: p for p in Product.query.all()}
    customers_dict = {c.id: c for c in Customer.query.all()}

    for fee_field, fee_label, cust_id_field in CONFIGS:
        # 查询有该项费用的运单，并按客户分组
        waybills = Waybill.query.filter(
            Waybill.order_time >= start_date,
            Waybill.order_time < end_date,
            getattr(Waybill, fee_field) > 0,
            getattr(Waybill, cust_id_field) != None
        ).all()

        if not waybills:
            continue

        # 按 customer_id 分组
        groups = {}
        for wb in waybills:
            cid = getattr(wb, cust_id_field)
            if cid not in groups:
                groups[cid] = []
            groups[cid].append(wb)

        # 为每个客户生成一个新 Excel 账单 (不再套用模板)
        for cid, group_wbs in groups.items():
            customer = customers_dict.get(cid)
            if not customer:
                continue

            # 计算总金额
            total_amount = sum(float(getattr(wb, fee_field)) for wb in group_wbs)
            
            # 生成文件名和账单号
            invoice_no = f"TKTX-{customer.short_name}-{month_str}"
            file_name = f"{invoice_no}-{fee_label}.xlsx"
            file_path = os.path.join(invoice_folder, file_name)

            # 创建新的 Workbook
            wb_new = Workbook()
            
            # --- 1. Summary Sheet ---
            ws_sum = wb_new.active
            ws_sum.title = "Summary"
            
            # 基础样式
            title_font = Font(name='黑体', size=14, bold=True)
            header_font = Font(name='黑体', size=10, bold=True)
            normal_font = Font(name='黑体', size=10)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center', wrapText=True)
            thin_side = Side(border_style="thin", color="000000")
            full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            # 标题与基本信息
            ws_sum.merge_cells('A1:E1')
            ws_sum['A1'] = f"应收账单汇总 - {fee_label}"
            ws_sum['A1'].font = title_font
            ws_sum['A1'].alignment = center_align

            # 客户信息
            ws_sum['A3'] = "客户全称:"
            ws_sum['B3'] = customer.full_name
            ws_sum['A4'] = "联 系 人:"
            ws_sum['B4'] = customer.contact_person or "-"
            ws_sum['A5'] = "联系邮箱:"
            ws_sum['B5'] = customer.email or "-"
            
            for r in range(3, 6):
                ws_sum.cell(row=r, column=1).font = header_font
                ws_sum.cell(row=r, column=2).font = normal_font

            # 账单信息
            ws_sum['D3'] = "账单编号:"
            ws_sum['E3'] = invoice_no
            ws_sum['D4'] = "生成日期:"
            ws_sum['E4'] = now_str
            ws_sum['D5'] = "账单周期:"
            ws_sum['E5'] = f"{year}年{month}月"
            
            for r in range(3, 6):
                ws_sum.cell(row=r, column=4).font = header_font
                ws_sum.cell(row=r, column=5).font = normal_font

            # 表格表头
            if fee_label in ['差价收费', '尾程收费']:
                headers = ["序号", "项目名称", "件数", "重量", "金额"]
                ws_sum.column_dimensions['A'].width = 8
                ws_sum.column_dimensions['B'].width = 35
                ws_sum.column_dimensions['C'].width = 12
                ws_sum.column_dimensions['D'].width = 15
                ws_sum.column_dimensions['E'].width = 20
            else:
                headers = ["序号", "项目名称", "件数/重量", "单价", "金额"]
                ws_sum.column_dimensions['A'].width = 8
                ws_sum.column_dimensions['B'].width = 35
                ws_sum.column_dimensions['C'].width = 15
                ws_sum.column_dimensions['D'].width = 15
                ws_sum.column_dimensions['E'].width = 20

            for c, h in enumerate(headers, 1):
                cell = ws_sum.cell(row=8, column=c, value=h)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = full_border

            # 汇总逻辑
            summary_data = []
            if fee_label == '单号收费':
                p_groups = {}
                for wb in group_wbs:
                    pid = wb.product_id
                    if pid not in p_groups: p_groups[pid] = {'count': 0, 'fee': 0, 'price': wb.unit_fee}
                    p_groups[pid]['count'] += 1
                    p_groups[pid]['fee'] += float(wb.unit_fee)
                for pid, d in p_groups.items():
                    prod = products_dict.get(pid)
                    summary_data.append([f"{fee_label}({prod.name if prod else '未知'})", d['count'], float(d['price']), d['fee']])
            
            elif fee_label == '头程收费':
                p_groups = {}
                for wb in group_wbs:
                    pid = wb.product_id
                    if pid not in p_groups:
                        price = get_quote_unit_price(wb, fee_label)
                        p_groups[pid] = {'weight': 0, 'fee': 0, 'price': price}
                    p_groups[pid]['weight'] += float(wb.weight)
                    p_groups[pid]['fee'] += float(wb.first_leg_fee)
                for pid, d in p_groups.items():
                    prod = products_dict.get(pid)
                    summary_data.append([f"{fee_label}({prod.name if prod else '未知'})", round(d['weight'], 3), float(d['price']), d['fee']])

            elif fee_label in ['差价收费', '尾程收费']:
                p_groups = {}
                for wb in group_wbs:
                    pid = wb.product_id
                    if pid not in p_groups: p_groups[pid] = {'count': 0, 'weight': 0, 'fee': 0}
                    p_groups[pid]['count'] += 1
                    p_groups[pid]['weight'] += float(wb.weight)
                    p_groups[pid]['fee'] += float(getattr(wb, fee_field))
                for pid, d in p_groups.items():
                    prod = products_dict.get(pid)
                    summary_data.append([f"{fee_label}({prod.name if prod else '未知'})", d['count'], round(d['weight'], 3), d['fee']])
            
            # 填充数据行
            current_row = 9
            for idx, row in enumerate(summary_data, 1):
                data_row = [idx] + row
                for c, val in enumerate(data_row, 1):
                    cell = ws_sum.cell(row=current_row, column=c, value=val)
                    cell.font = normal_font
                    cell.alignment = center_align
                    cell.border = full_border
                    # 动态判断数字格式
                    if fee_label in ['差价收费', '尾程收费']:
                        if c == 4: cell.number_format = '0.000' # 重量
                        if c == 5: cell.number_format = '0.00'  # 金额
                    else:
                        if c == 5 or (isinstance(val, (float, int)) and c == 4):
                            cell.number_format = '0.00'
                current_row += 1

            # 总计行
            ws_sum.cell(row=current_row, column=4, value="总计金额:").font = header_font
            ws_sum.cell(row=current_row, column=4).alignment = center_align
            ws_sum.cell(row=current_row, column=4).border = full_border
            
            c_total = ws_sum.cell(row=current_row, column=5, value=total_amount)
            c_total.font = Font(name='黑体', size=11, bold=True)
            c_total.alignment = center_align
            c_total.border = full_border
            c_total.number_format = '0.00'
            current_row += 2

            # 付款信息
            if fee_label in ['单号收费', '差价收费', '头程收费']:
                payment_info = "请付款至以下账号：\n开户银行: 平安银行深圳宝城支行\n开户名称：深圳市淘客天下贸易有限公司\n账户：15346748610064"
            else:
                payment_info = "请付款至以下账号：\n公司名称：易泰通供应链（深圳）有限公司\n开户银行：建设银行高新园支行\n银行账号：44250100004800000716"
            
            ws_sum.merge_cells(start_row=current_row, start_column=1, end_row=current_row+3, end_column=3)
            cell_pay = ws_sum.cell(row=current_row, column=1, value=payment_info)
            cell_pay.font = normal_font
            cell_pay.alignment = left_align
            
            extra_info = "如有疑问，请联系对账组：\n邮箱: acc@postchain56.com\n感谢您的支持！"
            ws_sum.merge_cells(start_row=current_row, start_column=4, end_row=current_row+3, end_column=5)
            cell_extra = ws_sum.cell(row=current_row, column=4, value=extra_info)
            cell_extra.font = normal_font
            cell_extra.alignment = left_align

            # 设置列宽
            ws_sum.column_dimensions['A'].width = 8
            ws_sum.column_dimensions['B'].width = 35
            ws_sum.column_dimensions['C'].width = 15
            ws_sum.column_dimensions['D'].width = 15
            ws_sum.column_dimensions['E'].width = 20

            # --- 2. Details Sheet ---
            ws_det = wb_new.create_sheet(title="Details")
            
            if fee_label == '头程收费':
                headers = ["订单号", "转单号", "产品名称", "下单日期", "重量(kg)", "单价", "总金额"]
            elif fee_label == '单号收费':
                headers = ["订单号", "下单日期", "产品名称", "单价", "金额"]
            elif fee_label == '差价收费':
                headers = ["订单号", "转单号", "产品名称", "下单时间", "重量(kg)", "其他收费", "差价金额"]
            else: # 尾程
                headers = ["订单号", "转单号", "产品名称", "下单日期", "重量(kg)", "快递费", "挂号费", "其他费用", "总费用"]

            for c, h in enumerate(headers, 1):
                cell = ws_det.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = full_border

            for i, wb in enumerate(group_wbs, 2):
                prod = products_dict.get(wb.product_id)
                p_name = prod.name if prod else "-"
                
                if fee_label == '头程收费':
                    price = get_quote_unit_price(wb, fee_label)
                    row = [wb.order_no, wb.transfer_no, p_name, wb.order_time.strftime("%Y-%m-%d"), float(wb.weight), price, float(wb.first_leg_fee)]
                elif fee_label == '单号收费':
                    row = [wb.order_no, wb.order_time.strftime("%Y-%m-%d"), p_name, float(wb.unit_fee), float(wb.unit_fee)]
                elif fee_label == '差价收费':
                    row = [wb.order_no, wb.transfer_no, p_name, wb.order_time.strftime("%Y-%m-%d"), float(wb.weight), float(wb.other_fee or 0), float(wb.differential_fee)]
                else: # 尾程
                    express, reg = get_quote_split(wb, "尾程报价")
                    row = [wb.order_no, wb.transfer_no, p_name, wb.order_time.strftime("%Y-%m-%d"), float(wb.weight), express, reg, float(wb.other_fee or 0), float(wb.last_leg_fee)]
                
                for c, val in enumerate(row, 1):
                    cell = ws_det.cell(row=i, column=c, value=val)
                    cell.font = normal_font
                    cell.alignment = center_align
                    cell.border = full_border
                    if isinstance(val, float):
                        # 判断该列是否为重量
                        is_weight_col = "重量" in headers[c-1]
                        cell.number_format = '0.000' if is_weight_col else '0.00'

            for c in range(1, len(headers) + 1):
                ws_det.column_dimensions[ws_det.cell(row=1, column=c).column_letter].width = 18

            wb_new.save(file_path)

            # 覆盖逻辑
            old_invoices = Invoice.query.filter_by(
                customer_id=cid,
                fee_type=fee_label,
                year=year,
                month=month
            ).all()
            for old_inv in old_invoices:
                if old_inv.file_name and old_inv.file_name != file_name:
                    old_path = os.path.join(invoice_folder, old_inv.file_name)
                    if os.path.exists(old_path):
                        try: os.remove(old_path)
                        except: pass
                db.session.delete(old_inv)
            
            db.session.flush()
            new_invoice = Invoice(
                customer_id=cid, fee_type=fee_label, year=year, month=month,
                amount=total_amount, file_name=file_name, created_at=datetime.utcnow()
            )
            db.session.add(new_invoice)
            count += 1

    db.session.commit()
    return count
