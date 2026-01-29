from datetime import datetime, timedelta
from sqlalchemy import func, extract
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import os
load_dotenv()
import json
from flask import Flask, render_template, redirect, url_for, request, jsonify, session, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from celery import Celery
from waybill_import_handler import (
    validate_and_process_waybill_import,
    calculate_waybill_fees
)
from invoice_handler import generate_customer_invoices, generate_supplier_invoices

app = Flask(__name__)
app.secret_key = "dev-change-this-secret"  # 用于会话管理，后续可根据需要修改

# Celery 配置
app.config['broker_url'] = 'redis://127.0.0.1:6379/0'
app.config['result_backend'] = 'redis://127.0.0.1:6379/0'

def make_celery(app):
    celery = Celery(
        app.import_name,
        backend=app.config.get('result_backend', 'redis://127.0.0.1:6379/0'),
        broker=app.config.get('broker_url', 'redis://127.0.0.1:6379/0')
    )
    celery.conf.update(app.config)

    class ContextTask(celery.Task):
        def __call__(self, *args, **kwargs):
            with app.app_context():
                return self.run(*args, **kwargs)

    celery.Task = ContextTask
    return celery

celery = make_celery(app)

# 配置Celery Beat定时任务
from celery.schedules import crontab

celery.conf.beat_schedule = {
    # 每小时执行一次自动获取轨迹
    'auto-fetch-tracking-hourly': {
        'task': 'app.auto_fetch_tracking_task',
        'schedule': crontab(minute=0),  # 每小时的0分钟执行
    },
}

celery.conf.timezone = 'UTC'

# Celery任务路由配置（注释掉，使用默认celery队列）
# celery.conf.task_routes = {
#     'app.async_fetch_tracking_task': {'queue': 'tracking'},
#     'app.async_fetch_lastmile_tracking_task': {'queue': 'tracking'},
# }

# 并发控制
celery.conf.worker_prefetch_multiplier = 1  # 每次只预取一个任务
celery.conf.task_acks_late = True  # 任务执行完才确认

# 上传文件配置
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB最大文件大小

# 账单存储配置
INVOICE_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'customer invoices')
if not os.path.exists(INVOICE_FOLDER):
    os.makedirs(INVOICE_FOLDER)
app.config['INVOICE_FOLDER'] = INVOICE_FOLDER

SUPPLIER_INVOICE_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'supplier invoices')
if not os.path.exists(SUPPLIER_INVOICE_FOLDER):
    os.makedirs(SUPPLIER_INVOICE_FOLDER)
app.config['SUPPLIER_INVOICE_FOLDER'] = SUPPLIER_INVOICE_FOLDER

# 数据库配置 - 使用环境变量
DB_USER = os.environ.get('DB_USER', 'root')  # 本地使用 root
DB_PASSWORD = os.environ.get('DB_PASSWORD', '123456')  # 本地使用 123456
DB_HOST = os.environ.get('DB_HOST', '127.0.0.1')
DB_PORT = os.environ.get('DB_PORT', '3308')  # 本地使用 3308
DB_NAME = os.environ.get('DB_NAME', 'unitransDB')

app.config["SQLALCHEMY_DATABASE_URI"] = (
    f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class TaskRecord(db.Model):
    __tablename__ = "task_records"

    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.String(64), unique=True, nullable=False)  # Celery 任务 ID
    task_name = db.Column(db.String(128), nullable=False)
    status = db.Column(db.String(32), default="PENDING")  # PENDING, PROCESSING, SUCCESS, FAILURE
    result_msg = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class Role(db.Model):
    __tablename__ = "roles"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=True, nullable=False)
    description = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    permissions = db.relationship(
        "RolePagePermission",
        backref="role",
        cascade="all, delete-orphan",
        lazy="joined",
    )


class RolePagePermission(db.Model):
    __tablename__ = "role_page_permissions"

    id = db.Column(db.Integer, primary_key=True)
    role_id = db.Column(db.Integer, db.ForeignKey("roles.id"), nullable=False)
    page_key = db.Column(db.String(64), nullable=False)
    can_view = db.Column(db.Boolean, default=False, nullable=False)  # 可查看（控制菜单显示）
    can_create = db.Column(db.Boolean, default=False, nullable=False)
    can_update = db.Column(db.Boolean, default=False, nullable=False)
    can_delete = db.Column(db.Boolean, default=False, nullable=False)
    # 存储字段级权限的JSON数据
    field_permissions = db.Column(db.Text, nullable=True)  # JSON格式存储字段权限


class User(db.Model):
    __tablename__ = "users"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    password = db.Column(db.String(128), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey("roles.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    role = db.relationship("Role", backref="users", lazy="joined")


class Country(db.Model):
    __tablename__ = "countries"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False)
    code = db.Column(db.String(2), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Product(db.Model):
    __tablename__ = "products"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False)
    description = db.Column(db.String(300))  # 产品描述，允许为空
    # 收费类别：用逗号分隔，如：“单号收费,头程收费”
    fee_types = db.Column(db.String(255), nullable=False)
    supplier_id = db.Column(db.Integer, db.ForeignKey("suppliers.id"))  # 绑定的供应商ID（仅当有差价收费时）
    tracking_interface_id = db.Column(db.Integer, db.ForeignKey("tracking_interfaces.id"))  # 绑定的轨迹接口ID（仅当有尾程收费时）
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    # 关联供应商和轨迹接口
    supplier = db.relationship("Supplier", backref="products", lazy="joined")
    tracking_interface = db.relationship("TrackingInterface", backref="products", lazy="joined")


class Customer(db.Model):
    __tablename__ = "customers"

    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(128), nullable=False)  # 客户全称
    short_name = db.Column(db.String(64), nullable=False)  # 客户简称
    # 客户类别：用逗号分隔，如："单号客户,头程客户"
    customer_types = db.Column(db.String(255), nullable=False)
    contact_person = db.Column(db.String(64))  # 联系人
    email = db.Column(db.String(128))  # 邮箱
    remark = db.Column(db.String(500))  # 备注
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Supplier(db.Model):
    __tablename__ = "suppliers"

    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(128), nullable=False)  # 供应商全称
    short_name = db.Column(db.String(64), nullable=False)  # 供应商简称
    contact_person = db.Column(db.String(64))  # 联系人
    email = db.Column(db.String(128))  # 邮箱
    remark = db.Column(db.String(500))  # 备注
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class TrackingNode(db.Model):
    """轨迹节点状态管理"""
    __tablename__ = "tracking_nodes"

    id = db.Column(db.Integer, primary_key=True)
    status_code = db.Column(db.String(32), unique=True, nullable=False)  # 状态代码
    status_description = db.Column(db.String(128), nullable=False)  # 状态说明
    default_city = db.Column(db.String(64))  # 默认城市
    default_country_code = db.Column(db.String(3))  # 默认国家代码（如：CN、US）
    default_airport_code = db.Column(db.String(3))  # 默认机场三字代码（如：PVG、LAX）
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class TrackingInterface(db.Model):
    """轨迹接口管理"""
    __tablename__ = "tracking_interfaces"

    id = db.Column(db.Integer, primary_key=True)
    interface_name = db.Column(db.String(128), unique=True, nullable=False)  # 轨迹接口名称
    request_url = db.Column(db.String(512), nullable=False)  # 轨迹请求地址
    auth_params = db.Column(db.Text)  # 轨迹接口验证信息，JSON格式：{"api_id":"123456","key":"xxxx"}
    status_mapping = db.Column(db.Text)  # 头程状态映射表，JSON格式：[{"supplier_status":"xxx","supplier_description":"","system_status_code":"xxx"},...]
    response_key_params = db.Column(db.Text)  # 关键信息代码参数，JSON格式：{"time_key":"changeDate","status_key":"status","description_key":"record","city_key":"city","country_key":"country"}
    fetch_interval = db.Column(db.Numeric(5, 2), nullable=False)  # 获取频率（小时），如 1.2 表示 1小时12分钟
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class LastmileStatusMapping(db.Model):
    """尾程轨迹状态映射表（独立表）"""
    __tablename__ = "lastmile_status_mappings"

    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(255), nullable=True, index=True, default='')  # 尾程轨迹描述（对应报文的"description"），非必填
    sub_status = db.Column(db.String(64), nullable=False, index=True)  # 尾程轨迹状态（对应报文的"sub_status"）
    system_status_code = db.Column(db.String(32), nullable=False)  # 系统状态代码
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    # 不再使用唯一约束，因为 description 可以为空，后端代码会手动检查重复


class TrackingInfo(db.Model):
    """轨迹信息存储"""
    __tablename__ = "tracking_info"

    id = db.Column(db.Integer, primary_key=True)
    waybill_id = db.Column(db.Integer, db.ForeignKey("waybills.id"), nullable=False, index=True)  # 运单ID
    order_no = db.Column(db.String(64), nullable=False, index=True)  # 订单号
    transfer_no = db.Column(db.String(64), index=True)  # 转单号
    tracking_interface_id = db.Column(db.Integer, db.ForeignKey("tracking_interfaces.id"), nullable=False)  # 轨迹接口ID
    
    tracking_description = db.Column(db.Text)  # 轨迹描述
    status_code = db.Column(db.String(32))  # 轨迹状态代码（系统状态代码）
    tracking_time = db.Column(db.DateTime)  # 时间节点
    
    raw_response = db.Column(db.Text)  # 接口原始报文(JSON格式)
    
    last_fetch_time = db.Column(db.DateTime)  # 最新获取时间（从供应商接口获取）
    last_push_time = db.Column(db.DateTime)  # 最新推送时间（推送到上家）
    
    # 停止自动跟踪相关字段
    stop_tracking = db.Column(db.Boolean, default=False)  # 是否停止自动跟踪
    stop_tracking_reason = db.Column(db.String(255))  # 停止跟踪原因
    stop_tracking_time = db.Column(db.DateTime)  # 停止跟踪时间
    
    # 尾程轨迹相关字段
    lastmile_no = db.Column(db.String(64), index=True)  # 尾程单号
    lastmile_raw_response = db.Column(db.Text)  # 尾程接口原始报文(JSON格式)，兼容字段
    lastmile_last_fetch_time = db.Column(db.DateTime)  # 尾程最新获取时间
    lastmile_register_response = db.Column(db.Text(length=4294967295))  # 尾程注册报文(JSON格式) - LONGTEXT
    lastmile_tracking_response = db.Column(db.Text(length=4294967295))  # 尾程单号报文(JSON格式) - LONGTEXT
    
    # 推送报文（JSON数组，存储所有轨迹节点）
    push_events = db.Column(db.Text(length=4294967295))  # 推送报文(JSON格式) - LONGTEXT
    szpost_response = db.Column(db.Text(length=4294967295))  # 深邮响应报文(JSON格式) - LONGTEXT
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    # 关联关系
    waybill = db.relationship("Waybill", backref="tracking_info", lazy="joined")
    tracking_interface = db.relationship("TrackingInterface", backref="tracking_records", lazy="joined")


class CustomerQuote(db.Model):
    __tablename__ = "customer_quotes"

    id = db.Column(db.Integer, primary_key=True)
    quote_name = db.Column(db.String(128), unique=True, nullable=False)  # 报价名称
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)  # 客户ID
    quote_type = db.Column(db.String(32), nullable=False)  # 报价类别：单号报价/头程报价/尾程报价
    product_ids = db.Column(db.String(255), nullable=True)  # 产品ID集合（逗号分隔）
    
    # 报价明细（根据类型不同，使用不同字段）
    unit_fee = db.Column(db.Numeric(10, 2))  # 单号费（元/单）
    air_freight = db.Column(db.Numeric(10, 2))  # 空运费（元/kg）
    express_fee = db.Column(db.Numeric(10, 2))  # 快递费（元/kg）
    registration_fee = db.Column(db.Numeric(10, 2))  # 挂号费（元/单）
    
    # 专线处理费
    dedicated_line_weight_fee = db.Column(db.Numeric(10, 2))  # 重量收费（元/kg）
    dedicated_line_piece_fee = db.Column(db.Numeric(10, 2))  # 单件收费（元/件）
    
    # 有效期
    valid_from = db.Column(db.DateTime, nullable=False)  # 开始时间
    valid_to = db.Column(db.DateTime, nullable=False)  # 结束时间
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    # 关联客户和产品
    customer = db.relationship("Customer", backref="quotes", lazy="joined")


class SupplierQuote(db.Model):
    __tablename__ = "supplier_quotes"

    id = db.Column(db.Integer, primary_key=True)
    quote_name = db.Column(db.String(128), unique=True, nullable=False)  # 报价名称
    supplier_id = db.Column(db.Integer, db.ForeignKey("suppliers.id"), nullable=False)  # 供应商ID
    product_id = db.Column(db.Integer, db.ForeignKey("products.id"), nullable=False)  # 产品ID
    
    # 新阶梯报价逻辑：旧的 express_fee/registration_fee 保留但不主要使用，或者作为默认值
    express_fee = db.Column(db.Numeric(10, 2), nullable=True)  # 快递费（元/kg）
    registration_fee = db.Column(db.Numeric(10, 2), nullable=True)  # 挂号费（元/单）
    
    # 阶梯报价扩展字段
    min_weight = db.Column(db.Numeric(10, 3), default=0)  # 最低计费重量
    price_tiers = db.Column(db.Text)  # 价格阶梯，JSON 字符串存储：[{"start": 0, "end": 0.1, "express": 86, "reg": 19}, ...]
    
    # 有效期
    valid_from = db.Column(db.DateTime, nullable=False)  # 开始时间
    valid_to = db.Column(db.DateTime, nullable=False)  # 结束时间
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    # 关联供应商和产品
    supplier = db.relationship("Supplier", backref="quotes", lazy="joined")
    product = db.relationship("Product", backref="supplier_quotes", lazy="joined")


class Waybill(db.Model):
    __tablename__ = "waybills"

    id = db.Column(db.Integer, primary_key=True)
    order_no = db.Column(db.String(64), unique=True, nullable=False)  # 订单号
    transfer_no = db.Column(db.String(64))  # 转单号
    weight = db.Column(db.Numeric(10, 3), default=0)  # 重量（kg，保疙3位小数）
    order_time = db.Column(db.DateTime, nullable=False, index=True)  # 下单时间
    product_id = db.Column(db.Integer, db.ForeignKey("products.id"), index=True)  # 产品ID
    
    # 客户信息
    unit_customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), index=True)  # 单号客户ID
    first_leg_customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), index=True)  # 头程客户ID
    last_leg_customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), index=True)  # 尾程客户ID
    differential_customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), index=True)  # 差价客户ID
    
    supplier_id = db.Column(db.Integer, db.ForeignKey("suppliers.id"))  # 供应商ID
    
    # 费用信息
    unit_fee = db.Column(db.Numeric(10, 2), default=0)  # 单号收费
    first_leg_fee = db.Column(db.Numeric(10, 2), default=0)  # 头程收费
    last_leg_fee = db.Column(db.Numeric(10, 2), default=0)  # 尾程收费
    differential_fee = db.Column(db.Numeric(10, 2), default=0)  # 差价收费
    dedicated_line_fee = db.Column(db.Numeric(10, 2), default=0)  # 专线处理费
    supplier_cost = db.Column(db.Numeric(10, 2), default=0)  # 供应商成本
    other_fee = db.Column(db.Numeric(10, 2), default=0)  # 其他费用
    
    remark = db.Column(db.Text)  # 备注
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    # 关联关系
    product = db.relationship("Product", backref="waybills", lazy="joined")
    unit_customer = db.relationship("Customer", foreign_keys=[unit_customer_id], backref="unit_waybills", lazy="joined")
    first_leg_customer = db.relationship("Customer", foreign_keys=[first_leg_customer_id], backref="first_leg_waybills", lazy="joined")
    last_leg_customer = db.relationship("Customer", foreign_keys=[last_leg_customer_id], backref="last_leg_waybills", lazy="joined")
    differential_customer = db.relationship("Customer", foreign_keys=[differential_customer_id], backref="differential_waybills", lazy="joined")
    supplier = db.relationship("Supplier", backref="waybills", lazy="joined")


class Invoice(db.Model):
    __tablename__ = "invoices"

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)
    fee_type = db.Column(db.String(64), nullable=False)  # 单号收费, 头程收费, 尾程收费, 差价收费
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    amount = db.Column(db.Numeric(10, 2), default=0)
    file_name = db.Column(db.String(255))
    is_paid = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    customer = db.relationship("Customer", backref="invoices", lazy="joined")


class SupplierInvoice(db.Model):
    __tablename__ = "supplier_invoices"

    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey("suppliers.id"), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    amount = db.Column(db.Numeric(10, 2), default=0)
    file_name = db.Column(db.String(255))
    is_paid = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    supplier = db.relationship("Supplier", backref="invoices", lazy="joined")

class Payment(db.Model):
    __tablename__ = "payments"

    id = db.Column(db.Integer, primary_key=True)
    target_type = db.Column(db.String(20), nullable=False)  # 'customer' or 'supplier'
    target_id = db.Column(db.Integer, nullable=False)
    payment_type = db.Column(db.String(10), nullable=False)  # '收款' or '付款'
    payment_date = db.Column(db.DateTime, nullable=False)
    amount = db.Column(db.Numeric(12, 2), default=0, nullable=False)
    receipt_path = db.Column(db.String(255))  # 水单图片路径
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoices.id'))  # 关联应收账单
    supplier_invoice_id = db.Column(db.Integer, db.ForeignKey('supplier_invoices.id'))  # 关联应付账单
    remark = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    @property
    def target_name(self):
        if self.target_type == 'customer':
            c = Customer.query.get(self.target_id)
            return c.full_name if c else "未知客户"
        else:
            s = Supplier.query.get(self.target_id)
            return s.full_name if s else "未知供应商"


# 简单的内存用户与角色模型（后续可迁移到数据库）
USERS = {
    "admin": {
        "password": "123456",
        "role": "系统管理员",
        "permissions": "ALL",  # 默认拥有所有页面的权限
    }
}


# ==================== 停止跟踪相关函数 ====================

def should_stop_tracking(waybill, tracking_info):
    """
    检查是否应该停止跟踪
    
    停止条件：
    1. 运单导入系统超过45天
    2. 运单状态代码是O_016
    3. 运单停止更新超过20天
    
    Args:
        waybill: Waybill对象
        tracking_info: TrackingInfo对象，可能为None
    
    Returns:
        tuple: (should_stop: bool, reason: str or None)
    """
    now = datetime.utcnow()
    
    # 条件1：运单导入系统超过45天
    if waybill.created_at:
        days_since_import = (now - waybill.created_at).days
        if days_since_import > 45:
            return True, f"运单导入超过45天（{days_since_import}天）"
    
    # 条件2：状态代码是O_016
    if tracking_info and tracking_info.status_code == 'O_016':
        return True, "状态代码为O_016"
    
    # 条件3：停止更新超过20天
    if tracking_info and tracking_info.updated_at:
        days_since_update = (now - tracking_info.updated_at).days
        if days_since_update > 20:
            return True, f"停止更新超过20天（{days_since_update}天）"
    
    return False, None


def batch_check_stop_tracking():
    """
    批量检查所有运单，标记满足停止条件的运单
    该函数可用于定时任务
    
    Returns:
        dict: 包含处理结果的字典
    """
    stopped_count = 0
    checked_count = 0
    
    # 查询所有未停止跟踪的运单
    query = db.session.query(
        Waybill,
        TrackingInfo
    ).outerjoin(
        TrackingInfo, Waybill.id == TrackingInfo.waybill_id
    ).filter(
        db.or_(
            TrackingInfo.stop_tracking == False,
            TrackingInfo.stop_tracking.is_(None)
        )
    )
    
    for waybill, tracking_info in query.all():
        checked_count += 1
        should_stop, reason = should_stop_tracking(waybill, tracking_info)
        
        if should_stop:
            if tracking_info:
                # 更新现有轨迹记录
                tracking_info.stop_tracking = True
                tracking_info.stop_tracking_reason = reason
                tracking_info.stop_tracking_time = datetime.utcnow()
                stopped_count += 1
            else:
                # 如果还没有轨迹记录，也需要创建一个标记为停止
                # 这种情况可能是运单导入后从未获取过轨迹，但已超过45天
                # 获取产品的轨迹接口ID
                if waybill.product and waybill.product.tracking_interface_id:
                    new_tracking = TrackingInfo(
                        waybill_id=waybill.id,
                        order_no=waybill.order_no,
                        transfer_no=waybill.transfer_no,
                        tracking_interface_id=waybill.product.tracking_interface_id,
                        stop_tracking=True,
                        stop_tracking_reason=reason,
                        stop_tracking_time=datetime.utcnow()
                    )
                    db.session.add(new_tracking)
                    stopped_count += 1
    
    try:
        db.session.commit()
        return {
            "success": True,
            "checked_count": checked_count,
            "stopped_count": stopped_count,
            "message": f"检查了 {checked_count} 条运单，标记停止跟踪 {stopped_count} 条"
        }
    except Exception as e:
        db.session.rollback()
        return {
            "success": False,
            "message": f"批量检查失败: {str(e)}"
        }


# ==================== Celery 异步任务 ====================

@celery.task(bind=True, name='app.async_generate_customer_invoices')
def async_generate_customer_invoices(self, year, month, customer_id=None):
    """异步生成客户账单任务"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"========== 开始执行任务: {self.request.id} ==========")
    
    with app.app_context():
        try:
            logger.info("步骤 1: 查询任务记录")
            task_record = TaskRecord.query.filter_by(task_id=self.request.id).first()
            if task_record:
                logger.info("步骤 2: 更新任务状态为 PROCESSING")
                task_record.status = "PROCESSING"
                db.session.commit()
            else:
                logger.warning("未找到任务记录")
            
            logger.info(f"步骤 3: 开始生成账单 {year}-{month}, customer_id={customer_id}")
            invoice_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'customer invoices')
            logger.info(f"账单目录: {invoice_folder}")
            
            count = generate_customer_invoices(int(year), int(month), db, {
                'Waybill': Waybill,
                'Product': Product,
                'Customer': Customer,
                'Invoice': Invoice,
                'CustomerQuote': CustomerQuote
            }, invoice_folder, customer_id=customer_id)
            
            logger.info(f"步骤 4: 账单生成完成，数量: {count}")
            
            if task_record:
                task_record.status = "SUCCESS"
                task_record.result_msg = f"成功生成了 {count} 份账单"
                db.session.commit()
                logger.info("步骤 5: 任务状态已更新为 SUCCESS")
            
            logger.info("========== 任务执行成功 ==========")
            return {"success": True, "count": count}
            
        except Exception as e:
            logger.error(f"========== 任务执行失败: {str(e)} ==========")
            logger.exception("详细错误信息:")
            if task_record:
                task_record.status = "FAILURE"
                task_record.result_msg = str(e)
                db.session.commit()
            raise e

@celery.task(bind=True, name='app.async_generate_supplier_invoices')
def async_generate_supplier_invoices(self, year, month):
    """异步生成供应商账单任务"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"========== 开始执行供应商账单任务: {self.request.id} ==========")
    
    with app.app_context():
        try:
            task_record = TaskRecord.query.filter_by(task_id=self.request.id).first()
            if task_record:
                task_record.status = "PROCESSING"
                db.session.commit()

            supplier_invoice_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'supplier invoices')
            logger.info(f"供应商账单目录: {supplier_invoice_folder}")
            
            count = generate_supplier_invoices(int(year), int(month), db, {
                'Waybill': Waybill,
                'Supplier': Supplier,
                'SupplierInvoice': SupplierInvoice,
                'SupplierQuote': SupplierQuote,
                'Product': Product
            }, supplier_invoice_folder)
            
            if task_record:
                task_record.status = "SUCCESS"
                task_record.result_msg = f"成功生成了 {count} 份供应商账单"
                db.session.commit()
            
            logger.info("========== 供应商账单任务执行成功 ==========")
            return {"success": True, "count": count}
        except Exception as e:
            logger.error(f"========== 供应商账单任务失败: {str(e)} ==========")
            logger.exception("详细错误:")
            if task_record:
                task_record.status = "FAILURE"
                task_record.result_msg = str(e)
                db.session.commit()
            raise e


def generate_push_events(tracking, interface, raw_response_json):
    """
    生成推送报文（头程轨迹）
    
    Args:
        tracking: TrackingInfo 对象
        interface: TrackingInterface 对象
        raw_response_json: 原始报文 JSON 对象
    
    Returns:
        list: 推送事件列表
    """
    import json
    from datetime import datetime
    
    push_events = []
    
    try:
        # 解析关键信息代码参数
        response_key_params = json.loads(interface.response_key_params) if interface.response_key_params else {}
        time_key = response_key_params.get('time_key', 'changeDate')  # 默认为changeDate
        status_key = response_key_params.get('status_key', 'status')
        description_key = response_key_params.get('description_key', 'record')
        # city_key 和 country_key 不使用默认值，空字符串就是空
        city_key = response_key_params.get('city_key', '')
        country_key = response_key_params.get('country_key', '')
        
        # 解析头程状态映射表
        status_mapping = json.loads(interface.status_mapping) if interface.status_mapping else []
        
        # 遍历轨迹节点 - 修复：进入 trackInfo 数组
        tracks = raw_response_json.get('tracks', [])
        for track in tracks:
            track_info_list = track.get('trackInfo', [])
            
            for track_info in track_info_list:
                # 提取时间戳并转换为 ISO 格式（北京时间）
                time_value = track_info.get(time_key)
                tracking_time = ''
                if time_value:
                    try:
                        # 如果是数值类型，则转换为时间格式（假设为毫秒时间戳）
                        if isinstance(time_value, (int, float)):
                            from datetime import timedelta
                            dt = datetime.utcfromtimestamp(time_value / 1000)
                            # 转换为北京时间 (UTC+8)
                            dt_beijing = dt + timedelta(hours=8)
                            tracking_time = dt_beijing.isoformat()
                        # 如果已经是字符串格式
                        elif isinstance(time_value, str):
                            # 如果包含 Z 后缀，说明是 UTC 时间，需要转换
                            if time_value.endswith('Z'):
                                from datetime import timedelta
                                dt = datetime.fromisoformat(time_value.replace('Z', '+00:00'))
                                # 转换为北京时间 (UTC+8)
                                dt_beijing = dt + timedelta(hours=8)
                                tracking_time = dt_beijing.replace(tzinfo=None).isoformat()
                            else:
                                # 已经是北京时间或无时区信息，直接使用
                                tracking_time = time_value
                    except:
                        tracking_time = ''
                
                supplier_status = str(track_info.get(status_key, ''))
                supplier_description = track_info.get(description_key, '')
                
                # 匹配系统状态代码：优先匹配轨迹描述，其次匹配状态代码
                system_status_code = ''
                
                # 第1优先级：用 record 匹配映射表中的 supplier_description（包含匹配）
                if supplier_description:
                    for mapping in status_mapping:
                        mapping_desc = mapping.get('supplier_description', '').strip()
                        if mapping_desc and mapping_desc in supplier_description.strip():
                            system_status_code = mapping.get('system_status_code', '')
                            break
                
                # 第2优先级：如果描述没匹配到，用 status 匹配 supplier_status
                if not system_status_code:
                    for mapping in status_mapping:
                        if str(mapping.get('supplier_status', '')) == supplier_status:
                            system_status_code = mapping.get('system_status_code', '')
                            break
                
                # 根据匹配到的 system_status_code 获取该状态码的默认城市和国家
                default_city = ''
                default_country = ''
                if system_status_code:
                    node = TrackingNode.query.filter_by(status_code=system_status_code).first()
                    if node:
                        default_city = node.default_city or ''
                        default_country = node.default_country_code or ''
                
                # 从关键参数指定的字段提取城市和国家
                city = ''
                country = ''
                if city_key:
                    city = track_info.get(city_key, '')
                if country_key:
                    country = track_info.get(country_key, '')
                
                # 如果提取到的城市/国家为空，使用该状态码的默认值
                if not city:
                    city = default_city
                if not country:
                    country = default_country
                
                # 只有匹配到状态代码时才生成推送事件
                if system_status_code:
                    event = {
                        'order_no': tracking.order_no,
                        'tracking_time': tracking_time,
                        'status_code': system_status_code,
                        'description': supplier_description,
                        'city': city,
                        'country': country,
                        'source': 'headhaul'  # 标记来源为头程
                    }
                    push_events.append(event)
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"生成头程推送报文失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
    
    return push_events


def merge_lastmile_push_events(tracking, existing_events):
    """
    合并尾程轨迹到推送报文
    
    Args:
        tracking: TrackingInfo 对象
        existing_events: 现有的推送事件列表
    
    Returns:
        list: 合并后的推送事件列表
    """
    import json
    from datetime import datetime
    
    if not tracking.lastmile_tracking_response:
        return existing_events
    
    try:
        lastmile_data = json.loads(tracking.lastmile_tracking_response)
        lastmile_tracks = lastmile_data.get('data', {}).get('accepted', [])
        
        # 解析尾程轨迹（只处理当前订单的尾程单号）
        for track_item in lastmile_tracks:
            # 检查是否为当前订单的尾程单号
            track_number = track_item.get('number', '')
            if track_number != tracking.lastmile_no:
                # 跳过不属于当前订单的尾程单号
                continue
            
            track_info = track_item.get('track_info', {})
            tracking_detail = track_info.get('tracking', {}).get('providers', [{}])[0] if track_info.get('tracking') else {}
            events = tracking_detail.get('events', [])
            
            for event in events:
                time_iso = event.get('time_iso', '')
                
                # 将 ISO 8601 格式转换为 yyyy-MM-dd HH:mm:ss
                tracking_time = ''
                if time_iso:
                    try:
                        from datetime import datetime
                        # 解析 ISO 8601 格式（带时区）
                        dt = datetime.fromisoformat(time_iso.replace('Z', '+00:00'))
                        # 转换为 yyyy-MM-dd HH:mm:ss 格式
                        tracking_time = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        tracking_time = time_iso
                
                sub_status = event.get('sub_status', '')
                description = event.get('description', '')
                
                # 从 address 对象中提取 city 和 country
                address = event.get('address', {})
                city = ''
                country = ''
                
                if isinstance(address, dict):
                    city = address.get('city', '')
                    country = address.get('country', '')
                
                # 如果 country 为空，默认使用 US
                if not country:
                    country = 'US'
                
                # 如果 address 中没有 city，尝试从 location 解析（降级方案）
                if not city:
                    location = event.get('location', {})
                    if isinstance(location, dict):
                        city = location.get('city', '')
                    elif isinstance(location, str):
                        # location 是字符串，可能为 "ALTOONA, PA, US" 类型格式
                        city = location
                
                # 匹配系统状态代码（使用尾程映射表）
                system_status_code = ''
                matched_node = None
                
                # 优先匹配 description
                if description:
                    mapping = LastmileStatusMapping.query.filter_by(
                        description=description,
                        sub_status=sub_status
                    ).first()
                    if mapping:
                        system_status_code = mapping.system_status_code
                        # 获取对应的轨迹节点
                        matched_node = TrackingNode.query.filter_by(status_code=system_status_code).first()
                
                # 如果没匹配到，直接匹配 sub_status
                if not system_status_code:
                    mapping = LastmileStatusMapping.query.filter(
                        LastmileStatusMapping.sub_status == sub_status,
                        db.or_(
                            LastmileStatusMapping.description == '',
                            LastmileStatusMapping.description.is_(None)
                        )
                    ).first()
                    if mapping:
                        system_status_code = mapping.system_status_code
                        # 获取对应的轨迹节点
                        matched_node = TrackingNode.query.filter_by(status_code=system_status_code).first()
                
                # 如果 city 或 country 为空，使用轨迹节点的默认值
                if matched_node:
                    if not city:
                        city = matched_node.default_city or ''
                    if not country or country == 'US':  # 如果是默认的 US，尝试使用节点的默认国家
                        if matched_node.default_country_code:
                            country = matched_node.default_country_code
                
                # 检查是否与头程轨迹时间相同或状态码相同
                # 如果时间相同或状态码相同，删除头程轨迹，只保留尾程
                for i in range(len(existing_events) - 1, -1, -1):  # 从后往前遍历，安全删除
                    existing_event = existing_events[i]
                    # 删除条件：1) 时间相同  2) 状态码相同且都是头程
                    if (existing_event.get('tracking_time') == tracking_time or 
                        (existing_event.get('status_code') == system_status_code and 
                         existing_event.get('source') == 'headhaul')):
                        del existing_events[i]
                
                # 只有匹配到状态代码时才添加尾程事件
                if system_status_code:
                    new_event = {
                        'order_no': tracking.order_no,
                        'tracking_time': tracking_time,
                        'status_code': system_status_code,
                        'description': description,
                        'city': city,
                        'country': country,
                        'source': 'lastmile'
                    }
                    existing_events.append(new_event)
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"合并尾程推送报文失败: {str(e)}")
    
    # 按时间排序
    existing_events.sort(key=lambda x: x.get('tracking_time', ''))
    
    return existing_events


@celery.task(bind=True, name='app.async_fetch_tracking_task')
def async_fetch_tracking_task(self, waybill_ids):
    """
    异步获取头程轨迹任务
    
    Args:
        waybill_ids: 运单ID列表
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[异步轨迹] 任务启动，ID: {self.request.id}, 运单数: {len(waybill_ids)}")
    
    with app.app_context():
        try:
            # 更新任务状态
            task_record = TaskRecord.query.filter_by(task_id=self.request.id).first()
            if task_record:
                task_record.status = 'PROCESSING'
                db.session.commit()
            
            from tracking_handler.tracking_handler_manager import batch_fetch_tracking_by_interface
            import json
            from datetime import datetime
            
            # 按轨迹接口分组
            interface_groups = {}
            
            for waybill_id in waybill_ids:
                tracking = TrackingInfo.query.filter_by(waybill_id=waybill_id).first()
                if not tracking or not tracking.tracking_interface_id:
                    continue
                
                interface_id = tracking.tracking_interface_id
                if interface_id not in interface_groups:
                    interface_groups[interface_id] = []
                
                interface_groups[interface_id].append({
                    'waybill_id': waybill_id,
                    'order_no': tracking.order_no,
                    'transfer_no': tracking.transfer_no
                })
            
            total_success = 0
            total_failed = 0
            error_details = []
            now = datetime.utcnow()
            
            # 按接口批量处理
            for interface_id, waybill_list in interface_groups.items():
                interface = TrackingInterface.query.get(interface_id)
                if not interface:
                    continue
                
                logger.info(f"[异步轨迹] 处理接口: {interface.interface_name}, 运单数: {len(waybill_list)}")
                
                interface_config = {
                    'interface_name': interface.interface_name,
                    'request_url': interface.request_url,
                    'auth_params': interface.auth_params
                }
                
                status_mapping = json.loads(interface.status_mapping) if interface.status_mapping else []
                response_key_params = json.loads(interface.response_key_params) if interface.response_key_params else None
                
                # 批量获取轨迹
                results = batch_fetch_tracking_by_interface(
                    waybill_list,
                    interface_config,
                    status_mapping,
                    response_key_params
                )
                
                # 更新数据库
                for result in results:
                    waybill_id = result.get('waybill_id')
                    if result.get('success'):
                        tracking = TrackingInfo.query.filter_by(waybill_id=waybill_id).first()
                        if tracking:
                            tracking.tracking_description = result.get('tracking_description', '')
                            tracking.status_code = result.get('status_code', '')
                            tracking.tracking_time = result.get('tracking_time')
                            tracking.raw_response = result.get('raw_response', '')
                            tracking.last_fetch_time = now
                            
                            # 从 json中提取尾程单号并保存
                            try:
                                raw_data = json.loads(result.get('raw_response', '{}'))
                                if "tracks" in raw_data and raw_data["tracks"]:
                                    lastmile_no = raw_data["tracks"][0].get("transferNo", "")
                                    if lastmile_no:
                                        tracking.lastmile_no = lastmile_no
                                
                                # 生成推送报文
                                push_events = generate_push_events(tracking, interface, raw_data)
                                
                                # 如果已有尾程轨迹，合并
                                if tracking.lastmile_tracking_response:
                                    push_events = merge_lastmile_push_events(tracking, push_events)
                                
                                # 保存推送报文
                                tracking.push_events = json.dumps(push_events, ensure_ascii=False)
                                
                            except Exception as e:
                                logger.error(f"生成推送报文失败: {str(e)}")
                            
                            total_success += 1
                            logger.info(f"[异步轨迹] 运单 {result.get('order_no')} 获取成功")
                    else:
                        total_failed += 1
                        error_msg = f"运单ID {waybill_id}: {result.get('message')}"
                        error_details.append(error_msg)
                        logger.warning(f"[异步轨迹] {error_msg}")
            
            db.session.commit()
            
            # 更新任务状态
            if task_record:
                task_record.status = 'SUCCESS'
                task_record.result_msg = f"成功: {total_success}单, 失败: {total_failed}单"
                db.session.commit()
            
            logger.info(f"[异步轨迹] 任务完成，成功: {total_success}, 失败: {total_failed}")
            
            return {
                "success": True,
                "total_success": total_success,
                "total_failed": total_failed,
                "error_details": error_details,
                "message": f"成功获取 {total_success} 条轨迹数据"
            }
            
        except Exception as e:
            logger.error(f"[异步轨迹] 任务异常: {str(e)}")
            import traceback
            traceback.print_exc()
            
            if task_record:
                task_record.status = 'FAILURE'
                task_record.result_msg = str(e)
                db.session.commit()
            
            raise


@celery.task(bind=True, name='app.async_fetch_lastmile_tracking_task')
def async_fetch_lastmile_tracking_task(self, waybill_ids):
    """
    异步获取尾程轨迹任务
    
    Args:
        waybill_ids: 运单ID列表
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[异步尾程轨迹] 任务启动，ID: {self.request.id}, 运单数: {len(waybill_ids)}")
    
    with app.app_context():
        try:
            # 更新任务状态
            task_record = TaskRecord.query.filter_by(task_id=self.request.id).first()
            if task_record:
                task_record.status = 'PROCESSING'
                db.session.commit()
            
            from tracking_handler.tracking_lastmile_handler import batch_fetch_lastmile_tracking, fetch_lastmile_tracking
            from datetime import datetime
            import json
            
            # 查询运单和尾程单号
            waybills = db.session.query(
                Waybill.id,
                Waybill.order_no,
                TrackingInfo.id.label('tracking_id'),
                TrackingInfo.lastmile_no,
                TrackingInfo.lastmile_register_response
            ).outerjoin(
                TrackingInfo, Waybill.id == TrackingInfo.waybill_id
            ).filter(
                Waybill.id.in_(waybill_ids)
            ).all()
            
            already_registered_list = []
            to_register_list = []
            
            for wb in waybills:
                if wb.lastmile_no:
                    tracking = TrackingInfo.query.get(wb.tracking_id)
                    if tracking and tracking.lastmile_register_response:
                        already_registered_list.append({
                            'waybill_id': wb.id,
                            'lastmile_no': wb.lastmile_no,
                            'tracking': tracking
                        })
                    else:
                        to_register_list.append({
                            'waybill_id': wb.id,
                            'lastmile_no': wb.lastmile_no
                        })
            
            success_count = 0
            error_details = []
            now = datetime.utcnow()
            
            # 处理已注册的单号
            if already_registered_list:
                logger.info(f"[异步尾程轨迹] 处理已注册单号 {len(already_registered_list)}单")
                
                registered_numbers = [item['lastmile_no'] for item in already_registered_list]
                
                try:
                    fetch_result = fetch_lastmile_tracking(registered_numbers)
                    
                    if fetch_result.get('success'):
                        for item in already_registered_list:
                            tracking = item['tracking']
                            tracking.lastmile_tracking_response = fetch_result.get('raw_response', '')
                            tracking.lastmile_last_fetch_time = now
                            
                            # 合并到推送报文
                            try:
                                existing_push_events = json.loads(tracking.push_events) if tracking.push_events else []
                                merged_events = merge_lastmile_push_events(tracking, existing_push_events)
                                tracking.push_events = json.dumps(merged_events, ensure_ascii=False)
                            except Exception as e:
                                logger.error(f"合并尾程推送报文失败: {str(e)}")
                            
                            success_count += 1
                    else:
                        error_details.append(f"已注册单号查询失败: {fetch_result.get('message')}")
                except Exception as e:
                    error_details.append(f"已注册单号查询异常: {str(e)}")
            
            # 处理未注册的单号
            if to_register_list:
                logger.info(f"[异步尾程轨迹] 处理未注册单号 {len(to_register_list)}单")
                
                try:
                    results = batch_fetch_lastmile_tracking(to_register_list)
                    
                    for result in results:
                        if result.get('success'):
                            waybill_id = result.get('waybill_id')
                            tracking = TrackingInfo.query.filter_by(waybill_id=waybill_id).first()
                            if tracking:
                                tracking.lastmile_register_response = result.get('register_response', '')
                                tracking.lastmile_tracking_response = result.get('tracking_response', '')
                                tracking.lastmile_last_fetch_time = now
                                
                                # 合并到推送报文
                                try:
                                    existing_push_events = json.loads(tracking.push_events) if tracking.push_events else []
                                    merged_events = merge_lastmile_push_events(tracking, existing_push_events)
                                    tracking.push_events = json.dumps(merged_events, ensure_ascii=False)
                                except Exception as e:
                                    logger.error(f"合并尾程推送报文失败: {str(e)}")
                                
                                success_count += 1
                        else:
                            error_details.append(f"运单ID {result.get('waybill_id')}: {result.get('message')}")
                except Exception as e:
                    error_details.append(f"未注册单号处理异常: {str(e)}")
            
            db.session.commit()
            
            # 更新任务状态
            if task_record:
                task_record.status = 'SUCCESS'
                task_record.result_msg = f"成功: {success_count}单, 失败: {len(error_details)}单"
                db.session.commit()
            
            logger.info(f"[异步尾程轨迹] 任务完成，成功: {success_count}, 失败: {len(error_details)}")
            
            return {
                "success": True,
                "total_success": success_count,
                "total_failed": len(error_details),
                "error_details": error_details,
                "message": f"成功获取 {success_count} 条尾程轨迹数据"
            }
            
        except Exception as e:
            logger.error(f"[异步尾程轨迹] 任务异常: {str(e)}")
            import traceback
            traceback.print_exc()
            
            if task_record:
                task_record.status = 'FAILURE'
                task_record.result_msg = str(e)
                db.session.commit()
            
            raise


@celery.task(name='app.auto_fetch_tracking_task')
def auto_fetch_tracking_task():
    """
    自动获取头程轨迹定时任务（同时获取尾程轨迹）
    每小时执行一次，检查需要获取轨迹的运单
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info("========== 开始执行自动获取轨迹任务 ==========")
    
    with app.app_context():
        try:
            from tracking_handler.tracking_handler_manager import batch_fetch_tracking_by_interface
            from tracking_handler.tracking_lastmile_handler import batch_fetch_lastmile_tracking, fetch_lastmile_tracking
            import json
            from datetime import datetime, timedelta
            from decimal import Decimal
            
            now = datetime.utcnow()
            
            # 查询所有轨迹接口
            interfaces = TrackingInterface.query.all()
            
            total_fetched = 0
            total_skipped = 0
            total_failed = 0
            
            # 尾程轨迹统计
            total_lastmile_fetched = 0
            total_lastmile_failed = 0
            
            for interface in interfaces:
                logger.info(f"处理接口: {interface.interface_name}")
                
                # 计算时间阈值：现在 - fetch_interval
                fetch_interval_hours = float(interface.fetch_interval)
                threshold_time = now - timedelta(hours=fetch_interval_hours)
                
                # 查询需要获取轨迹的运单
                # 条件：
                # 1. 产品绑定了该接口
                # 2. 未停止跟踪（stop_tracking=False）
                # 3. last_fetch_time < threshold_time 或者 last_fetch_time 为 NULL
                query = db.session.query(
                    TrackingInfo.id,
                    TrackingInfo.waybill_id,
                    TrackingInfo.order_no,
                    TrackingInfo.transfer_no,
                    TrackingInfo.last_fetch_time,
                    TrackingInfo.lastmile_no,
                    TrackingInfo.lastmile_register_response
                ).filter(
                    TrackingInfo.tracking_interface_id == interface.id,
                    db.or_(
                        TrackingInfo.stop_tracking == False,
                        TrackingInfo.stop_tracking.is_(None)
                    ),
                    db.or_(
                        TrackingInfo.last_fetch_time < threshold_time,
                        TrackingInfo.last_fetch_time.is_(None)
                    )
                )
                
                trackings = query.all()
                
                if not trackings:
                    logger.info(f"  接口 {interface.interface_name} 暂无需要获取的运单")
                    continue
                
                logger.info(f"  找到 {len(trackings)} 个需要获取轨迹的运单")
                
                # 构造运单列表
                waybill_list = []
                for tracking in trackings:
                    waybill_list.append({
                        'waybill_id': tracking.waybill_id,
                        'order_no': tracking.order_no,
                        'transfer_no': tracking.transfer_no
                    })
                
                # 准备接口配置
                interface_config = {
                    'interface_name': interface.interface_name,
                    'request_url': interface.request_url,
                    'auth_params': interface.auth_params
                }
                
                status_mapping = json.loads(interface.status_mapping) if interface.status_mapping else []
                response_key_params = json.loads(interface.response_key_params) if interface.response_key_params else None
                
                # 批量获取头程轨迹
                results = batch_fetch_tracking_by_interface(
                    waybill_list,
                    interface_config,
                    status_mapping,
                    response_key_params
                )
                
                # 收集有尾程单号的运单
                lastmile_registered_list = []  # 已注册的尾程单号
                lastmile_to_register_list = []  # 未注册的尾程单号
                
                # 处理头程轨迹结果
                for result in results:
                    if not result.get('success'):
                        total_failed += 1
                        logger.warning(f"    运单 {result.get('order_no')} 获取失败: {result.get('message')}")
                        continue
                    
                    waybill_id = result['waybill_id']
                    
                    # 查询运单和轨迹信息
                    waybill = Waybill.query.get(waybill_id)
                    tracking = TrackingInfo.query.filter_by(waybill_id=waybill_id).first()
                    
                    if not waybill or not tracking:
                        total_failed += 1
                        continue
                    
                    # 更新头程轨迹信息
                    tracking.tracking_description = result.get('tracking_description', '')
                    tracking.status_code = result.get('status_code', '')
                    tracking.tracking_time = result.get('tracking_time')
                    tracking.raw_response = result.get('raw_response', '')
                    tracking.last_fetch_time = now
                    
                    # 从头程报文中提取尾程单号
                    try:
                        raw_data = json.loads(result.get('raw_response', '{}'))
                        if "tracks" in raw_data and raw_data["tracks"]:
                            lastmile_no = raw_data["tracks"][0].get("transferNo", "")
                            if lastmile_no:
                                tracking.lastmile_no = lastmile_no
                                
                                # 判断是否需要获取尾程轨迹
                                if tracking.lastmile_register_response:
                                    # 已注册，直接查询
                                    lastmile_registered_list.append({
                                        'lastmile_no': lastmile_no,
                                        'tracking': tracking
                                    })
                                else:
                                    # 未注册，需要注册
                                    lastmile_to_register_list.append({
                                        'waybill_id': waybill_id,
                                        'lastmile_no': lastmile_no
                                    })
                    except:
                        pass
                    
                    # 检查是否应该停止跟踪
                    should_stop, reason = should_stop_tracking(waybill, tracking)
                    if should_stop:
                        tracking.stop_tracking = True
                        tracking.stop_tracking_reason = reason
                        tracking.stop_tracking_time = now
                        logger.info(f"    运单 {tracking.order_no} 满足停止条件: {reason}")
                    
                    total_fetched += 1
            
                # 批量获取尾程轨迹（已注册的单号）
                if lastmile_registered_list:
                    logger.info(f"  开始获取已注册尾程单号 {len(lastmile_registered_list)} 个")
                    registered_numbers = [item['lastmile_no'] for item in lastmile_registered_list]
                    
                    try:
                        fetch_result = fetch_lastmile_tracking(registered_numbers)
                        if fetch_result.get('success'):
                            for item in lastmile_registered_list:
                                tracking = item['tracking']
                                tracking.lastmile_tracking_response = fetch_result.get('raw_response', '')
                                tracking.lastmile_last_fetch_time = now
                                total_lastmile_fetched += 1
                        else:
                            total_lastmile_failed += len(lastmile_registered_list)
                            logger.warning(f"  已注册尾程单号查询失败: {fetch_result.get('message')}")
                    except Exception as e:
                        total_lastmile_failed += len(lastmile_registered_list)
                        logger.error(f"  已注册尾程单号查询异常: {str(e)}")
                
                # 批量获取尾程轨迹（未注册的单号）
                if lastmile_to_register_list:
                    logger.info(f"  开始注册并获取尾程单号 {len(lastmile_to_register_list)} 个")
                    
                    try:
                        lastmile_results = batch_fetch_lastmile_tracking(lastmile_to_register_list)
                        for result in lastmile_results:
                            if result.get('success'):
                                waybill_id = result.get('waybill_id')
                                tracking = TrackingInfo.query.filter_by(waybill_id=waybill_id).first()
                                if tracking:
                                    tracking.lastmile_register_response = result.get('register_response', '')
                                    tracking.lastmile_tracking_response = result.get('tracking_response', '')
                                    tracking.lastmile_last_fetch_time = now
                                    total_lastmile_fetched += 1
                            else:
                                total_lastmile_failed += 1
                                logger.warning(f"  运单 {result.get('waybill_id')} 尾程轨迹获取失败: {result.get('message')}")
                    except Exception as e:
                        total_lastmile_failed += len(lastmile_to_register_list)
                        logger.error(f"  未注册尾程单号处理异常: {str(e)}")
            
            # 提交所有更新
            db.session.commit()
            
            logger.info(f"========== 任务执行完成 ==========")
            logger.info(f"头程轨迹 - 成功: {total_fetched} 条, 失败: {total_failed} 条")
            logger.info(f"尾程轨迹 - 成功: {total_lastmile_fetched} 条, 失败: {total_lastmile_failed} 条")
            
            return {
                "success": True,
                "fetched": total_fetched,
                "failed": total_failed,
                "lastmile_fetched": total_lastmile_fetched,
                "lastmile_failed": total_lastmile_failed,
                "message": f"头程成功 {total_fetched} 条，尾程成功 {total_lastmile_fetched} 条"
            }
            
        except Exception as e:
            logger.error(f"========== 任务执行失败: {str(e)} ==========")
            logger.exception("详细错误:")
            db.session.rollback()
            return {
                "success": False,
                "message": f"执行失败: {str(e)}"
            }


@celery.task(name='app.auto_fetch_lastmile_tracking_task')
def auto_fetch_lastmile_tracking_task():
    """
    尾程轨迹自动获取任务（已在 auto_fetch_tracking_task 中实现）
    此函数保留以兼容历史配置，实际不会被调用
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info("尾程轨迹自动获取已合并到头程轨迹任务中，无需单独执行")
    return {
        "success": True,
        "message": "尾程轨迹获取已集成在头程轨迹任务中"
    }


@app.route('/pictures/<path:filename>')
def serve_pictures(filename):
    """根据路径获取图片文件"""
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pictures'), filename)


@app.route('/uploads/<path:filename>')
def serve_uploads(filename):
    """根据路径获取上传的文件（如水单图片）"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route("/")
def index():
    """根路径：如果已登录，进入主界面；否则跳转到登录页"""
    if "user" in session:
        return redirect(url_for("app_main"))
    return redirect(url_for("login"))


@app.route("/login")
def login():
    """登录页（只返回前端登录界面，逻辑在 /api/login 中处理）"""
    if "user" in session:
        return redirect(url_for("app_main"))
    return render_template("login.html")


@app.post("/api/login")
def api_login():
    """处理登录请求：支持默认账号 admin 和数据库用户"""
    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()

    # 先检查是否是默认管理员账号
    if username in USERS:
        user = USERS[username]
        if user["password"] == password:
            session["user"] = {
                "username": username,
                "role": user["role"],
                "permissions": user["permissions"],
            }
            return jsonify({"success": True, "redirect": url_for("app_main")})
    
    # 再检查数据库中的用户
    db_user = User.query.filter_by(username=username).first()
    if db_user and db_user.password == password:
        session["user"] = {
            "username": username,
            "role": db_user.role.name if db_user.role else "普通用户",
            "permissions": "ROLE_BASED",  # 基于角色的权限
        }
        return jsonify({"success": True, "redirect": url_for("app_main")})
    
    # 登录失败
    return jsonify({"success": False, "message": "用户名或密码错误"}), 400


@app.post("/api/logout")
def api_logout():
    """退出登录"""
    session.clear()
    return jsonify({"success": True})


@app.get("/api/user-permissions")
def api_get_user_permissions():
    """获取当前用户的页面权限"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    # 如果是 admin 管理员，返回所有权限
    if current_user.get("permissions") == "ALL":
        return jsonify({
            "success": True,
            "permissions": "ALL",
            "pages": []
        })
    
    # 获取数据库用户的角色权限
    db_user = User.query.filter_by(username=current_user["username"]).first()
    if not db_user or not db_user.role:
        return jsonify({
            "success": True,
            "permissions": "NONE",
            "pages": []
        })
    
    # 返回角色的权限列表
    permissions_data = []
    for perm in db_user.role.permissions:
        permission_item = {
            "page_key": perm.page_key,
            "can_view": perm.can_view,
            "can_create": perm.can_create,
            "can_update": perm.can_update,
            "can_delete": perm.can_delete,
        }
        
        # 添加字段权限（如果存在）
        if perm.field_permissions:
            try:
                permission_item["field_permissions"] = json.loads(perm.field_permissions)
            except:
                permission_item["field_permissions"] = {}
        else:
            permission_item["field_permissions"] = {}
        
        permissions_data.append(permission_item)
    
    return jsonify({
        "success": True,
        "permissions": "ROLE_BASED",
        "pages": permissions_data
    })


@app.post("/api/change-password")
def api_change_password():
    """修改密码"""
    # 检查登录状态
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    data = request.get_json() or {}
    old_password = (data.get("oldPassword") or "").strip()
    new_password = (data.get("newPassword") or "").strip()
    
    # 验证新密码格式
    if len(new_password) < 6:
        return jsonify({
            "success": False,
            "message": "密码长度至少6位",
            "field": "newPassword"
        }), 400
    
    username = user["username"]
    
    # 检查是否是内存中的默认管理员
    if username in USERS:
        # 验证旧密码
        if USERS[username]["password"] != old_password:
            return jsonify({
                "success": False,
                "message": "当前密码错误",
                "field": "oldPassword"
            }), 400
        
        # 更新内存中的密码（注意：重启后会恢复）
        USERS[username]["password"] = new_password
        return jsonify({"success": True, "message": "密码修改成功"})
    
    # 检查数据库用户
    db_user = User.query.filter_by(username=username).first()
    if not db_user:
        return jsonify({
            "success": False,
            "message": "用户不存在"
        }), 404
    
    # 验证旧密码
    if db_user.password != old_password:
        return jsonify({
            "success": False,
            "message": "当前密码错误",
            "field": "oldPassword"
        }), 400
    
    # 更新数据库中的密码
    db_user.password = new_password
    db.session.commit()
    
    return jsonify({"success": True, "message": "密码修改成功"})


@app.post("/api/roles")
def api_create_role():
    """创建角色，并保存页面级增删改查权限与创建时间"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    permissions_data = data.get("permissions") or []

    if not name:
        return jsonify({"success": False, "message": "角色名称不能为空"}), 400

    # 简单：只有系统管理员可以创建角色（后续可按权限表做更精细控制）
    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建角色"}), 403

    if Role.query.filter_by(name=name).first():
        return jsonify({"success": False, "message": "角色名称已存在"}), 400

    role = Role(name=name, description=description)

    for item in permissions_data:
        page_key = (item.get("pageKey") or "").strip()
        if not page_key:
            continue
        can_view = bool(item.get("canView"))
        can_create = bool(item.get("canCreate"))
        can_update = bool(item.get("canUpdate"))
        can_delete = bool(item.get("canDelete"))
        
        # 获取字段权限
        field_permissions = item.get("field_permissions")
        field_permissions_json = None
        if field_permissions:
            field_permissions_json = json.dumps(field_permissions)
        
        # 如果四个权限全为 False，且没有字段权限，就不保存该页面的记录
        if not (can_view or can_create or can_update or can_delete) and not field_permissions_json:
            continue

        role.permissions.append(
            RolePagePermission(
                page_key=page_key,
                can_view=can_view,
                can_create=can_create,
                can_update=can_update,
                can_delete=can_delete,
                field_permissions=field_permissions_json
            )
        )

    db.session.add(role)
    db.session.commit()

    return jsonify({"success": True, "id": role.id})


@app.get("/api/roles")
def api_get_roles():
    """获取角色列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = Role.query.order_by(Role.id.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        roles = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        roles = query.all()
        pagination_data = None

    roles_data = []
    for role in roles:
        roles_data.append({
            "id": role.id,
            "name": role.name,
            "description": role.description,
            "created_at": role.created_at.isoformat() if role.created_at else None,
            "permissions": [
                {
                    "page_key": p.page_key,
                    "can_view": p.can_view,
                    "can_create": p.can_create,
                    "can_update": p.can_update,
                    "can_delete": p.can_delete,
                    # 添加字段权限（如果存在）
                    "field_permissions": json.loads(p.field_permissions) if p.field_permissions else {}
                }
                for p in role.permissions
            ],
        })

    return jsonify({
        "success": True, 
        "roles": roles_data,
        "pagination": pagination_data
    })


@app.put("/api/roles/<int:role_id>")
def api_update_role(role_id):
    """更新角色"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑角色"}), 403

    role = Role.query.get(role_id)
    if not role:
        return jsonify({"success": False, "message": "角色不存在"}), 404

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    permissions_data = data.get("permissions") or []

    if not name:
        return jsonify({"success": False, "message": "角色名称不能为空"}), 400

    # 检查名称是否与其他角色重复
    existing = Role.query.filter_by(name=name).first()
    if existing and existing.id != role_id:
        return jsonify({"success": False, "message": "角色名称已存在"}), 400

    role.name = name
    role.description = description

    # 删除旧权限
    RolePagePermission.query.filter_by(role_id=role_id).delete()

    # 添加新权限
    for item in permissions_data:
        page_key = (item.get("pageKey") or "").strip()
        if not page_key:
            continue
        can_view = bool(item.get("canView"))
        can_create = bool(item.get("canCreate"))
        can_update = bool(item.get("canUpdate"))
        can_delete = bool(item.get("canDelete"))
        
        # 获取字段权限
        field_permissions = item.get("field_permissions")
        field_permissions_json = None
        if field_permissions:
            field_permissions_json = json.dumps(field_permissions)
        
        if not (can_view or can_create or can_update or can_delete) and not field_permissions_json:
            continue

        role.permissions.append(
            RolePagePermission(
                page_key=page_key,
                can_view=can_view,
                can_create=can_create,
                can_update=can_update,
                can_delete=can_delete,
                field_permissions=field_permissions_json
            )
        )

    db.session.commit()
    return jsonify({"success": True})


@app.delete("/api/roles/<int:role_id>")
def api_delete_role(role_id):
    """删除角色"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除角色"}), 403

    role = Role.query.get(role_id)
    if not role:
        return jsonify({"success": False, "message": "角色不存在"}), 404

    db.session.delete(role)
    db.session.commit()
    return jsonify({"success": True})


# ==================== 用户管理 API ====================

@app.get("/api/users")
def api_get_users():
    """获取用户列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = User.query.order_by(User.id.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        users = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        users = query.all()
        pagination_data = None

    users_data = []
    for user in users:
        users_data.append({
            "id": user.id,
            "username": user.username,
            "role_id": user.role_id,
            "role_name": user.role.name if user.role else "",
            "created_at": user.created_at.isoformat() if user.created_at else None,
        })

    return jsonify({
        "success": True, 
        "users": users_data,
        "pagination": pagination_data
    })


@app.post("/api/users")
def api_create_user():
    """创建用户，默认密码 654321"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建用户"}), 403

    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    role_id = data.get("role_id")

    if not username:
        return jsonify({"success": False, "message": "用户名不能为空", "field": "username"}), 400

    if not role_id:
        return jsonify({"success": False, "message": "请选择角色", "field": "role_id"}), 400

    # 检查用户名是否已存在
    if User.query.filter_by(username=username).first():
        return jsonify({"success": False, "message": "用户名已存在", "field": "username"}), 400

    # 检查角色是否存在
    if not Role.query.get(role_id):
        return jsonify({"success": False, "message": "角色不存在", "field": "role_id"}), 400

    # 创建用户，默认密码 654321
    user = User(username=username, password="654321", role_id=role_id)
    db.session.add(user)
    db.session.commit()

    return jsonify({"success": True, "id": user.id})


@app.put("/api/users/<int:user_id>")
def api_update_user(user_id):
    """编辑用户（仅修改角色）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑用户"}), 403

    user = User.query.get(user_id)
    if not user:
        return jsonify({"success": False, "message": "用户不存在"}), 404

    data = request.get_json() or {}
    role_id = data.get("role_id")

    if not role_id:
        return jsonify({"success": False, "message": "请选择角色", "field": "role_id"}), 400

    # 检查角色是否存在
    if not Role.query.get(role_id):
        return jsonify({"success": False, "message": "角色不存在", "field": "role_id"}), 400

    user.role_id = role_id
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/users/<int:user_id>")
def api_delete_user(user_id):
    """删除用户"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除用户"}), 403

    user = User.query.get(user_id)
    if not user:
        return jsonify({"success": False, "message": "用户不存在"}), 404

    db.session.delete(user)
    db.session.commit()

    return jsonify({"success": True})


@app.post("/api/users/<int:user_id>/reset-password")
def api_reset_user_password(user_id):
    """重置用户密码为 654321"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限重置密码"}), 403

    user = User.query.get(user_id)
    if not user:
        return jsonify({"success": False, "message": "用户不存在"}), 404

    user.password = "654321"
    db.session.commit()

    return jsonify({"success": True})


# ==================== 目的国管理 API ====================

@app.get("/api/countries")
def api_get_countries():
    """获取目的国列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = Country.query.order_by(Country.id.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        countries = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        countries = query.all()
        pagination_data = None

    countries_data = []
    for country in countries:
        countries_data.append({
            "id": country.id,
            "name": country.name,
            "code": country.code,
            "created_at": country.created_at.isoformat() if country.created_at else None,
        })

    return jsonify({
        "success": True, 
        "countries": countries_data,
        "pagination": pagination_data
    })


@app.post("/api/countries")
def api_create_country():
    """创建目的国"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建目的国"}), 403

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip().upper()

    if not name:
        return jsonify({"success": False, "message": "国家名称不能为空", "field": "name"}), 400

    if not code:
        return jsonify({"success": False, "message": "国家二字代码不能为空", "field": "code"}), 400

    if len(code) != 2:
        return jsonify({"success": False, "message": "国家代码必须是2个字符", "field": "code"}), 400

    # 检查代码是否已存在
    if Country.query.filter_by(code=code).first():
        return jsonify({"success": False, "message": "国家代码已存在", "field": "code"}), 400

    country = Country(name=name, code=code)
    db.session.add(country)
    db.session.commit()

    return jsonify({"success": True, "id": country.id})


@app.put("/api/countries/<int:country_id>")
def api_update_country(country_id):
    """更新目的国"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑目的国"}), 403

    country = Country.query.get(country_id)
    if not country:
        return jsonify({"success": False, "message": "目的国不存在"}), 404

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip().upper()

    if not name:
        return jsonify({"success": False, "message": "国家名称不能为空", "field": "name"}), 400

    if not code:
        return jsonify({"success": False, "message": "国家二字代码不能为空", "field": "code"}), 400

    if len(code) != 2:
        return jsonify({"success": False, "message": "国家代码必须是2个字符", "field": "code"}), 400

    # 检查代码是否与其他国家重复
    existing = Country.query.filter_by(code=code).first()
    if existing and existing.id != country_id:
        return jsonify({"success": False, "message": "国家代码已存在", "field": "code"}), 400

    country.name = name
    country.code = code
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/countries/<int:country_id>")
def api_delete_country(country_id):
    """删除目的国"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除目的国"}), 403

    country = Country.query.get(country_id)
    if not country:
        return jsonify({"success": False, "message": "目的国不存在"}), 404

    db.session.delete(country)
    db.session.commit()

    return jsonify({"success": True})


@app.post("/api/countries/import")
def api_import_countries():
    """导入目的国数据（支持 xls, xlsx, csv）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限导入数据"}), 403

    if "file" not in request.files:
        return jsonify({"success": False, "message": "请选择文件"}), 400

    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"success": False, "message": "请选择文件"}), 400

    # 检查文件格式
    filename = file.filename.lower()
    if not (filename.endswith(".xls") or filename.endswith(".xlsx") or filename.endswith(".csv")):
        return jsonify({"success": False, "message": "仅支持 xls, xlsx, csv 格式"}), 400

    try:
        # 读取文件内容
        file_content = file.read()
        
        # 根据文件类型解析
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            df = pd.read_excel(io.BytesIO(file_content))

        # 检查表头
        required_columns = ["国家", "国家二字代码"]
        if not all(col in df.columns for col in required_columns):
            return jsonify({
                "success": False,
                "message": f"文件表头必须包含：{', '.join(required_columns)}"
            }), 400

        # 统计数据
        success_count = 0
        skip_count = 0
        error_rows = []

        # 逐行处理
        for index, row in df.iterrows():
            name = str(row["国家"]).strip() if pd.notna(row["国家"]) else ""
            code = str(row["国家二字代码"]).strip().upper() if pd.notna(row["国家二字代码"]) else ""

            # 验证数据
            if not name or not code:
                error_rows.append(f"第{index + 2}行：国家名称或代码为空")
                continue

            if len(code) != 2:
                error_rows.append(f"第{index + 2}行：代码 '{code}' 不是2个字符")
                continue

            # 检查是否已存在
            existing = Country.query.filter_by(code=code).first()
            if existing:
                skip_count += 1
                continue

            # 创建新记录
            country = Country(name=name, code=code)
            db.session.add(country)
            success_count += 1

        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"导入完成：成功 {success_count} 条，跳过 {skip_count} 条",
            "success_count": success_count,
            "skip_count": skip_count,
            "errors": error_rows[:10]  # 最多返回前10条错误
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"文件解析失败：{str(e)}"
        }), 400


# ==================== 产品管理 API ====================

@app.get("/api/products")
def api_get_products():
    """获取产品列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = Product.query.order_by(Product.id.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        products = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        products = query.all()
        pagination_data = None

    products_data = []
    for product in products:
        products_data.append({
            "id": product.id,
            "name": product.name,
            "description": product.description or "",
            "fee_types": product.fee_types.split(",") if product.fee_types else [],
            "supplier_id": product.supplier_id,
            "supplier_name": product.supplier.short_name if product.supplier else "",
            "created_at": product.created_at.isoformat() if product.created_at else None,
        })

    return jsonify({
        "success": True, 
        "products": products_data,
        "pagination": pagination_data
    })


@app.post("/api/products")
def api_create_product():
    """创建产品"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建产品"}), 403

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    fee_types = data.get("fee_types") or []
    supplier_id = data.get("supplier_id")
    tracking_interface_id = data.get("tracking_interface_id")

    if not name:
        return jsonify({"success": False, "message": "产品名称不能为空", "field": "name"}), 400

    if not fee_types or len(fee_types) == 0:
        return jsonify({"success": False, "message": "请至少选择一种收费类别", "field": "fee_types"}), 400

    # 验证描述长度
    if description and len(description) > 100:
        return jsonify({"success": False, "message": "产品描述最多100字", "field": "description"}), 400
    
    # 如果包含差价收费，必须选择供应商
    if "差价收费" in fee_types:
        if not supplier_id:
            return jsonify({"success": False, "message": "包含差价收费的产品必须选择供应商", "field": "supplier_id"}), 400
        # 验证供应商是否存在
        supplier = Supplier.query.get(supplier_id)
        if not supplier:
            return jsonify({"success": False, "message": "供应商不存在", "field": "supplier_id"}), 400
    else:
        supplier_id = None  # 不包含差价收费时，不绑定供应商
    
    # 如果包含尾程收费，可以选择轨迹接口（非必选）
    if "尾程收费" in fee_types and tracking_interface_id:
        # 验证轨迹接口是否存在
        interface = TrackingInterface.query.get(tracking_interface_id)
        if not interface:
            return jsonify({"success": False, "message": "轨迹接口不存在", "field": "tracking_interface_id"}), 400
    elif "尾程收费" not in fee_types:
        tracking_interface_id = None  # 不包含尾程收费时，不绑定轨迹接口

    # 将数组转为逗号分隔的字符串
    fee_types_str = ",".join(fee_types)

    product = Product(
        name=name, 
        description=description, 
        fee_types=fee_types_str, 
        supplier_id=supplier_id,
        tracking_interface_id=tracking_interface_id
    )
    db.session.add(product)
    db.session.commit()

    return jsonify({"success": True, "id": product.id})


@app.put("/api/products/<int:product_id>")
def api_update_product(product_id):
    """更新产品"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑产品"}), 403

    product = Product.query.get(product_id)
    if not product:
        return jsonify({"success": False, "message": "产品不存在"}), 404

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    fee_types = data.get("fee_types") or []
    supplier_id = data.get("supplier_id")
    tracking_interface_id = data.get("tracking_interface_id")

    if not name:
        return jsonify({"success": False, "message": "产品名称不能为空", "field": "name"}), 400

    if not fee_types or len(fee_types) == 0:
        return jsonify({"success": False, "message": "请至少选择一种收费类别", "field": "fee_types"}), 400

    # 验证描述长度
    if description and len(description) > 100:
        return jsonify({"success": False, "message": "产品描述最多100字", "field": "description"}), 400
    
    # 如果包含差价收费，必须选择供应商
    if "差价收费" in fee_types:
        if not supplier_id:
            return jsonify({"success": False, "message": "包含差价收费的产品必须选择供应商", "field": "supplier_id"}), 400
        # 验证供应商是否存在
        supplier = Supplier.query.get(supplier_id)
        if not supplier:
            return jsonify({"success": False, "message": "供应商不存在", "field": "supplier_id"}), 400
    else:
        supplier_id = None  # 不包含差价收费时，不绑定供应商
    
    # 如果包含尾程收费，可以选择轨迹接口（非必选）
    if "尾程收费" in fee_types and tracking_interface_id:
        # 验证轨迹接口是否存在
        interface = TrackingInterface.query.get(tracking_interface_id)
        if not interface:
            return jsonify({"success": False, "message": "轨迹接口不存在", "field": "tracking_interface_id"}), 400
    elif "尾程收费" not in fee_types:
        tracking_interface_id = None  # 不包含尾程收费时，不绑定轨迹接口

    product.name = name
    product.description = description
    product.fee_types = ",".join(fee_types)
    product.supplier_id = supplier_id
    product.tracking_interface_id = tracking_interface_id
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/products/<int:product_id>")
def api_delete_product(product_id):
    """删除产品"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除产品"}), 403

    product = Product.query.get(product_id)
    if not product:
        return jsonify({"success": False, "message": "产品不存在"}), 404

    db.session.delete(product)
    db.session.commit()

    return jsonify({"success": True})


# ==================== 轨迹节点状态管理 API ====================

@app.get("/api/tracking-nodes")
def api_get_tracking_nodes():
    """获取轨迹节点状态列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int, default=1)
    per_page = request.args.get("per_page", type=int, default=20)

    pagination = TrackingNode.query.order_by(TrackingNode.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    nodes_data = []
    for node in pagination.items:
        nodes_data.append({
            "id": node.id,
            "status_code": node.status_code,
            "status_description": node.status_description,
            "default_city": node.default_city or "",
            "default_country_code": node.default_country_code or "",
            "default_airport_code": node.default_airport_code or "",
            "created_at": node.created_at.isoformat() if node.created_at else None,
        })

    return jsonify({
        "success": True,
        "nodes": nodes_data,
        "pagination": {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    })


@app.post("/api/tracking-nodes")
def api_create_tracking_node():
    """创建轨迹节点状态"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建轨迹节点状态"}), 403

    data = request.get_json() or {}
    status_code = (data.get("status_code") or "").strip()
    status_description = (data.get("status_description") or "").strip()
    default_city = (data.get("default_city") or "").strip()
    default_country_code = (data.get("default_country_code") or "").strip()
    default_airport_code = (data.get("default_airport_code") or "").strip()

    if not status_code:
        return jsonify({"success": False, "message": "状态代码不能为空", "field": "status_code"}), 400

    if not status_description:
        return jsonify({"success": False, "message": "状态说明不能为空", "field": "status_description"}), 400

    # 检查状态代码是否已存在
    existing = TrackingNode.query.filter_by(status_code=status_code).first()
    if existing:
        return jsonify({"success": False, "message": f"状态代码'{status_code}'已存在", "field": "status_code"}), 400

    node = TrackingNode(
        status_code=status_code,
        status_description=status_description,
        default_city=default_city if default_city else None,
        default_country_code=default_country_code if default_country_code else None,
        default_airport_code=default_airport_code if default_airport_code else None
    )
    db.session.add(node)
    db.session.commit()

    return jsonify({"success": True, "id": node.id})


@app.put("/api/tracking-nodes/<int:node_id>")
def api_update_tracking_node(node_id):
    """更新轨迹节点状态"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑轨迹节点状态"}), 403

    node = TrackingNode.query.get(node_id)
    if not node:
        return jsonify({"success": False, "message": "轨迹节点状态不存在"}), 404

    data = request.get_json() or {}
    status_code = (data.get("status_code") or "").strip()
    status_description = (data.get("status_description") or "").strip()
    default_city = (data.get("default_city") or "").strip()
    default_country_code = (data.get("default_country_code") or "").strip()
    default_airport_code = (data.get("default_airport_code") or "").strip()

    if not status_code:
        return jsonify({"success": False, "message": "状态代码不能为空", "field": "status_code"}), 400

    if not status_description:
        return jsonify({"success": False, "message": "状态说明不能为空", "field": "status_description"}), 400

    # 检查状态代码是否与其他记录冲突
    existing = TrackingNode.query.filter(TrackingNode.status_code == status_code, TrackingNode.id != node_id).first()
    if existing:
        return jsonify({"success": False, "message": f"状态代码'{status_code}'已存在", "field": "status_code"}), 400

    node.status_code = status_code
    node.status_description = status_description
    node.default_city = default_city if default_city else None
    node.default_country_code = default_country_code if default_country_code else None
    node.default_airport_code = default_airport_code if default_airport_code else None
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/tracking-nodes/<int:node_id>")
def api_delete_tracking_node(node_id):
    """删除轨迹节点状态"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除轨迹节点状态"}), 403

    node = TrackingNode.query.get(node_id)
    if not node:
        return jsonify({"success": False, "message": "轨迹节点状态不存在"}), 404

    db.session.delete(node)
    db.session.commit()

    return jsonify({"success": True})


# ==================== 轨迹接口管理 API ====================

@app.get("/api/tracking-interfaces")
def api_get_tracking_interfaces():
    """获取轨迹接口列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = TrackingInterface.query.order_by(TrackingInterface.id.desc())
    
    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        interfaces = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        interfaces = query.all()
        pagination_data = None
    
    interfaces_data = []
    for interface in interfaces:
        interfaces_data.append({
            "id": interface.id,
            "interface_name": interface.interface_name,
            "request_url": interface.request_url,
            "auth_params": interface.auth_params or "",
            "status_mapping": interface.status_mapping or "",
            "response_key_params": interface.response_key_params or "",
            "fetch_interval": float(interface.fetch_interval) if interface.fetch_interval else 0,
            "created_at": interface.created_at.isoformat() if interface.created_at else None,
        })

    return jsonify({
        "success": True,
        "interfaces": interfaces_data,
        "pagination": pagination_data
    })


@app.post("/api/tracking-interfaces")
def api_create_tracking_interface():
    """创建轨迹接口"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建轨迹接口"}), 403

    data = request.get_json() or {}
    interface_name = (data.get("interface_name") or "").strip()
    request_url = (data.get("request_url") or "").strip()
    auth_params = data.get("auth_params") or ""
    status_mapping = data.get("status_mapping") or ""
    response_key_params = data.get("response_key_params") or ""
    fetch_interval = data.get("fetch_interval")

    if not interface_name:
        return jsonify({"success": False, "message": "轨迹接口名称不能为空", "field": "interface_name"}), 400

    if not request_url:
        return jsonify({"success": False, "message": "轨迹请求地址不能为空", "field": "request_url"}), 400

    if not fetch_interval:
        return jsonify({"success": False, "message": "获取频率不能为空", "field": "fetch_interval"}), 400

    try:
        fetch_interval = float(fetch_interval)
        if fetch_interval <= 0:
            return jsonify({"success": False, "message": "获取频率必须大于0", "field": "fetch_interval"}), 400
    except:
        return jsonify({"success": False, "message": "获取频率格式错误", "field": "fetch_interval"}), 400

    # 检查接口名称是否已存在
    existing = TrackingInterface.query.filter_by(interface_name=interface_name).first()
    if existing:
        return jsonify({"success": False, "message": f"轨迹接口名称'{interface_name}'已存在", "field": "interface_name"}), 400

    # 验证JSON格式
    if auth_params:
        try:
            json.loads(auth_params)
        except:
            return jsonify({"success": False, "message": "验证信息JSON格式错误", "field": "auth_params"}), 400

    if status_mapping:
        try:
            json.loads(status_mapping)
        except:
            return jsonify({"success": False, "message": "状态映射表JSON格式错误", "field": "status_mapping"}), 400
    
    if response_key_params:
        try:
            json.loads(response_key_params)
        except:
            return jsonify({"success": False, "message": "关键参数JSON格式错误", "field": "response_key_params"}), 400

    interface = TrackingInterface(
        interface_name=interface_name,
        request_url=request_url,
        auth_params=auth_params,
        status_mapping=status_mapping,
        response_key_params=response_key_params,
        fetch_interval=fetch_interval
    )
    db.session.add(interface)
    db.session.commit()

    return jsonify({"success": True, "id": interface.id})


@app.put("/api/tracking-interfaces/<int:interface_id>")
def api_update_tracking_interface(interface_id):
    """更新轨迹接口"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑轨迹接口"}), 403

    interface = TrackingInterface.query.get(interface_id)
    if not interface:
        return jsonify({"success": False, "message": "轨迹接口不存在"}), 404

    data = request.get_json() or {}
    interface_name = (data.get("interface_name") or "").strip()
    request_url = (data.get("request_url") or "").strip()
    auth_params = data.get("auth_params") or ""
    status_mapping = data.get("status_mapping") or ""
    response_key_params = data.get("response_key_params") or ""
    fetch_interval = data.get("fetch_interval")

    if not interface_name:
        return jsonify({"success": False, "message": "轨迹接口名称不能为空", "field": "interface_name"}), 400

    if not request_url:
        return jsonify({"success": False, "message": "轨迹请求地址不能为空", "field": "request_url"}), 400

    if not fetch_interval:
        return jsonify({"success": False, "message": "获取频率不能为空", "field": "fetch_interval"}), 400

    try:
        fetch_interval = float(fetch_interval)
        if fetch_interval <= 0:
            return jsonify({"success": False, "message": "获取频率必须大于0", "field": "fetch_interval"}), 400
    except:
        return jsonify({"success": False, "message": "获取频率格式错误", "field": "fetch_interval"}), 400

    # 检查接口名称是否与其他记录冲突
    existing = TrackingInterface.query.filter(TrackingInterface.interface_name == interface_name, TrackingInterface.id != interface_id).first()
    if existing:
        return jsonify({"success": False, "message": f"轨迹接口名称'{interface_name}'已存在", "field": "interface_name"}), 400

    # 验证JSON格式
    if auth_params:
        try:
            json.loads(auth_params)
        except:
            return jsonify({"success": False, "message": "验证信息JSON格式错误", "field": "auth_params"}), 400

    if status_mapping:
        try:
            json.loads(status_mapping)
        except:
            return jsonify({"success": False, "message": "状态映射表JSON格式错误", "field": "status_mapping"}), 400
    
    if response_key_params:
        try:
            json.loads(response_key_params)
        except:
            return jsonify({"success": False, "message": "关键参数JSON格式错误", "field": "response_key_params"}), 400

    interface.interface_name = interface_name
    interface.request_url = request_url
    interface.auth_params = auth_params
    interface.status_mapping = status_mapping
    interface.response_key_params = response_key_params
    interface.fetch_interval = fetch_interval
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/tracking-interfaces/<int:interface_id>")
def api_delete_tracking_interface(interface_id):
    """删除轨迹接口"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除轨迹接口"}), 403

    interface = TrackingInterface.query.get(interface_id)
    if not interface:
        return jsonify({"success": False, "message": "轨迹接口不存在"}), 404

    db.session.delete(interface)
    db.session.commit()

    return jsonify({"success": True})


# ==================== 尾程轨迹状态映射表 API ====================

@app.get("/api/lastmile-status-mappings")
def api_get_lastmile_status_mappings():
    """获取尾程轨迹状态映射表列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=50)

    query = LastmileStatusMapping.query.order_by(LastmileStatusMapping.id.desc())
    
    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        mappings = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        mappings = query.all()
        pagination_data = None
    
    mappings_data = []
    for mapping in mappings:
        mappings_data.append({
            "id": mapping.id,
            "description": mapping.description,
            "sub_status": mapping.sub_status,
            "system_status_code": mapping.system_status_code,
            "created_at": mapping.created_at.isoformat() if mapping.created_at else None,
        })

    return jsonify({
        "success": True,
        "mappings": mappings_data,
        "pagination": pagination_data
    })


@app.post("/api/lastmile-status-mappings")
def api_create_lastmile_status_mapping():
    """创建尾程轨迹状态映射"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建尾程映射"}), 403

    data = request.get_json() or {}
    description = (data.get("description") or "").strip()  # description 为非必填，允许为空
    sub_status = (data.get("sub_status") or "").strip()
    system_status_code = (data.get("system_status_code") or "").strip()

    # description 为非必填，不需要验证

    if not sub_status:
        return jsonify({"success": False, "message": "尾程轨迹状态不能为空", "field": "sub_status"}), 400

    if not system_status_code:
        return jsonify({"success": False, "message": "系统状态不能为空", "field": "system_status_code"}), 400

    # 检查系统状态代码是否存在
    node = TrackingNode.query.filter_by(status_code=system_status_code).first()
    if not node:
        return jsonify({"success": False, "message": f"系统状态代码'{system_status_code}'不存在", "field": "system_status_code"}), 400

    # 检查是否已存在相同的映射（description 可为空，所以需要同时比较 description 和 sub_status）
    if description:  # 如果 description 有值，检查 description + sub_status 组合
        existing = LastmileStatusMapping.query.filter_by(description=description, sub_status=sub_status).first()
        if existing:
            return jsonify({"success": False, "message": f"相同的映射已存在（description='{description}', sub_status='{sub_status}'）"}), 400
    else:  # 如果 description 为空，检查 sub_status 是否已存在（description 也为空的情况）
        existing = LastmileStatusMapping.query.filter(
            LastmileStatusMapping.sub_status == sub_status,
            db.or_(LastmileStatusMapping.description == '', LastmileStatusMapping.description.is_(None))
        ).first()
        if existing:
            return jsonify({"success": False, "message": f"相同的映射已存在（sub_status='{sub_status}'，且 description 为空）"}), 400

    mapping = LastmileStatusMapping(
        description=description,
        sub_status=sub_status,
        system_status_code=system_status_code
    )
    db.session.add(mapping)
    db.session.commit()

    return jsonify({"success": True, "id": mapping.id})


@app.put("/api/lastmile-status-mappings/<int:mapping_id>")
def api_update_lastmile_status_mapping(mapping_id):
    """更新尾程轨迹状态映射"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑尾程映射"}), 403

    mapping = LastmileStatusMapping.query.get(mapping_id)
    if not mapping:
        return jsonify({"success": False, "message": "映射不存在"}), 404

    data = request.get_json() or {}
    description = (data.get("description") or "").strip()  # description 为非必填，允许为空
    sub_status = (data.get("sub_status") or "").strip()
    system_status_code = (data.get("system_status_code") or "").strip()

    # description 为非必填，不需要验证

    if not sub_status:
        return jsonify({"success": False, "message": "尾程轨迹状态不能为空", "field": "sub_status"}), 400

    if not system_status_code:
        return jsonify({"success": False, "message": "系统状态不能为空", "field": "system_status_code"}), 400

    # 检查系统状态代码是否存在
    node = TrackingNode.query.filter_by(status_code=system_status_code).first()
    if not node:
        return jsonify({"success": False, "message": f"系统状态代码'{system_status_code}'不存在", "field": "system_status_code"}), 400

    # 检查是否与其他记录冲突
    if description:  # 如果 description 有值，检查 description + sub_status 组合
        existing = LastmileStatusMapping.query.filter(
            LastmileStatusMapping.description == description,
            LastmileStatusMapping.sub_status == sub_status,
            LastmileStatusMapping.id != mapping_id
        ).first()
        if existing:
            return jsonify({"success": False, "message": f"相同的映射已存在（description='{description}', sub_status='{sub_status}'）"}), 400
    else:  # 如果 description 为空，检查 sub_status 是否已存在（description 也为空的情况）
        existing = LastmileStatusMapping.query.filter(
            LastmileStatusMapping.sub_status == sub_status,
            db.or_(LastmileStatusMapping.description == '', LastmileStatusMapping.description.is_(None)),
            LastmileStatusMapping.id != mapping_id
        ).first()
        if existing:
            return jsonify({"success": False, "message": f"相同的映射已存在（sub_status='{sub_status}'，且 description 为空）"}), 400

    mapping.description = description
    mapping.sub_status = sub_status
    mapping.system_status_code = system_status_code
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/lastmile-status-mappings/<int:mapping_id>")
def api_delete_lastmile_status_mapping(mapping_id):
    """删除尾程轨迹状态映射"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除尾程映射"}), 403

    mapping = LastmileStatusMapping.query.get(mapping_id)
    if not mapping:
        return jsonify({"success": False, "message": "映射不存在"}), 404

    db.session.delete(mapping)
    db.session.commit()

    return jsonify({"success": True})


# ==================== 轨迹数据管理 API ====================

@app.get("/api/tracking-data")
def api_get_tracking_data():
    """获取轨迹数据列表（只显示有轨迹接口绑定的运单）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int, default=1)
    per_page = request.args.get("per_page", type=int, default=200)
    
    # 搜索条件
    interface_name = request.args.get("interface_name", type=str, default="")
    stop_tracking = request.args.get("stop_tracking", type=str, default="")  # "true" 或 "false"
    order_nos = request.args.get("order_nos", type=str, default="")  # 订单号搜索（多个，逗号分隔）
    transfer_nos = request.args.get("transfer_nos", type=str, default="")  # 转单号搜索（多个，逗号分隔）
    start_date = request.args.get("start_date", type=str, default="")  # 开始日期
    end_date = request.args.get("end_date", type=str, default="")  # 结束日期

    # 查询有轨迹接口绑定的运单
    query = db.session.query(
        Waybill.id,
        Waybill.order_no,
        Waybill.transfer_no,
        Waybill.order_time,
        TrackingInfo.status_code,
        TrackingInfo.last_fetch_time,
        TrackingInfo.last_push_time,
        TrackingInfo.raw_response,
        TrackingInfo.lastmile_no,
        TrackingInfo.lastmile_register_response,
        TrackingInfo.lastmile_tracking_response,
        TrackingInfo.push_events,
        TrackingInfo.szpost_response,
        TrackingInfo.stop_tracking,
        TrackingInfo.id.label('tracking_id'),
        TrackingInterface.interface_name
    ).join(
        Product, Waybill.product_id == Product.id
    ).outerjoin(
        TrackingInfo, Waybill.id == TrackingInfo.waybill_id
    ).outerjoin(
        TrackingInterface, Product.tracking_interface_id == TrackingInterface.id
    ).filter(
        Product.tracking_interface_id.isnot(None)
    )
    
    # 按接口名称筛选
    if interface_name:
        query = query.filter(TrackingInterface.interface_name == interface_name)
    
    # 按订单号筛选（支持多个，换行或逗号分隔）
    if order_nos:
        # 将换行符或逗号分隔的订单号分割为列表
        order_list = [o.strip() for o in order_nos.replace('\n', ',').replace('\r', '').split(',') if o.strip()]
        if order_list:
            query = query.filter(Waybill.order_no.in_(order_list))
    
    # 按转单号筛选（支持多个，换行或逗号分隔）
    if transfer_nos:
        # 将换行符或逗号分隔的转单号分割为列表
        transfer_list = [t.strip() for t in transfer_nos.replace('\n', ',').replace('\r', '').split(',') if t.strip()]
        if transfer_list:
            query = query.filter(Waybill.transfer_no.in_(transfer_list))
    
    # 按下单日期筛选
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Waybill.order_time >= start_dt)
        except:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Waybill.order_time < end_dt)
        except:
            pass
    
    # 按停止跟踪状态筛选
    if stop_tracking == "true":
        query = query.filter(TrackingInfo.stop_tracking == True)
    elif stop_tracking == "false":
        query = query.filter(db.or_(
            TrackingInfo.stop_tracking == False,
            TrackingInfo.stop_tracking.is_(None)
        ))
    
    query = query.order_by(Waybill.order_time.desc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    tracking_data = []
    for row in pagination.items:
        has_raw_response = bool(row.raw_response)
        # 检查是否有尾程报文（注册或查询任意一个有值即可）
        has_lastmile_response = bool(row.lastmile_register_response or row.lastmile_tracking_response)
        
        # 优先使用数据库存储的尾程单号，如果没有则从原始报文解析
        last_mile_no = row.lastmile_no or ""
        if not last_mile_no and row.raw_response:
            try:
                raw_data = json.loads(row.raw_response)
                # 从通邮接口报文中提取transferNo
                if "tracks" in raw_data and raw_data["tracks"]:
                    last_mile_no = raw_data["tracks"][0].get("transferNo", "")
            except:
                pass
        
        # 获取状态码对应的中文描述（比较头程和尾程，取最新的）
        status_description = ""
        final_status_code = ""
        final_tracking_time = None
        
        # 头程时间和状态码
        headhaul_time = None
        headhaul_status_code = row.status_code or ""
        if row.raw_response:
            try:
                raw_data = json.loads(row.raw_response)
                if "tracks" in raw_data and raw_data["tracks"]:
                    track_info_list = raw_data["tracks"][0].get("trackInfo", [])
                    if track_info_list:
                        # 按时间排序取最新
                        track_info_list.sort(key=lambda x: x.get('changeDate', 0), reverse=True)
                        latest_track = track_info_list[0]
                        change_date = latest_track.get('changeDate')
                        if change_date:
                            from datetime import timezone
                            # 添加UTC时区信息，使其成为aware datetime
                            headhaul_time = datetime.fromtimestamp(change_date / 1000.0, tz=timezone.utc)
            except:
                pass
        
        # 尾程时间和状态码
        lastmile_time = None
        lastmile_status_code = ""
        if row.lastmile_tracking_response:
            try:
                lastmile_data = json.loads(row.lastmile_tracking_response)
                if isinstance(lastmile_data, dict) and lastmile_data.get("code") == 0:
                    data_content = lastmile_data.get("data", {})
                    if isinstance(data_content, dict):
                        accepted = data_content.get("accepted", [])
                        if accepted and len(accepted) > 0:
                            track_info = accepted[0].get("track_info", {})
                            latest_event = track_info.get("latest_event", {})
                            if latest_event:
                                time_iso = latest_event.get("time_iso", "")
                                sub_status = latest_event.get("sub_status", "")
                                description = latest_event.get("description", "")
                                
                                # 解析时间
                                if time_iso:
                                    try:
                                        lastmile_time = datetime.fromisoformat(time_iso.replace('Z', '+00:00'))
                                    except:
                                        pass
                                
                                # 匹配尾程状态码
                                if sub_status:
                                    mapping = None
                                    if description:
                                        mapping = LastmileStatusMapping.query.filter_by(
                                            description=description,
                                            sub_status=sub_status
                                        ).first()
                                    if not mapping:
                                        mapping = LastmileStatusMapping.query.filter(
                                            LastmileStatusMapping.sub_status == sub_status,
                                            db.or_(
                                                LastmileStatusMapping.description == '',
                                                LastmileStatusMapping.description.is_(None)
                                            )
                                        ).first()
                                    if mapping:
                                        lastmile_status_code = mapping.system_status_code
            except:
                pass
        
        # 比较时间，取最新的状态码
        if headhaul_time and lastmile_time:
            if lastmile_time > headhaul_time:
                final_status_code = lastmile_status_code
                final_tracking_time = lastmile_time
            else:
                final_status_code = headhaul_status_code
                final_tracking_time = headhaul_time
        elif headhaul_time:
            final_status_code = headhaul_status_code
            final_tracking_time = headhaul_time
        elif lastmile_time:
            final_status_code = lastmile_status_code
            final_tracking_time = lastmile_time
        else:
            final_status_code = headhaul_status_code  # 都没有时间，使用头程
        
        # 根据最终状态码获取描述
        if final_status_code:
            node = TrackingNode.query.filter_by(status_code=final_status_code).first()
            if node:
                status_description = node.status_description
            else:
                status_description = final_status_code
        
        tracking_data.append({
            "waybill_id": row.id,
            "order_no": row.order_no,
            "transfer_no": row.transfer_no or "",
            "last_mile_no": last_mile_no,
            "order_time": row.order_time.isoformat() if row.order_time else None,
            "status_code": final_status_code or "",
            "status_description": status_description,
            "last_fetch_time": row.last_fetch_time.isoformat() if row.last_fetch_time else None,
            "last_push_time": row.last_push_time.isoformat() if row.last_push_time else None,
            "stop_tracking": row.stop_tracking or False,
            "tracking_id": row.tracking_id,
            "interface_name": row.interface_name or "",
            "has_raw_response": has_raw_response,
            "has_lastmile_response": has_lastmile_response,
            "has_push_events": bool(row.push_events),
            "has_szpost_response": bool(row.szpost_response)
        })

    return jsonify({
        "success": True,
        "tracking_data": tracking_data,
        "pagination": {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    })


@app.get("/api/tracking-data/<int:tracking_id>/details")
def api_get_tracking_details(tracking_id):
    """获取轨迹详情"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    tracking = TrackingInfo.query.get(tracking_id)
    if not tracking:
        return jsonify({"success": False, "message": "轨迹信息不存在"}), 404

    return jsonify({
        "success": True,
        "tracking": {
            "id": tracking.id,
            "order_no": tracking.order_no,
            "transfer_no": tracking.transfer_no or "",
            "tracking_description": tracking.tracking_description or "",
            "status_code": tracking.status_code or "",
            "tracking_time": tracking.tracking_time.isoformat() if tracking.tracking_time else None,
            "last_fetch_time": tracking.last_fetch_time.isoformat() if tracking.last_fetch_time else None,
            "last_push_time": tracking.last_push_time.isoformat() if tracking.last_push_time else None,
            "raw_response": tracking.raw_response or "",
            "lastmile_register_response": tracking.lastmile_register_response or "",
            "lastmile_tracking_response": tracking.lastmile_tracking_response or ""
        }
    })


@app.post("/api/tracking-data/push")
def api_push_tracking_data():
    """手动推送轨迹数据到深邮接口"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限推送轨迹"}), 403

    data = request.get_json() or {}
    tracking_ids = data.get("tracking_ids") or []

    if not tracking_ids:
        return jsonify({"success": False, "message": "请选择要推送的轨迹"}), 400

    try:
        from tracking_handler.push_szpost_handler import push_tracking_to_szpost
        import json
        from datetime import datetime
        
        # 查询轨迹信息和推送报文
        trackings = TrackingInfo.query.filter(TrackingInfo.id.in_(tracking_ids)).all()
        
        if not trackings:
            return jsonify({"success": False, "message": "未找到轨迹信息"}), 404
        
        # 获取所有轨迹节点映射
        all_status_codes = set()
        for tracking in trackings:
            if tracking.push_events:
                events = json.loads(tracking.push_events)
                for event in events:
                    status_code = event.get('status_code')
                    if status_code:
                        all_status_codes.add(status_code)
        
        # 查询轨迹节点信息
        tracking_nodes = TrackingNode.query.filter(TrackingNode.status_code.in_(all_status_codes)).all()
        tracking_nodes_map = {node.status_code: node for node in tracking_nodes}
        
        # 合并所有推送事件
        all_push_events = []
        for tracking in trackings:
            if tracking.push_events:
                events = json.loads(tracking.push_events)
                all_push_events.extend(events)
        
        if not all_push_events:
            return jsonify({"success": False, "message": "没有可推送的轨迹数据"}), 400
        
        # 推送到深邮接口
        result = push_tracking_to_szpost(all_push_events, tracking_nodes_map)
        
        # 保存响应报文
        now = datetime.utcnow()
        response_json = json.dumps(result, ensure_ascii=False)
        
        for tracking in trackings:
            tracking.szpost_response = response_json
            tracking.last_push_time = now
        
        db.session.commit()
        
        if result.get('success'):
            return jsonify({
                "success": True,
                "message": f"成功推送 {len(tracking_ids)} 条轨迹数据"
            })
        else:
            return jsonify({
                "success": False,
                "message": f"推送失败: {result.get('message')}"
            }), 500
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"推送异常: {str(e)}"
        }), 500


@app.post("/api/tracking-data/fetch")
def api_fetch_tracking_data():
    """手动获取轨迹数据（异步）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限获取轨迹"}), 403

    data = request.get_json() or {}
    waybill_ids = data.get("waybill_ids") or []

    if not waybill_ids:
        return jsonify({"success": False, "message": "请选择要获取轨迹的运单"}), 400
    
    try:
        # 提交异步任务
        task = async_fetch_tracking_task.delay(waybill_ids)
        
        # 记录任务
        new_task = TaskRecord(
            task_id=task.id,
            task_name=f"获取头程轨迹({len(waybill_ids)}单)",
            status="PENDING"
        )
        db.session.add(new_task)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"已提交异步任务，共{len(waybill_ids)}单，请稍后查看结果",
            "task_id": task.id
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"提交任务失败：{str(e)}"
        }), 500

    # 查询运单信息和关联的接口配置
    waybills_query = db.session.query(
        Waybill.id,
        Waybill.order_no,
        Waybill.transfer_no,
        Waybill.product_id,
        Product.tracking_interface_id,
        TrackingInterface.id.label('interface_id'),
        TrackingInterface.interface_name,
        TrackingInterface.request_url,
        TrackingInterface.auth_params,
        TrackingInterface.status_mapping,
        TrackingInterface.response_key_params
    ).join(
        Product, Waybill.product_id == Product.id
    ).outerjoin(
        TrackingInterface, Product.tracking_interface_id == TrackingInterface.id
    ).filter(
        Waybill.id.in_(waybill_ids)
    ).all()

    if not waybills_query:
        return jsonify({"success": False, "message": "未找到运单信息"}), 404

    # 按接口分组
    from tracking_handler.tracking_handler_manager import batch_fetch_tracking_by_interface
    import json
    from datetime import datetime

    interface_groups = {}
    for row in waybills_query:
        if not row.interface_id:
            continue
        
        interface_id = row.interface_id
        if interface_id not in interface_groups:
            interface_groups[interface_id] = {
                'config': {
                    'interface_name': row.interface_name,
                    'request_url': row.request_url,
                    'auth_params': row.auth_params
                },
                'status_mapping': json.loads(row.status_mapping) if row.status_mapping else [],
                'response_key_params': json.loads(row.response_key_params) if row.response_key_params else None,
                'waybills': []
            }
        
        interface_groups[interface_id]['waybills'].append({
            'waybill_id': row.id,
            'order_no': row.order_no,
            'transfer_no': row.transfer_no
        })

    # 批量获取轨迹
    success_count = 0
    fail_count = 0
    error_details = []  # 记录失败详情
    
    for interface_id, group_data in interface_groups.items():
        results = batch_fetch_tracking_by_interface(
            group_data['waybills'],
            group_data['config'],
            group_data['status_mapping'],
            group_data.get('response_key_params')
        )
        
        # 保存结果到数据库
        for result in results:
            if not result.get('success'):
                fail_count += 1
                # 记录失败原因
                error_details.append({
                    'order_no': result.get('order_no', 'Unknown'),
                    'transfer_no': result.get('transfer_no', 'Unknown'),
                    'error': result.get('message', '未知错误')
                })
                continue
            
            waybill_id = result['waybill_id']
            order_no = result['order_no']
            transfer_no = result['transfer_no']
            
            # 查询运单信息，用于检查停止条件
            waybill = Waybill.query.get(waybill_id)
            if not waybill:
                fail_count += 1
                error_details.append({
                    'order_no': order_no,
                    'transfer_no': transfer_no,
                    'error': '运单不存在'
                })
                continue
            
            # 查找或创建轨迹记录
            tracking = TrackingInfo.query.filter_by(
                waybill_id=waybill_id
            ).first()
            
            now = datetime.utcnow()
            
            if tracking:
                # 更新现有记录
                tracking.tracking_description = result.get('tracking_description', '')
                tracking.status_code = result.get('status_code', '')
                tracking.tracking_time = result.get('tracking_time')
                tracking.raw_response = result.get('raw_response', '')
                tracking.last_fetch_time = now
                
                # 从json中提取尾程单号并保存
                try:
                    raw_data = json.loads(result.get('raw_response', '{}'))
                    if "tracks" in raw_data and raw_data["tracks"]:
                        lastmile_no = raw_data["tracks"][0].get("transferNo", "")
                        if lastmile_no:
                            tracking.lastmile_no = lastmile_no
                except:
                    pass
            else:
                # 创建新记录
                # 先提取尾程单号
                lastmile_no = ""
                try:
                    raw_data = json.loads(result.get('raw_response', '{}'))
                    if "tracks" in raw_data and raw_data["tracks"]:
                        lastmile_no = raw_data["tracks"][0].get("transferNo", "")
                except:
                    pass
                
                tracking = TrackingInfo(
                    waybill_id=waybill_id,
                    order_no=order_no,
                    transfer_no=transfer_no,
                    tracking_interface_id=interface_id,
                    tracking_description=result.get('tracking_description', ''),
                    status_code=result.get('status_code', ''),
                    tracking_time=result.get('tracking_time'),
                    raw_response=result.get('raw_response', ''),
                    lastmile_no=lastmile_no if lastmile_no else None,
                    last_fetch_time=now
                )
                db.session.add(tracking)
            
            # 刷新tracking对象，确保status_code已更新
            db.session.flush()
            db.session.refresh(tracking)
            
            # 检查是否应该停止跟踪（在status_code更新后检查）
            should_stop, reason = should_stop_tracking(waybill, tracking)
            if should_stop:
                tracking.stop_tracking = True
                tracking.stop_tracking_reason = reason
                tracking.stop_tracking_time = now
            
            success_count += 1
    
    try:
        db.session.commit()
        message = f"成功获取 {success_count} 条运单轨迹数据"
        if fail_count > 0:
            message += f"，{fail_count} 条失败"
            # 添加错误详情
            if error_details:
                error_summary = "; ".join([f"{e['transfer_no']}: {e['error']}" for e in error_details[:3]])
                message += f" ({error_summary})"
        
        return jsonify({
            "success": True,
            "message": message,
            "error_details": error_details  # 返回详细错误信息
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"保存失败: {str(e)}"}), 500


@app.post("/api/tracking-data/batch-check-stop")
def api_batch_check_stop_tracking():
    """批量检查运单，标记满足停止条件的运单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401


@app.get("/api/tracking-data/<int:tracking_id>/push-events")
def api_get_push_events(tracking_id):
    """获取推送报文"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    tracking = TrackingInfo.query.get(tracking_id)
    if not tracking:
        return jsonify({"success": False, "message": "轨迹信息不存在"}), 404
    
    import json
    push_events = json.loads(tracking.push_events) if tracking.push_events else []
    
    return jsonify({
        "success": True,
        "push_events": push_events,
        "order_no": tracking.order_no
    })


@app.put("/api/tracking-data/<int:tracking_id>/push-events")
def api_update_push_events(tracking_id):
    """更新推送报文（删除某个事件）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    tracking = TrackingInfo.query.get(tracking_id)
    if not tracking:
        return jsonify({"success": False, "message": "轨迹信息不存在"}), 404
    
    import json
    data = request.get_json() or {}
    push_events = data.get("push_events", [])
    
    # 更新推送报文
    tracking.push_events = json.dumps(push_events, ensure_ascii=False)
    db.session.commit()
    
    return jsonify({
        "success": True,
        "message": "更新成功"
    })


@app.get("/api/tracking-data/<int:tracking_id>/szpost-response")
def api_get_szpost_response(tracking_id):
    """获取深邮响应报文"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    tracking = TrackingInfo.query.get(tracking_id)
    if not tracking:
        return jsonify({"success": False, "message": "轨迹信息不存在"}), 404
    
    return jsonify({
        "success": True,
        "szpost_response": tracking.szpost_response or '',
        "order_no": tracking.order_no
    })


@app.post("/api/tracking-data/batch-check-stop")
def api_batch_check_stop_tracking_old():
    """批量检查运单，标记满足停止条件的运单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限执行此操作"}), 403

    try:
        result = batch_check_stop_tracking()
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify(result), 500
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"执行失败: {str(e)}"
        }), 500


@app.post("/api/tracking-data/fetch-lastmile")
def api_fetch_lastmile_tracking():
    """获取尾程轨迹数据（异步）"""
    try:
        print(f"[尾程轨迹] 收到请求，方法: {request.method}")
        
        current_user = session.get("user")
        if not current_user:
            print("[尾程轨迹] 错误：未登录")
            return jsonify({"success": False, "message": "未登录"}), 401

        if current_user.get("role") != "系统管理员":
            print(f"[尾程轨迹] 错误：权限不足，用户角色: {current_user.get('role')}")
            return jsonify({"success": False, "message": "无权限获取尾程轨迹"}), 403

        data = request.get_json() or {}
        waybill_ids = data.get("waybill_ids") or []
        print(f"[尾程轨迹] 运单ID列表: {waybill_ids}")
        
        if not waybill_ids:
            print("[尾程轨迹] 错误：未选择运单")
            return jsonify({"success": False, "message": "请选择要获取尾程轨迹的运单"}), 400
        
        # 提交异步任务
        task = async_fetch_lastmile_tracking_task.delay(waybill_ids)
        print(f"[尾程轨迹] 任务已提交，Task ID: {task.id}")
        
        # 记录任务
        new_task = TaskRecord(
            task_id=task.id,
            task_name=f"获取尾程轨迹({len(waybill_ids)}单)",
            status="PENDING"
        )
        db.session.add(new_task)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"已提交异步任务，共{len(waybill_ids)}单，请稍后查看结果",
            "task_id": task.id
        })
    
    except Exception as e:
        error_msg = f"提交任务失败: {str(e)}"
        print(f"[尾程轨迹] 异常: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": error_msg
        }), 500




@app.post("/api/tracking-data/import-lastmile")
def api_import_lastmile_numbers():
    """导入尾程单号"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限导入"}), 403

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "没有上传文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "文件名为空"}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "请上传Excel文件"}), 400

    try:
        import pandas as pd
        from datetime import datetime
        
        # 读取Excel
        df = pd.read_excel(file, dtype=str)
        
        # 验证表头
        if '订单号' not in df.columns or '尾程单号' not in df.columns:
            return jsonify({"success": False, "message": "Excel表头必须包含：订单号、尾程单号"}), 400
        
        # 验证每一行
        errors = []
        to_update = []
        
        for idx, row in df.iterrows():
            order_no = str(row['订单号']).strip() if pd.notna(row['订单号']) else ""
            lastmile_no = str(row['尾程单号']).strip() if pd.notna(row['尾程单号']) else ""
            
            if not order_no:
                errors.append(f"第{idx+2}行：订单号为空")
                break
            
            if not lastmile_no:
                errors.append(f"第{idx+2}行：尾程单号为空")
                break
            
            # 检查订单号是否存在
            waybill = Waybill.query.filter_by(order_no=order_no).first()
            if not waybill:
                errors.append(f"第{idx+2}行：订单号'{order_no}'不存在")
                break
            
            # 查找或创建轨迹记录（UPSERT逻辑）
            tracking = TrackingInfo.query.filter_by(waybill_id=waybill.id).first()
            if not tracking:
                # 如果没有轨迹记录（例如删除后重新导入），创建一条新记录
                product = Product.query.get(waybill.product_id)
                tracking_interface_id = product.tracking_interface_id if product else None
                
                tracking = TrackingInfo(
                    waybill_id=waybill.id,
                    order_no=waybill.order_no,
                    transfer_no=waybill.transfer_no,
                    tracking_interface_id=tracking_interface_id,
                    lastmile_no=lastmile_no
                )
                db.session.add(tracking)
                # 立即flush以获取ID，避免后续操作冲突
                db.session.flush()
            
            to_update.append({
                'tracking': tracking,
                'lastmile_no': lastmile_no,
                'waybill_id': waybill.id
            })
        
        # 如果有错误，返回错误
        if errors:
            return jsonify({"success": False, "message": errors[0]}), 400
        
        # 先保存尾程单号到数据库，立即返回
        now = datetime.utcnow()
        for item in to_update:
            item['tracking'].lastmile_no = item['lastmile_no']
        
        db.session.commit()
        
        # 返回成功，先不触发获取（用户可以手动点击“获取尾程单轨迹”按钮）
        return jsonify({
            "success": True,
            "message": f"成功导入{len(to_update)}条尾程单号，请勾选运单后点击“获取尾程单轨迹”按钮获取轨迹信息"
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"导入失败：{str(e)}"}), 500


@app.get("/api/tracking-data/lastmile-template")
def api_download_lastmile_template():
    """下载尾程单号导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "尾程单号导入"
    
    # 设置表头
    headers = ['订单号', '尾程单号']
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    
    # 添加示例数据
    ws.append(['ZC12345678', 'GFUS01028251402241'])
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='尾程单号导入模板.xlsx'
    )


# ==================== 客户管理 API ====================
@app.get("/api/customers")
def api_get_customers():
    """获取客户列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = Customer.query.order_by(Customer.id.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        customers = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        customers = query.all()
        pagination_data = None

    customers_data = []
    for customer in customers:
        customers_data.append({
            "id": customer.id,
            "full_name": customer.full_name,
            "short_name": customer.short_name,
            "customer_types": customer.customer_types.split(",") if customer.customer_types else [],
            "contact_person": customer.contact_person or "",
            "email": customer.email or "",
            "remark": customer.remark or "",
            "created_at": customer.created_at.isoformat() if customer.created_at else None,
        })

    return jsonify({
        "success": True, 
        "customers": customers_data,
        "pagination": pagination_data
    })


@app.post("/api/customers")
def api_create_customer():
    """创建客户"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建客户"}), 403

    data = request.get_json() or {}
    full_name = (data.get("full_name") or "").strip()
    short_name = (data.get("short_name") or "").strip()
    customer_types = data.get("customer_types") or []
    contact_person = (data.get("contact_person") or "").strip()
    email = (data.get("email") or "").strip()
    remark = (data.get("remark") or "").strip()

    # 验证必填字段
    if not full_name:
        return jsonify({"success": False, "message": "客户全称不能为空", "field": "full_name"}), 400

    if not short_name:
        return jsonify({"success": False, "message": "客户简称不能为空", "field": "short_name"}), 400

    if not customer_types or len(customer_types) == 0:
        return jsonify({"success": False, "message": "请至少选择一种客户类别", "field": "customer_types"}), 400

    # 验证邮箱格式（如果填写了）
    if email:
        import re
        # 支持多个邮箱，用中英文分号分隔
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        # 分隔邮箱（支持; 和 ；）
        emails = re.split(r'[;；]', email)
        for e in emails:
            e = e.strip()
            if e and not re.match(email_pattern, e):
                return jsonify({"success": False, "message": f"邮箱格式不正确：{e}", "field": "email"}), 400

    # 将数组转为逗号分隔的字符串
    customer_types_str = ",".join(customer_types)

    customer = Customer(
        full_name=full_name,
        short_name=short_name,
        customer_types=customer_types_str,
        contact_person=contact_person,
        email=email,
        remark=remark
    )
    db.session.add(customer)
    db.session.commit()

    return jsonify({"success": True, "id": customer.id})


@app.put("/api/customers/<int:customer_id>")
def api_update_customer(customer_id):
    """更新客户"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑客户"}), 403

    customer = Customer.query.get(customer_id)
    if not customer:
        return jsonify({"success": False, "message": "客户不存在"}), 404

    data = request.get_json() or {}
    full_name = (data.get("full_name") or "").strip()
    short_name = (data.get("short_name") or "").strip()
    customer_types = data.get("customer_types") or []
    contact_person = (data.get("contact_person") or "").strip()
    email = (data.get("email") or "").strip()
    remark = (data.get("remark") or "").strip()

    # 验证必填字段
    if not full_name:
        return jsonify({"success": False, "message": "客户全称不能为空", "field": "full_name"}), 400

    if not short_name:
        return jsonify({"success": False, "message": "客户简称不能为空", "field": "short_name"}), 400

    if not customer_types or len(customer_types) == 0:
        return jsonify({"success": False, "message": "请至少选择一种客户类别", "field": "customer_types"}), 400

    # 验证邮箱格式
    if email:
        import re
        # 支持多个邮箱，用中英文分号分隔
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        # 分隔邮箱（支持; 和 ；）
        emails = re.split(r'[;；]', email)
        for e in emails:
            e = e.strip()
            if e and not re.match(email_pattern, e):
                return jsonify({"success": False, "message": f"邮箱格式不正确：{e}", "field": "email"}), 400

    customer.full_name = full_name
    customer.short_name = short_name
    customer.customer_types = ",".join(customer_types)
    customer.contact_person = contact_person
    customer.email = email
    customer.remark = remark
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/customers/<int:customer_id>")
def api_delete_customer(customer_id):
    """删除客户"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除客户"}), 403

    customer = Customer.query.get(customer_id)
    if not customer:
        return jsonify({"success": False, "message": "客户不存在"}), 404

    db.session.delete(customer)
    db.session.commit()

    return jsonify({"success": True})


# ==================== 供应商管理 API ====================

@app.get("/api/suppliers")
def api_get_suppliers():
    """获取供应商列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = Supplier.query.order_by(Supplier.id.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        suppliers = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        suppliers = query.all()
        pagination_data = None

    suppliers_data = []
    for supplier in suppliers:
        suppliers_data.append({
            "id": supplier.id,
            "full_name": supplier.full_name,
            "short_name": supplier.short_name,
            "contact_person": supplier.contact_person or "",
            "email": supplier.email or "",
            "remark": supplier.remark or "",
            "created_at": supplier.created_at.isoformat() if supplier.created_at else None,
        })

    return jsonify({
        "success": True, 
        "suppliers": suppliers_data,
        "pagination": pagination_data
    })


@app.post("/api/suppliers")
def api_create_supplier():
    """创建供应商"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建供应商"}), 403

    data = request.get_json() or {}
    full_name = (data.get("full_name") or "").strip()
    short_name = (data.get("short_name") or "").strip()
    contact_person = (data.get("contact_person") or "").strip()
    email = (data.get("email") or "").strip()
    remark = (data.get("remark") or "").strip()

    # 验证必填字段
    if not full_name:
        return jsonify({"success": False, "message": "供应商全称不能为空", "field": "full_name"}), 400

    if not short_name:
        return jsonify({"success": False, "message": "供应商简称不能为空", "field": "short_name"}), 400

    # 验证邮箱格式（如果填写了）
    if email:
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return jsonify({"success": False, "message": "邮箱格式不正确", "field": "email"}), 400

    supplier = Supplier(
        full_name=full_name,
        short_name=short_name,
        contact_person=contact_person,
        email=email,
        remark=remark
    )
    db.session.add(supplier)
    db.session.commit()

    return jsonify({"success": True, "id": supplier.id})


@app.put("/api/suppliers/<int:supplier_id>")
def api_update_supplier(supplier_id):
    """更新供应商"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑供应商"}), 403

    supplier = Supplier.query.get(supplier_id)
    if not supplier:
        return jsonify({"success": False, "message": "供应商不存在"}), 404

    data = request.get_json() or {}
    full_name = (data.get("full_name") or "").strip()
    short_name = (data.get("short_name") or "").strip()
    contact_person = (data.get("contact_person") or "").strip()
    email = (data.get("email") or "").strip()
    remark = (data.get("remark") or "").strip()

    # 验证必填字段
    if not full_name:
        return jsonify({"success": False, "message": "供应商全称不能为空", "field": "full_name"}), 400

    if not short_name:
        return jsonify({"success": False, "message": "供应商简称不能为空", "field": "short_name"}), 400

    # 验证邮箱格式
    if email:
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return jsonify({"success": False, "message": "邮箱格式不正确", "field": "email"}), 400

    supplier.full_name = full_name
    supplier.short_name = short_name
    supplier.contact_person = contact_person
    supplier.email = email
    supplier.remark = remark
    db.session.commit()

    return jsonify({"success": True})


@app.delete("/api/suppliers/<int:supplier_id>")
def api_delete_supplier(supplier_id):
    """删除供应商"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除供应商"}), 403

    supplier = Supplier.query.get(supplier_id)
    if not supplier:
        return jsonify({"success": False, "message": "供应商不存在"}), 404

    db.session.delete(supplier)
    db.session.commit()

    return jsonify({"success": True})


# ==================== 客户报价管理 API ====================

@app.get("/api/customer-quotes")
def api_get_customer_quotes():
    """获取客户报价列表（支持搜索）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    # 获取搜索参数
    customer_id = request.args.get("customer_id", type=int)
    quote_type = request.args.get("quote_type", "")
    valid_date = request.args.get("valid_date", "")  # YYYY-MM-DD
    status = request.args.get("status", "")  # 生效中/已失效
    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    # 构建查询
    query = CustomerQuote.query
    now = datetime.now()

    if customer_id:
        query = query.filter(CustomerQuote.customer_id == customer_id)
    
    if quote_type:
        query = query.filter(CustomerQuote.quote_type == quote_type)
    
    if valid_date:
        try:
            search_date = datetime.strptime(valid_date, "%Y-%m-%d")
            # 查找在该日期有效的报价
            query = query.filter(
                CustomerQuote.valid_from <= search_date,
                CustomerQuote.valid_to >= search_date
            )
        except:
            pass

    if status == "生效中":
        query = query.filter(CustomerQuote.valid_from <= now, CustomerQuote.valid_to >= now)
    elif status == "已失效":
        query = query.filter((CustomerQuote.valid_from > now) | (CustomerQuote.valid_to < now))

    # 按有效期起始时间降序排列（最近的在最上面）
    query = query.order_by(CustomerQuote.valid_from.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        quotes = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        quotes = query.all()
        pagination_data = None
    
    quotes_data = []
    now = datetime.now()
    for quote in quotes:
        is_effective = quote.valid_from <= now <= quote.valid_to
        quote_data = {
            "id": quote.id,
            "quote_name": quote.quote_name,
            "customer_id": quote.customer_id,
            "customer_name": quote.customer.full_name if quote.customer else "",
            "customer_short_name": quote.customer.short_name if quote.customer else "",
            "quote_type": quote.quote_type,
            "valid_from": quote.valid_from.isoformat() if quote.valid_from else None,
            "valid_to": quote.valid_to.isoformat() if quote.valid_to else None,
            "is_effective": is_effective,
            "status": "生效中" if is_effective else "已失效",
            "created_at": quote.created_at.isoformat() if quote.created_at else None,
        }
        
        # 根据类型添加报价明细
        if quote.quote_type == "单号报价":
            quote_data["unit_fee"] = float(quote.unit_fee) if quote.unit_fee else 0
        elif quote.quote_type == "头程报价":
            quote_data["air_freight"] = float(quote.air_freight) if quote.air_freight else 0
            quote_data["product_ids"] = [int(x) for x in quote.product_ids.split(",")] if quote.product_ids else []
            # 获取产品名称列表
            if quote.product_ids:
                p_ids = quote.product_ids.split(",")
                p_names = [p.name for p in Product.query.filter(Product.id.in_(p_ids)).all()]
                quote_data["product_names"] = ",".join(p_names)
            else:
                quote_data["product_names"] = ""
        elif quote.quote_type == "尾程报价":
            quote_data["express_fee"] = float(quote.express_fee) if quote.express_fee else 0
            quote_data["registration_fee"] = float(quote.registration_fee) if quote.registration_fee else 0
            quote_data["product_ids"] = [int(x) for x in quote.product_ids.split(",")] if quote.product_ids else []
            # 获取产品名称列表
            if quote.product_ids:
                p_ids = quote.product_ids.split(",")
                p_names = [p.name for p in Product.query.filter(Product.id.in_(p_ids)).all()]
                quote_data["product_names"] = ",".join(p_names)
            else:
                quote_data["product_names"] = ""
        elif quote.quote_type == "专线处理费":
            quote_data["dedicated_line_weight_fee"] = float(quote.dedicated_line_weight_fee) if quote.dedicated_line_weight_fee else 0
            quote_data["dedicated_line_piece_fee"] = float(quote.dedicated_line_piece_fee) if quote.dedicated_line_piece_fee else 0
            quote_data["product_ids"] = [int(x) for x in quote.product_ids.split(",")] if quote.product_ids else []
            # 获取产品名称列表
            if quote.product_ids:
                p_ids = quote.product_ids.split(",")
                p_names = [p.name for p in Product.query.filter(Product.id.in_(p_ids)).all()]
                quote_data["product_names"] = ",".join(p_names)
            else:
                quote_data["product_names"] = ""
        
        quotes_data.append(quote_data)

    return jsonify({
        "success": True, 
        "quotes": quotes_data,
        "pagination": pagination_data
    })


@app.post("/api/customer-quotes")
def api_create_customer_quote():
    """创建客户报价"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建报价"}), 403

    data = request.get_json() or {}
    quote_name = (data.get("quote_name") or "").strip()
    customer_id = data.get("customer_id")
    quote_type = data.get("quote_type", "")
    valid_from = data.get("valid_from", "")
    valid_to = data.get("valid_to", "")

    # 验证必填字段
    if not quote_name:
        return jsonify({"success": False, "message": "报价名称不能为空", "field": "quote_name"}), 400

    # 检查报价名称是否已存在
    existing = CustomerQuote.query.filter_by(quote_name=quote_name).first()
    if existing:
        return jsonify({"success": False, "message": "报价名称已存在，请使用其他名称", "field": "quote_name"}), 400

    if not customer_id:
        return jsonify({"success": False, "message": "请选择报价客户", "field": "customer_id"}), 400

    if not quote_type:
        return jsonify({"success": False, "message": "请选择报价类别", "field": "quote_type"}), 400

    if not valid_from or not valid_to:
        return jsonify({"success": False, "message": "请选择有效期", "field": "valid_from"}), 400

    # 验证客户类别与报价类别是否匹配
    customer = Customer.query.get(customer_id)
    if not customer:
        return jsonify({"success": False, "message": "客户不存在"}), 404

    customer_types = customer.customer_types.split(",") if customer.customer_types else []
    type_mapping = {
        "单号报价": "单号客户",
        "头程报价": "头程客户",
        "尾程报价": "尾程客户",
        "专线处理费": "差价客户"
    }
    
    required_customer_type = type_mapping.get(quote_type)
    if required_customer_type not in customer_types:
        return jsonify({
            "success": False, 
            "message": f"该客户不是{required_customer_type}，无法创建{quote_type}",
            "field": "quote_type"
        }), 400

    # 解析日期（直接使用本地时间，不进行时区转换）
    try:
        # 移除时区信息，直接使用提供的日期时间
        valid_from_str = valid_from.replace("Z", "").split("+")[0].split(".")[0]
        valid_to_str = valid_to.replace("Z", "").split("+")[0].split(".")[0]
        valid_from_dt = datetime.strptime(valid_from_str, "%Y-%m-%dT%H:%M:%S")
        valid_to_dt = datetime.strptime(valid_to_str, "%Y-%m-%dT%H:%M:%S")
    except:
        return jsonify({"success": False, "message": "日期格式错误"}), 400

    # 检查同一客户、同一类别、同一时期是否已有报价
    conflict_query = CustomerQuote.query.filter(
        CustomerQuote.customer_id == customer_id,
        CustomerQuote.quote_type == quote_type,
        db.or_(
            db.and_(
                CustomerQuote.valid_from <= valid_from_dt,
                CustomerQuote.valid_to >= valid_from_dt
            ),
            db.and_(
                CustomerQuote.valid_from <= valid_to_dt,
                CustomerQuote.valid_to >= valid_to_dt
            ),
            db.and_(
                CustomerQuote.valid_from >= valid_from_dt,
                CustomerQuote.valid_to <= valid_to_dt
            )
        )
    )
    
    product_ids_list = data.get("product_ids") or []
    product_ids_str = ",".join(map(str, sorted(product_ids_list))) if product_ids_list else None

    # 需要检查产品ID冲突的类型
    if quote_type in ["头程报价", "尾程报价", "专线处理费"]:
        if not product_ids_list:
            return jsonify({"success": False, "message": "请选择产品", "field": "product_ids"}), 400
        
        # 检查重叠
        potential_conflicts = conflict_query.all()
        for pc in potential_conflicts:
            if pc.product_ids:
                pc_list = pc.product_ids.split(",")
                overlap = set(map(str, product_ids_list)) & set(pc_list)
                if overlap:
                    # 将overlap中的ID转为整数列表
                    overlap_ids = [int(pid) for pid in overlap]
                    conflicting_product_names = [p.name for p in Product.query.filter(Product.id.in_(overlap_ids)).all()]
                    return jsonify({
                        "success": False, 
                        "message": f"该客户在此时期内已有包含产品({', '.join(conflicting_product_names)})的{quote_type}"
                    }), 400
    else:
        # 对于单号报价，直接检查时间冲突
        if conflict_query.first():
            return jsonify({
                "success": False,
                "message": f"该客户在此时期内已有{quote_type}"
            }), 400
    
    # 获取报价明细
    quote = CustomerQuote(
        quote_name=quote_name,
        customer_id=customer_id,
        quote_type=quote_type,
        product_ids=product_ids_str,
        valid_from=valid_from_dt,
        valid_to=valid_to_dt
    )

    # 根据类型设置报价明细
    if quote_type == "单号报价":
        unit_fee = data.get("unit_fee")
        if unit_fee is None or unit_fee == "":
            return jsonify({"success": False, "message": "请输入单号费", "field": "unit_fee"}), 400
        quote.unit_fee = unit_fee
    
    elif quote_type == "头程报价":
        air_freight = data.get("air_freight")
        if air_freight is None or air_freight == "":
            return jsonify({"success": False, "message": "请输入空运费", "field": "air_freight"}), 400
        quote.air_freight = air_freight
    
    elif quote_type == "尾程报价":
        express_fee = data.get("express_fee")
        registration_fee = data.get("registration_fee")
        if express_fee is None or express_fee == "":
            return jsonify({"success": False, "message": "请输入快递费", "field": "express_fee"}), 400
        if registration_fee is None or registration_fee == "":
            return jsonify({"success": False, "message": "请输入挂号费", "field": "registration_fee"}), 400
        quote.express_fee = express_fee
        quote.registration_fee = registration_fee

    elif quote_type == "专线处理费":
        w_fee = data.get("dedicated_line_weight_fee")
        p_fee = data.get("dedicated_line_piece_fee")
        
        # 验证：不可同时为0
        try:
            w_val = float(w_fee) if w_fee else 0
            p_val = float(p_fee) if p_fee else 0
            if w_val == 0 and p_val == 0:
                return jsonify({"success": False, "message": "专线处理费的重量收费和单件收费不能同时为0", "field": "dedicated_line_weight_fee"}), 400
        except ValueError:
            return jsonify({"success": False, "message": "费用格式错误"}), 400
            
        quote.dedicated_line_weight_fee = w_fee
        quote.dedicated_line_piece_fee = p_fee

    db.session.add(quote)
    db.session.commit()

    return jsonify({"success": True, "id": quote.id})


@app.put("/api/customer-quotes/<int:quote_id>")
def api_update_customer_quote(quote_id):
    """更新客户报价"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑报价"}), 403

    quote = CustomerQuote.query.get(quote_id)
    if not quote:
        return jsonify({"success": False, "message": "报价不存在"}), 404

    data = request.get_json() or {}
    quote_name = (data.get("quote_name") or "").strip()
    customer_id = data.get("customer_id")
    quote_type = data.get("quote_type", "")
    valid_from = data.get("valid_from", "")
    valid_to = data.get("valid_to", "")

    # 验证
    if not quote_name:
        return jsonify({"success": False, "message": "报价名称不能为空", "field": "quote_name"}), 400

    # 检查名称是否被其他报价使用
    existing = CustomerQuote.query.filter(
        CustomerQuote.quote_name == quote_name,
        CustomerQuote.id != quote_id
    ).first()
    if existing:
        return jsonify({"success": False, "message": "报价名称已存在", "field": "quote_name"}), 400

    if not customer_id:
        return jsonify({"success": False, "message": "请选择报价客户", "field": "customer_id"}), 400

    if not quote_type:
        return jsonify({"success": False, "message": "请选择报价类别", "field": "quote_type"}), 400

    # 验证客户类别
    customer = Customer.query.get(customer_id)
    if not customer:
        return jsonify({"success": False, "message": "客户不存在"}), 404

    customer_types = customer.customer_types.split(",") if customer.customer_types else []
    type_mapping = {
        "单号报价": "单号客户",
        "头程报价": "头程客户",
        "尾程报价": "尾程客户",
        "专线处理费": "差价客户"
    }
    
    required_customer_type = type_mapping.get(quote_type)
    if required_customer_type not in customer_types:
        return jsonify({
            "success": False,
            "message": f"该客户不是{required_customer_type}，无法创建{quote_type}",
            "field": "quote_type"
        }), 400

    # 解析日期（直接使用本地时间，不进行时区转换）
    try:
        # 移除时区信息，直接使用提供的日期时间
        valid_from_str = valid_from.replace("Z", "").split("+")[0].split(".")[0]
        valid_to_str = valid_to.replace("Z", "").split("+")[0].split(".")[0]
        valid_from_dt = datetime.strptime(valid_from_str, "%Y-%m-%dT%H:%M:%S")
        valid_to_dt = datetime.strptime(valid_to_str, "%Y-%m-%dT%H:%M:%S")
    except:
        return jsonify({"success": False, "message": "日期格式错误"}), 400

    # 检查时间冲突
    conflict_query = CustomerQuote.query.filter(
        CustomerQuote.customer_id == customer_id,
        CustomerQuote.quote_type == quote_type,
        CustomerQuote.id != quote_id,
        db.or_(
            db.and_(
                CustomerQuote.valid_from <= valid_from_dt,
                CustomerQuote.valid_to >= valid_from_dt
            ),
            db.and_(
                CustomerQuote.valid_from <= valid_to_dt,
                CustomerQuote.valid_to >= valid_to_dt
            ),
            db.and_(
                CustomerQuote.valid_from >= valid_from_dt,
                CustomerQuote.valid_to <= valid_to_dt
            )
        )
    )
    
    product_ids_list = data.get("product_ids") or []
    product_ids_str = ",".join(map(str, sorted(product_ids_list))) if product_ids_list else None

    # 需要检查产品ID冲突的类型
    if quote_type in ["头程报价", "尾程报价", "专线处理费"]:
        if not product_ids_list:
            return jsonify({"success": False, "message": "请选择产品", "field": "product_ids"}), 400
        
        # 检查重叠
        potential_conflicts = conflict_query.all()
        for pc in potential_conflicts:
            if pc.product_ids:
                pc_list = pc.product_ids.split(",")
                overlap = set(map(str, product_ids_list)) & set(pc_list)
                if overlap:
                    # 将overlap中的ID转为整数列表
                    overlap_ids = [int(pid) for pid in overlap]
                    conflicting_product_names = [p.name for p in Product.query.filter(Product.id.in_(overlap_ids)).all()]
                    return jsonify({
                        "success": False, 
                        "message": f"该客户在此时期内已有包含产品({', '.join(conflicting_product_names)})的{quote_type}"
                    }), 400
    else:
        # 对于单号报价，直接检查时间冲突
        if conflict_query.first():
            return jsonify({
                "success": False,
                "message": f"该客户在此时期内已有{quote_type}"
            }), 400

    # 更新基本信息
    quote.quote_name = quote_name
    quote.customer_id = customer_id
    quote.quote_type = quote_type
    quote.product_ids = product_ids_str
    quote.valid_from = valid_from_dt
    quote.valid_to = valid_to_dt

    # 清空所有报价明细
    quote.unit_fee = None
    quote.air_freight = None
    quote.express_fee = None
    quote.registration_fee = None
    quote.dedicated_line_weight_fee = None
    quote.dedicated_line_piece_fee = None

    # 根据类型设置报价明细
    if quote_type == "单号报价":
        unit_fee = data.get("unit_fee")
        if unit_fee is None or unit_fee == "":
            return jsonify({"success": False, "message": "请输入单号费", "field": "unit_fee"}), 400
        quote.unit_fee = unit_fee
    
    elif quote_type == "头程报价":
        air_freight = data.get("air_freight")
        if air_freight is None or air_freight == "":
            return jsonify({"success": False, "message": "请输入空运费", "field": "air_freight"}), 400
        quote.air_freight = air_freight
    
    elif quote_type == "尾程报价":
        express_fee = data.get("express_fee")
        registration_fee = data.get("registration_fee")
        if express_fee is None or express_fee == "":
            return jsonify({"success": False, "message": "请输入快递费", "field": "express_fee"}), 400
        if registration_fee is None or registration_fee == "":
            return jsonify({"success": False, "message": "请输入挂号费", "field": "registration_fee"}), 400
        quote.express_fee = express_fee
        quote.registration_fee = registration_fee

    elif quote_type == "专线处理费":
        w_fee = data.get("dedicated_line_weight_fee")
        p_fee = data.get("dedicated_line_piece_fee")
        
        # 验证：不可同时为0
        try:
            w_val = float(w_fee) if w_fee else 0
            p_val = float(p_fee) if p_fee else 0
            if w_val == 0 and p_val == 0:
                return jsonify({"success": False, "message": "专线处理费的重量收费和单件收费不能同时为0", "field": "dedicated_line_weight_fee"}), 400
        except ValueError:
            return jsonify({"success": False, "message": "费用格式错误"}), 400
            
        quote.dedicated_line_weight_fee = w_fee
        quote.dedicated_line_piece_fee = p_fee

    db.session.commit()
    return jsonify({"success": True})


@app.delete("/api/customer-quotes/<int:quote_id>")
def api_delete_customer_quote(quote_id):
    """删除客户报价"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除报价"}), 403

    quote = CustomerQuote.query.get(quote_id)
    if not quote:
        return jsonify({"success": False, "message": "报价不存在"}), 404

    db.session.delete(quote)
    db.session.commit()
    return jsonify({"success": True})


# ==================== 供应商报价管理 API ====================

@app.get("/api/supplier-quotes")
def api_get_supplier_quotes():
    """获取供应商报价列表（支持搜索）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    # 获取搜索参数
    supplier_id = request.args.get("supplier_id", "")
    product_id = request.args.get("product_id", "")
    valid_date = request.args.get("valid_date", "")
    status = request.args.get("status", "")  # 生效中/已失效
    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", type=int, default=20)

    query = SupplierQuote.query
    now = datetime.now()

    if supplier_id:
        query = query.filter(SupplierQuote.supplier_id == supplier_id)
    
    if product_id:
        query = query.filter(SupplierQuote.product_id == product_id)
    
    if valid_date:
        try:
            search_date = datetime.strptime(valid_date, "%Y-%m-%d")
            query = query.filter(
                SupplierQuote.valid_from <= search_date,
                SupplierQuote.valid_to >= search_date
            )
        except:
            pass

    if status == "生效中":
        query = query.filter(SupplierQuote.valid_from <= now, SupplierQuote.valid_to >= now)
    elif status == "已失效":
        query = query.filter((SupplierQuote.valid_from > now) | (SupplierQuote.valid_to < now))

    # 按有效期起始时间降序排列
    query = query.order_by(SupplierQuote.valid_from.desc())

    if page:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        quotes = pagination.items
        pagination_data = {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    else:
        quotes = query.all()
        pagination_data = None
    
    quotes_data = []
    now = datetime.now()
    for quote in quotes:
        is_effective = quote.valid_from <= now <= quote.valid_to
        quotes_data.append({
            "id": quote.id,
            "quote_name": quote.quote_name,
            "supplier_id": quote.supplier_id,
            "supplier_name": quote.supplier.full_name if quote.supplier else "",
            "supplier_short_name": quote.supplier.short_name if quote.supplier else "",
            "product_id": quote.product_id,
            "product_name": quote.product.name if quote.product else "",
            "express_fee": float(quote.express_fee) if quote.express_fee else 0,
            "registration_fee": float(quote.registration_fee) if quote.registration_fee else 0,
            "min_weight": float(quote.min_weight) if quote.min_weight else 0,
            "price_tiers": json.loads(quote.price_tiers) if quote.price_tiers else [],
            "valid_from": quote.valid_from.isoformat() if quote.valid_from else None,
            "valid_to": quote.valid_to.isoformat() if quote.valid_to else None,
            "is_effective": is_effective,
            "status": "生效中" if is_effective else "已失效",
            "created_at": quote.created_at.isoformat() if quote.created_at else None,
        })

    return jsonify({
        "success": True, 
        "quotes": quotes_data,
        "pagination": pagination_data
    })


@app.post("/api/supplier-quotes")
def api_create_supplier_quote():
    """创建供应商报价"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限创建报价"}), 403

    data = request.get_json() or {}
    quote_name = (data.get("quote_name") or "").strip()
    supplier_id = data.get("supplier_id")
    product_id = data.get("product_id")
    min_weight = data.get("min_weight", 0)
    price_tiers = data.get("price_tiers", [])
    valid_from = data.get("valid_from", "")
    valid_to = data.get("valid_to", "")

    # 验证必填字段
    if not quote_name:
        return jsonify({"success": False, "message": "报价名称不能为空", "field": "quote_name"}), 400

    # 检查报价名称是否已存在
    existing = SupplierQuote.query.filter_by(quote_name=quote_name).first()
    if existing:
        return jsonify({"success": False, "message": "报价名称已存在，请使用其他名称", "field": "quote_name"}), 400

    if not supplier_id:
        return jsonify({"success": False, "message": "请选择供应商", "field": "supplier_id"}), 400

    if not product_id:
        return jsonify({"success": False, "message": "请选择产品", "field": "product_id"}), 400

    if not price_tiers or len(price_tiers) == 0:
        return jsonify({"success": False, "message": "请至少添加一个价格阶梯", "field": "price_tiers"}), 400

    if not valid_from or not valid_to:
        return jsonify({"success": False, "message": "请选择有效期", "field": "valid_from"}), 400

    # 解析日期
    try:
        valid_from_str = valid_from.replace("Z", "").split("+")[0].split(".")[0]
        valid_to_str = valid_to.replace("Z", "").split("+")[0].split(".")[0]
        valid_from_dt = datetime.strptime(valid_from_str, "%Y-%m-%dT%H:%M:%S")
        valid_to_dt = datetime.strptime(valid_to_str, "%Y-%m-%dT%H:%M:%S")
    except:
        return jsonify({"success": False, "message": "日期格式错误"}), 400

    # 检查同一供应商、同一产品、同一时期是否已有报价
    conflict = SupplierQuote.query.filter(
        SupplierQuote.supplier_id == supplier_id,
        SupplierQuote.product_id == product_id,
        db.or_(
            db.and_(
                SupplierQuote.valid_from <= valid_from_dt,
                SupplierQuote.valid_to >= valid_from_dt
            ),
            db.and_(
                SupplierQuote.valid_from <= valid_to_dt,
                SupplierQuote.valid_to >= valid_to_dt
            ),
            db.and_(
                SupplierQuote.valid_from >= valid_from_dt,
                SupplierQuote.valid_to <= valid_to_dt
            )
        )
    ).first()
    
    if conflict:
        return jsonify({
            "success": False,
            "message": "该供应商在此时期内已有相同产品的报价"
        }), 400

    # 创建报价
    quote = SupplierQuote(
        quote_name=quote_name,
        supplier_id=supplier_id,
        product_id=product_id,
        min_weight=min_weight,
        price_tiers=json.dumps(price_tiers),
        # 兼容旧字段，取第一个阶梯的值
        express_fee=price_tiers[0].get('express', 0) if price_tiers else 0,
        registration_fee=price_tiers[0].get('reg', 0) if price_tiers else 0,
        valid_from=valid_from_dt,
        valid_to=valid_to_dt
    )

    db.session.add(quote)
    db.session.commit()

    return jsonify({"success": True, "id": quote.id})


@app.put("/api/supplier-quotes/<int:quote_id>")
def api_update_supplier_quote(quote_id):
    """更新供应商报价"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限编辑报价"}), 403

    quote = SupplierQuote.query.get(quote_id)
    if not quote:
        return jsonify({"success": False, "message": "报价不存在"}), 404

    data = request.get_json() or {}
    quote_name = (data.get("quote_name") or "").strip()
    supplier_id = data.get("supplier_id")
    product_id = data.get("product_id")
    min_weight = data.get("min_weight", 0)
    price_tiers = data.get("price_tiers", [])
    valid_from = data.get("valid_from", "")
    valid_to = data.get("valid_to", "")

    # 验证
    if not quote_name:
        return jsonify({"success": False, "message": "报价名称不能为空", "field": "quote_name"}), 400

    # 检查名称是否被其他报价使用
    existing = SupplierQuote.query.filter(
        SupplierQuote.quote_name == quote_name,
        SupplierQuote.id != quote_id
    ).first()
    if existing:
        return jsonify({"success": False, "message": "报价名称已存在", "field": "quote_name"}), 400

    if not supplier_id:
        return jsonify({"success": False, "message": "请选择供应商", "field": "supplier_id"}), 400

    if not product_id:
        return jsonify({"success": False, "message": "请选择产品", "field": "product_id"}), 400

    if not price_tiers or len(price_tiers) == 0:
        return jsonify({"success": False, "message": "请至少添加一个价格阶梯", "field": "price_tiers"}), 400

    # 解析日期
    try:
        valid_from_str = valid_from.replace("Z", "").split("+")[0].split(".")[0]
        valid_to_str = valid_to.replace("Z", "").split("+")[0].split(".")[0]
        valid_from_dt = datetime.strptime(valid_from_str, "%Y-%m-%dT%H:%M:%S")
        valid_to_dt = datetime.strptime(valid_to_str, "%Y-%m-%dT%H:%M:%S")
    except:
        return jsonify({"success": False, "message": "日期格式错误"}), 400

    # 检查时间冲突
    conflict = SupplierQuote.query.filter(
        SupplierQuote.supplier_id == supplier_id,
        SupplierQuote.product_id == product_id,
        SupplierQuote.id != quote_id,
        db.or_(
            db.and_(
                SupplierQuote.valid_from <= valid_from_dt,
                SupplierQuote.valid_to >= valid_from_dt
            ),
            db.and_(
                SupplierQuote.valid_from <= valid_to_dt,
                SupplierQuote.valid_to >= valid_to_dt
            ),
            db.and_(
                SupplierQuote.valid_from >= valid_from_dt,
                SupplierQuote.valid_to <= valid_to_dt
            )
        )
    ).first()
    
    if conflict:
        return jsonify({
            "success": False,
            "message": "该供应商在此时期内已有相同产品的报价"
        }), 400

    # 更新
    quote.quote_name = quote_name
    quote.supplier_id = supplier_id
    quote.product_id = product_id
    quote.min_weight = min_weight
    quote.price_tiers = json.dumps(price_tiers)
    # 兼容旧字段
    quote.express_fee = price_tiers[0].get('express', 0) if price_tiers else 0
    quote.registration_fee = price_tiers[0].get('reg', 0) if price_tiers else 0
    quote.valid_from = valid_from_dt
    quote.valid_to = valid_to_dt

    db.session.commit()
    return jsonify({"success": True})


@app.delete("/api/supplier-quotes/<int:quote_id>")
def api_delete_supplier_quote(quote_id):
    """删除供应商报价"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除报价"}), 403

    quote = SupplierQuote.query.get(quote_id)
    if not quote:
        return jsonify({"success": False, "message": "报价不存在"}), 404

    db.session.delete(quote)
    db.session.commit()
    return jsonify({"success": True})


# ==================== 运单数据管理 API ====================

def apply_waybill_filters(query, data):
    """通用运单过滤逻辑，支持单号/转单号多行搜索"""
    customer_id = data.get("customer_id", "")
    supplier_id = data.get("supplier_id", "")
    product_id = data.get("product_id", "")
    order_time_start = data.get("order_time_start", "")
    order_time_end = data.get("order_time_end", "")
    order_nos_str = data.get("order_nos", "")
    transfer_nos_str = data.get("transfer_nos", "")

    # 客户搜索（四个客户字段中任意一个匹配）
    if customer_id:
        query = query.filter(
            db.or_(
                Waybill.unit_customer_id == customer_id,
                Waybill.first_leg_customer_id == customer_id,
                Waybill.last_leg_customer_id == customer_id,
                Waybill.differential_customer_id == customer_id
            )
        )
    
    if supplier_id:
        query = query.filter(Waybill.supplier_id == supplier_id)
    
    if product_id:
        query = query.filter(Waybill.product_id == product_id)
    
    # 下单时间范围搜索
    if order_time_start:
        try:
            start_dt = datetime.strptime(order_time_start, "%Y-%m-%d")
            query = query.filter(Waybill.order_time >= start_dt)
        except:
            pass
    
    if order_time_end:
        try:
            end_dt = datetime.strptime(order_time_end, "%Y-%m-%d")
            # 结束日期包含当天整天
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Waybill.order_time <= end_dt)
        except:
            pass

    # 订单号搜索（支持多行输入）
    if order_nos_str:
        order_nos = [x.strip() for x in order_nos_str.split('\n') if x.strip()]
        if order_nos:
            # 分片处理 IN 查询，避免 SQL 语句过长
            chunk_size = 1000
            chunks = [order_nos[i:i + chunk_size] for i in range(0, len(order_nos), chunk_size)]
            query = query.filter(db.or_(*[Waybill.order_no.in_(chunk) for chunk in chunks]))
            
    # 转单号搜索（支持多行输入）
    if transfer_nos_str:
        transfer_nos = [x.strip() for x in transfer_nos_str.split('\n') if x.strip()]
        if transfer_nos:
            # 分片处理 IN 查询，避免 SQL 语句过长
            chunk_size = 1000
            chunks = [transfer_nos[i:i + chunk_size] for i in range(0, len(transfer_nos), chunk_size)]
            query = query.filter(db.or_(*[Waybill.transfer_no.in_(chunk) for chunk in chunks]))
            
    return query

@app.route("/api/waybills", methods=["GET", "POST"])
def api_get_waybills():
    """获取运单列表（支持搜索和分页）"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    # 获取参数（优先从 JSON POST 中获取）
    if request.method == "POST":
        data = request.get_json() or {}
    else:
        data = request.args
    
    # 分页参数
    page = int(data.get("page", 1))
    page_size = int(data.get("page_size", 200))

    # 检查是否有搜索过滤条件 (排除分页参数)
    filter_keys = ["customer_id", "supplier_id", "product_id", "order_time_start", "order_time_end", "order_nos", "transfer_nos"]
    has_filter = any([data.get(k) for k in filter_keys])
    
    query = apply_waybill_filters(Waybill.query, data)

    if not has_filter:
        # 如果没有任何过滤条件，默认限制在最近的5万条数据内，以保证大表下的性能
        # 先统计总数，最大显示为50000
        full_count = query.count()
        total = min(full_count, 50000)
        # 只取最近的5万条进行分页
        query = query.order_by(Waybill.order_time.desc()).limit(50000)
    else:
        # 有搜索条件时，统计搜索结果的全量总数
        total = query.count()
        query = query.order_by(Waybill.order_time.desc())

    # 分页执行
    if not has_filter and full_count > 50000:
        # 对于超过50000条且无过滤条件的情况，直接使用limit和offset
        waybills = query.limit(page_size).offset((page - 1) * page_size).all()
    else:
        waybills = query.limit(page_size).offset((page - 1) * page_size).all()
    
    waybills_data = []
    for waybill in waybills:
        waybills_data.append({
            "id": waybill.id,
            "order_no": waybill.order_no,
            "transfer_no": waybill.transfer_no or "",
            "weight": float(waybill.weight) if waybill.weight else 0,
            "order_time": waybill.order_time.isoformat() if waybill.order_time else None,
            "product_id": waybill.product_id,
            "product_name": waybill.product.name if waybill.product else "",
            
            # 客户信息
            "unit_customer_id": waybill.unit_customer_id,
            "unit_customer_name": waybill.unit_customer.short_name if waybill.unit_customer else "",
            "first_leg_customer_id": waybill.first_leg_customer_id,
            "first_leg_customer_name": waybill.first_leg_customer.short_name if waybill.first_leg_customer else "",
            "last_leg_customer_id": waybill.last_leg_customer_id,
            "last_leg_customer_name": waybill.last_leg_customer.short_name if waybill.last_leg_customer else "",
            "differential_customer_id": waybill.differential_customer_id,
            "differential_customer_name": waybill.differential_customer.short_name if waybill.differential_customer else "",
            
            "supplier_id": waybill.supplier_id,
            "supplier_name": waybill.supplier.short_name if waybill.supplier else "",
            
            # 费用信息
            "unit_fee": float(waybill.unit_fee) if waybill.unit_fee else 0,
            "first_leg_fee": float(waybill.first_leg_fee) if waybill.first_leg_fee else 0,
            "last_leg_fee": float(waybill.last_leg_fee) if waybill.last_leg_fee else 0,
            "differential_fee": float(waybill.differential_fee) if waybill.differential_fee else 0,
            "dedicated_line_fee": float(waybill.dedicated_line_fee) if waybill.dedicated_line_fee else 0,
            "supplier_cost": float(waybill.supplier_cost) if waybill.supplier_cost else 0,
            "other_fee": float(waybill.other_fee) if waybill.other_fee else 0,
            
            "remark": waybill.remark or "",
        })

    return jsonify({
        "success": True,
        "waybills": waybills_data,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    })


@app.route("/api/waybills/recalculate", methods=["POST"])
def api_recalculate_waybills():
    """重算查询出来的运单费用"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限进行此操作"}), 403

    data = request.get_json() or {}
    
    # 1. 查找所有符合条件的运单
    query = Waybill.query
    query = apply_waybill_filters(query, data)
    
    waybills = query.all()
    if not waybills:
        return jsonify({"success": True, "message": "没有找到符合条件的运单", "count": 0})

    # 2. 预加载所有报价和基础数据并建立索引字典（大幅提升查询性能）
    products_map = {p.id: p for p in Product.query.all()}
    customer_quotes = CustomerQuote.query.all()
    customer_quotes_idx = {}
    for q in customer_quotes:
        key = (q.customer_id, q.quote_type)
        if key not in customer_quotes_idx:
            customer_quotes_idx[key] = []
        customer_quotes_idx[key].append(q)

    supplier_quotes = SupplierQuote.query.all()
    supplier_quotes_idx = {}
    for q in supplier_quotes:
        # 预解析阶梯报价 JSON
        if q.price_tiers and isinstance(q.price_tiers, str):
            try: q.parsed_tiers = json.loads(q.price_tiers)
            except: q.parsed_tiers = []
        else:
            q.parsed_tiers = []
            
        key = (q.supplier_id, q.product_id)
        if key not in supplier_quotes_idx:
            supplier_quotes_idx[key] = []
        supplier_quotes_idx[key].append(q)
    
    updated_count = 0
    errors = []
    
    try:
        # 3. 循环计算并更新
        for wb in waybills:
            fees, row_errors = calculate_waybill_fees(wb, products_map, customer_quotes_idx, supplier_quotes_idx)
            
            if row_errors:
                errors.append(f"单号 {wb.order_no}: {'; '.join(row_errors)}")
                continue # 如果计算出错，跳过该单更新，保持原有数据
            
            if fees:
                wb.unit_fee = fees["unit_fee"]
                wb.first_leg_fee = fees["first_leg_fee"]
                wb.last_leg_fee = fees["last_leg_fee"]
                wb.differential_fee = fees["differential_fee"]
                wb.dedicated_line_fee = fees["dedicated_line_fee"]
                wb.supplier_cost = fees["supplier_cost"]
                updated_count += 1
        
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": f"重算完成。成功更新 {updated_count} 条运单。",
            "total_queried": len(waybills),
            "error_count": len(errors),
            "errors": errors[:100] # 只返回前100个错误，避免数据量过大
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"重算失败：{str(e)}"}), 500


@app.route("/api/waybills/export", methods=["GET", "POST"])
def api_export_waybills():
    """导出运单列表为 Excel"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    # 获取参数
    if request.method == "POST":
        data = request.get_json() or {}
    else:
        data = request.args
    
    query = Waybill.query
    query = apply_waybill_filters(query, data)

    # 按下单时间降序
    waybills = query.order_by(Waybill.order_time.desc()).all()

    # 构造数据
    data = []
    for idx, w in enumerate(waybills):
        data.append({
            "序号": idx + 1,
            "订单号": w.order_no,
            "转单号": w.transfer_no or "",
            "重量(kg)": float(w.weight) if w.weight else 0,
            "下单时间": w.order_time.strftime("%Y-%m-%d %H:%M:%S") if w.order_time else "",
            "产品": w.product.name if w.product else "",
            "单号客户": w.unit_customer.short_name if w.unit_customer else "",
            "头程客户": w.first_leg_customer.short_name if w.first_leg_customer else "",
            "尾程客户": w.last_leg_customer.short_name if w.last_leg_customer else "",
            "差价客户": w.differential_customer.short_name if w.differential_customer else "",
            "供应商": w.supplier.short_name if w.supplier else "",
            "单号收费": float(w.unit_fee) if w.unit_fee else 0,
            "头程收费": float(w.first_leg_fee) if w.first_leg_fee else 0,
            "尾程收费": float(w.last_leg_fee) if w.last_leg_fee else 0,
            "差价收费": float(w.differential_fee) if w.differential_fee else 0,
            "供应商成本": float(w.supplier_cost) if w.supplier_cost else 0,
            "专线处理费": float(w.dedicated_line_fee) if w.dedicated_line_fee else 0,
            "其他费用": float(w.other_fee) if w.other_fee else 0,
            "备注": w.remark or ""
        })

    # 生成 Excel
    df = pd.DataFrame(data)
    output = io.BytesIO()
    # 使用 openpyxl 引擎
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='运单数据')
        
        # 获取工作表以设置列宽
        workbook = writer.book
        worksheet = writer.sheets['运单数据']
        
        # 简单设置列宽
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[chr(65 + i)].width = min(column_len, 50)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='waybills_export.xlsx'
    )


@app.get("/api/waybills/download-template")
def api_download_waybill_template():
    """下载运单导入模板"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "运单数据"
    
    # 设置表头
    headers = ["订单号", "转单号", "重量(kg)", "下单时间", "产品", 
                "单号客户", "头程客户", "尾程客户", "差价客户"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20  # 订单号
    ws.column_dimensions['B'].width = 20  # 转单号
    ws.column_dimensions['C'].width = 12  # 重量
    ws.column_dimensions['D'].width = 20  # 下单时间
    ws.column_dimensions['E'].width = 15  # 产品
    ws.column_dimensions['F'].width = 15  # 单号客户
    ws.column_dimensions['G'].width = 15  # 头程客户
    ws.column_dimensions['H'].width = 15  # 尾程客户
    ws.column_dimensions['I'].width = 15  # 差价客户
    
    # 添加示例数据（第2行）
    example_data = [
        "ORD20260115001",  # 订单号
        "TRN001",  # 转单号
        "12.500",  # 重量
        "2026-1-5 12:15:00",  # 下单时间
        "示例产品",  # 产品
        "客户A简称",  # 单号客户
        "客户B简称",  # 头程客户
        "客户C简称",  # 尾程客户
        ""  # 差价客户
    ]
    ws.append(example_data)
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'运单导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.post("/api/waybills/import")
def api_import_waybills():
    """导入运单数据"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限导入"}), 403
    
    # 检查是否有上传文件
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "请选择文件"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"success": False, "message": "请选择文件"}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "请上传Excel文件（.xlsx或.xls）"}), 400
    
    try:
        # 保存上传文件
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"waybill_import_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # 准备模型字典
        models = {
            'Waybill': Waybill,
            'Product': Product,
            'Customer': Customer,
            'Supplier': Supplier,
            'CustomerQuote': CustomerQuote,
            'SupplierQuote': SupplierQuote,
            'TrackingInfo': TrackingInfo  # 添加TrackingInfo模型
        }
        
        # 处理导入
        success, message, error_details = validate_and_process_waybill_import(
            file_path, db, models
        )
        
        # 删除上传的文件
        try:
            os.remove(file_path)
        except:
            pass
        
        if success:
            return jsonify({
                "success": True,
                "message": message
            })
        else:
            return jsonify({
                "success": False,
                "message": message,
                "errors": error_details
            }), 400
    
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"导入失败：{str(e)}"
        }), 500


@app.post("/api/waybills/batch-delete")
def api_batch_delete_waybills():
    """批量删除运单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401
    
    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除"}), 403
    
    try:
        data = request.get_json()
        ids = data.get("ids", [])
        
        if not ids:
            return jsonify({"success": False, "message": "请选择要删除的运单"}), 400
        
        # 查询要删除的运单
        waybills = Waybill.query.filter(Waybill.id.in_(ids)).all()
        
        if not waybills:
            return jsonify({"success": False, "message": "未找到要删除的运单"}), 404
        
        # 批量删除
        for waybill in waybills:
            db.session.delete(waybill)
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"成功删除{len(waybills)}条运单"
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": f"删除失败：{str(e)}"
        }), 500


# ==================== 账单管理 API ====================

# ==================== 供应商账单管理 API ====================

@app.get("/api/supplier-invoices")
def api_get_supplier_invoices():
    """获取供应商账单列表"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    supplier_name = request.args.get("supplier_name", "")
    supplier_id = request.args.get("supplier_id")
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    page = request.args.get("page", type=int, default=1)
    per_page = request.args.get("per_page", type=int, default=20)

    query = SupplierInvoice.query.join(Supplier)

    if supplier_id:
        query = query.filter(SupplierInvoice.supplier_id == supplier_id)
    elif supplier_name:
        query = query.filter(Supplier.short_name.like(f"%{supplier_name}%"))
    if year:
        query = query.filter(SupplierInvoice.year == year)
    if month:
        query = query.filter(SupplierInvoice.month == month)

    pagination = query.order_by(SupplierInvoice.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    invoices_data = []
    for inv in pagination.items:
        invoices_data.append({
            "id": inv.id,
            "supplier_name": inv.supplier.full_name,
            "period": f"{inv.year}-{inv.month:02d}",
            "amount": float(inv.amount),
            "file_name": inv.file_name,
            "is_paid": inv.is_paid,
            "created_at": inv.created_at.isoformat()
        })

    return jsonify({
        "success": True,
        "invoices": invoices_data,
        "pagination": {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    })


@app.post("/api/supplier-invoices/generate")
def api_generate_supplier_invoices():
    """批量异步生成供应商账单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    data = request.get_json() or {}
    year = data.get("year")
    month = data.get("month")

    if not year or not month:
        return jsonify({"success": False, "message": "请选择年份和月份"}), 400

    # 提交异步任务
    task = async_generate_supplier_invoices.delay(year, month)
    
    # 记录任务
    new_task = TaskRecord(
        task_id=task.id,
        task_name=f"生成供应商账单({year}-{month})",
        status="PENDING"
    )
    db.session.add(new_task)
    db.session.commit()

    return jsonify({
        "success": True, 
        "message": "账单生成任务已提交，请稍后查看结果",
        "task_id": task.id
    })


@app.delete("/api/supplier-invoices/<int:invoice_id>")
def api_delete_supplier_invoice(invoice_id):
    """删除供应商账单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除账单"}), 403

    invoice = SupplierInvoice.query.get(invoice_id)
    if not invoice:
        return jsonify({"success": False, "message": "账单不存在"}), 404

    if invoice.file_name:
        file_path = os.path.join(app.config['SUPPLIER_INVOICE_FOLDER'], invoice.file_name)
        if os.path.exists(file_path):
            try: os.remove(file_path)
            except: pass

    db.session.delete(invoice)
    db.session.commit()
    return jsonify({"success": True})


@app.post("/api/supplier-invoices/<int:invoice_id>/recalculate")
def api_recalculate_supplier_invoice(invoice_id):
    """批量异步重新计算供应商账单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    invoice = SupplierInvoice.query.get(invoice_id)
    if not invoice:
        return jsonify({"success": False, "message": "账单不存在"}), 404

    # 提交异步任务
    task = async_generate_supplier_invoices.delay(invoice.year, invoice.month)
    
    # 记录任务
    new_task = TaskRecord(
        task_id=task.id,
        task_name=f"重算供应商账单({invoice.year}-{invoice.month})",
        status="PENDING"
    )
    db.session.add(new_task)
    db.session.commit()

    return jsonify({
        "success": True, 
        "message": "重算任务已提交，请稍后查看结果",
        "task_id": task.id
    })


@app.get("/api/supplier-invoices/<int:invoice_id>/download")
def api_download_supplier_invoice(invoice_id):
    """下载供应商账单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    invoice = SupplierInvoice.query.get(invoice_id)
    if not invoice or not invoice.file_name:
        return jsonify({"success": False, "message": "文件不存在"}), 404

    file_path = os.path.join(app.config['SUPPLIER_INVOICE_FOLDER'], invoice.file_name)
    if not os.path.exists(file_path):
        return jsonify({"success": False, "message": "服务器文件已丢失"}), 404

    return send_file(file_path, as_attachment=True, download_name=invoice.file_name)


@app.get("/api/invoices")
def api_get_invoices():
    """获取账单列表，支持搜索"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    customer_id = request.args.get("customer_id")
    fee_type = request.args.get("fee_type")
    year = request.args.get("year")
    month = request.args.get("month")
    
    # 分页参数
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    query = Invoice.query

    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)
    if fee_type:
        query = query.filter(Invoice.fee_type == fee_type)
    if year:
        query = query.filter(Invoice.year == year)
    if month:
        query = query.filter(Invoice.month == month)

    # 排序：年份降序、月份降序、创建时间降序
    query = query.order_by(Invoice.year.desc(), Invoice.month.desc(), Invoice.created_at.desc())
    
    # 分页查询
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    invoices = pagination.items
    
    data = []
    for inv in invoices:
        data.append({
            "id": inv.id,
            "customer_id": inv.customer_id,
            "customer_name": inv.customer.full_name if inv.customer else "",
            "fee_type": inv.fee_type,
            "year": inv.year,
            "month": inv.month,
            "amount": float(inv.amount),
            "file_name": inv.file_name,
            "is_paid": inv.is_paid,
            "created_at": inv.created_at.isoformat()
        })

    return jsonify({
        "success": True, 
        "invoices": data,
        "total": pagination.total,
        "pages": pagination.pages,
        "current_page": pagination.page
    })


@app.post("/api/invoices/generate")
def api_generate_invoices():
    """批量异步生成应收账单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    data = request.get_json() or {}
    year = data.get("year")
    month = data.get("month")
    customer_id = data.get("customer_id")  # 新增客户ID参数

    if not year or not month:
        return jsonify({"success": False, "message": "请选择年份和月份"}), 400

    # 提交异步任务，传递customer_id参数
    task = async_generate_customer_invoices.delay(year, month, customer_id)
    
    # 记录任务
    task_name = f"生成应收账单({year}-{month})"
    if customer_id:
        customer = Customer.query.get(customer_id)
        if customer:
            task_name = f"生成应收账单({year}-{month})-{customer.short_name}"
    
    new_task = TaskRecord(
        task_id=task.id,
        task_name=task_name,
        status="PENDING"
    )
    db.session.add(new_task)
    db.session.commit()

    return jsonify({
        "success": True, 
        "message": "账单生成任务已提交，请稍后查看结果",
        "task_id": task.id
    })


@app.get("/api/tasks/status/<task_id>")
def api_get_task_status(task_id):
    """查询异步任务执行状态"""
    record = TaskRecord.query.filter_by(task_id=task_id).first()
    if not record:
        return jsonify({"success": False, "message": "任务记录不存在"}), 404
    
    return jsonify({
        "success": True,
        "task_id": record.task_id,
        "status": record.status,
        "result_msg": record.result_msg,
        "updated_at": record.updated_at.isoformat()
    })


@app.delete("/api/invoices/<int:invoice_id>")
def api_delete_invoice(invoice_id):
    """删除账单记录及文件"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除账单"}), 403

    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({"success": False, "message": "账单不存在"}), 404

    # 删除物理文件
    if invoice.file_name:
        file_path = os.path.join(app.config['INVOICE_FOLDER'], invoice.file_name)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

    db.session.delete(invoice)
    db.session.commit()
    return jsonify({"success": True})


@app.post("/api/invoices/<int:invoice_id>/recalculate")
def api_recalculate_invoice(invoice_id):
    """批量异步重新计算应收账单"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({"success": False, "message": "账单不存在"}), 404

    # 提交异步任务
    task = async_generate_customer_invoices.delay(invoice.year, invoice.month)
    
    # 记录任务
    new_task = TaskRecord(
        task_id=task.id,
        task_name=f"重算应收账单({invoice.year}-{invoice.month})",
        status="PENDING"
    )
    db.session.add(new_task)
    db.session.commit()

    return jsonify({
        "success": True, 
        "message": "重算任务已提交，请稍后查看结果",
        "task_id": task.id
    })


@app.get("/api/invoices/<int:invoice_id>/download")
def api_download_invoice(invoice_id):
    """下载账单文件"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    invoice = Invoice.query.get(invoice_id)
    if not invoice or not invoice.file_name:
        return jsonify({"success": False, "message": "文件不存在"}), 404

    file_path = os.path.join(app.config['INVOICE_FOLDER'], invoice.file_name)
    if not os.path.exists(file_path):
        return jsonify({"success": False, "message": "服务器文件已丢失"}), 404

    return send_file(file_path, as_attachment=True, download_name=invoice.file_name)


# ==================== 收付款管理 API ====================

@app.get("/api/payments")
def api_get_payments():
    """获取收付款记录"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    target_type = request.args.get("target_type")
    target_id = request.args.get("target_id")
    payment_type = request.args.get("payment_type")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    query = Payment.query

    # 打印调试信息到控制台（如果能看到的话）
    print(f"DEBUG: Fetching payments with filters - type: {target_type}, id: {target_id}, ptype: {payment_type}, start: {start_date}, end: {end_date}")

    if target_type and target_type.strip():
        query = query.filter(Payment.target_type == target_type.strip())
    if target_id and str(target_id).strip() and str(target_id) != 'undefined' and str(target_id).strip() != '':
        query = query.filter(Payment.target_id == target_id)
    if payment_type:
        query = query.filter(Payment.payment_type == payment_type)
    if start_date and start_date.strip():
        try:
            query = query.filter(Payment.payment_date >= datetime.strptime(start_date, "%Y-%m-%d"))
        except: pass
    if end_date and end_date.strip():
        try:
            query = query.filter(Payment.payment_date <= datetime.strptime(end_date, "%Y-%m-%d"))
        except: pass

    pagination = query.order_by(Payment.payment_date.desc(), Payment.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    print(f"DEBUG: Found {pagination.total} total records, current page items: {len(pagination.items)}")
    
    payments_data = []
    for p in pagination.items:
        linked_invoice = ""
        try:
            if p.target_type == 'customer' and p.invoice_id:
                inv = Invoice.query.get(p.invoice_id)
                if inv and inv.customer:
                    linked_invoice = f"{inv.year}年{inv.month}月-{inv.customer.short_name}-{inv.fee_type}"
            elif p.target_type == 'supplier' and p.supplier_invoice_id:
                sinv = SupplierInvoice.query.get(p.supplier_invoice_id)
                if sinv and sinv.supplier:
                    linked_invoice = f"{sinv.year}年{sinv.month}月-{sinv.supplier.short_name}"
        except Exception as e:
            print(f"Error processing payment invoice link: {e}")
            linked_invoice = "数据关联错误"

        payments_data.append({
            "id": p.id,
            "target_type": p.target_type,
            "target_id": p.target_id,
            "target_name": p.target_name,
            "payment_type": p.payment_type.strip() if p.payment_type else "",
            "payment_date": p.payment_date.strftime("%Y-%m-%d"),
            "amount": float(p.amount),
            "receipt_path": p.receipt_path,
            "invoice_id": p.invoice_id,
            "supplier_invoice_id": p.supplier_invoice_id,
            "linked_invoice": linked_invoice,
            "remark": p.remark,
            "created_at": p.created_at.isoformat()
        })

    return jsonify({
        "success": True,
        "payments": payments_data,
        "pagination": {
            "total": pagination.total,
            "pages": pagination.pages,
            "current_page": pagination.page,
            "per_page": pagination.per_page
        }
    })

@app.post("/api/payments")
def api_create_payment():
    """新增收付款记录"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    data = request.get_json() or {}
    try:
        invoice_id = data.get("invoice_id")
        supplier_invoice_id = data.get("supplier_invoice_id")

        payment = Payment(
            target_type=data.get("target_type"),
            target_id=data.get("target_id"),
            payment_type=data.get("payment_type"),
            payment_date=datetime.strptime(data.get("payment_date"), "%Y-%m-%d"),
            amount=data.get("amount"),
            receipt_path=data.get("receipt_path"),
            invoice_id=invoice_id,
            supplier_invoice_id=supplier_invoice_id,
            remark=data.get("remark")
        )
        
        # 标记账单为已收/已付
        if invoice_id:
            inv = Invoice.query.get(invoice_id)
            if inv: inv.is_paid = True
        if supplier_invoice_id:
            sinv = SupplierInvoice.query.get(supplier_invoice_id)
            if sinv: sinv.is_paid = True

        db.session.add(payment)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@app.put("/api/payments/<int:payment_id>")
def api_update_payment(payment_id):
    """更新收付款记录"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    payment = Payment.query.get(payment_id)
    if not payment:
        return jsonify({"success": False, "message": "记录不存在"}), 404

    data = request.get_json() or {}
    try:
        # 处理旧账单状态恢复
        if payment.invoice_id:
            old_inv = Invoice.query.get(payment.invoice_id)
            if old_inv: old_inv.is_paid = False
        if payment.supplier_invoice_id:
            old_sinv = SupplierInvoice.query.get(payment.supplier_invoice_id)
            if old_sinv: old_sinv.is_paid = False

        payment.target_type = data.get("target_type")
        payment.target_id = data.get("target_id")
        payment.payment_type = data.get("payment_type")
        payment.payment_date = datetime.strptime(data.get("payment_date"), "%Y-%m-%d")
        payment.amount = data.get("amount")
        payment.receipt_path = data.get("receipt_path")
        payment.invoice_id = data.get("invoice_id")
        payment.supplier_invoice_id = data.get("supplier_invoice_id")
        payment.remark = data.get("remark")

        # 标记新账单状态
        if payment.invoice_id:
            new_inv = Invoice.query.get(payment.invoice_id)
            if new_inv: new_inv.is_paid = True
        if payment.supplier_invoice_id:
            new_sinv = SupplierInvoice.query.get(payment.supplier_invoice_id)
            if new_sinv: new_sinv.is_paid = True

        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@app.delete("/api/payments/<int:payment_id>")
def api_delete_payment(payment_id):
    """删除收付款记录"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if current_user.get("role") != "系统管理员":
        return jsonify({"success": False, "message": "无权限删除记录"}), 403

    payment = Payment.query.get(payment_id)
    if not payment:
        return jsonify({"success": False, "message": "记录不存在"}), 404

    try:
        # 恢复账单状态
        if payment.invoice_id:
            inv = Invoice.query.get(payment.invoice_id)
            if inv: inv.is_paid = False
        if payment.supplier_invoice_id:
            sinv = SupplierInvoice.query.get(payment.supplier_invoice_id)
            if sinv: sinv.is_paid = False

        # 同时也尝试删除水单文件
        if payment.receipt_path:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(payment.receipt_path))
            if os.path.exists(file_path):
                try: os.remove(file_path)
                except: pass

        db.session.delete(payment)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@app.get("/api/unpaid-invoices")
def api_get_unpaid_invoices():
    """获取未核销的账单用于收付款绑定"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    target_type = request.args.get("target_type")
    target_id = request.args.get("target_id")
    
    invoices = []
    if target_type == 'customer':
        query = Invoice.query.filter_by(customer_id=target_id, is_paid=False)
        for inv in query.all():
            invoices.append({
                "id": inv.id,
                "label": f"{inv.year}年{inv.month}月-{inv.customer.short_name}-{inv.fee_type} (金额: {inv.amount})"
            })
    elif target_type == 'supplier':
        query = SupplierInvoice.query.filter_by(supplier_id=target_id, is_paid=False)
        for inv in query.all():
            invoices.append({
                "id": inv.id,
                "label": f"{inv.year}年{inv.month}月-{inv.supplier.short_name} (金额: {inv.amount})"
            })
            
    return jsonify({"success": True, "invoices": invoices})

@app.post("/api/payments/upload-receipt")
def api_upload_payment_receipt():
    """上传水单图片"""
    current_user = session.get("user")
    if not current_user:
        return jsonify({"success": False, "message": "未登录"}), 401

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "没有文件"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "文件名为空"}), 400
    
    if file:
        filename = secure_filename(file.filename)
        # 增加时间戳防止同名文件冲突
        unique_filename = f"receipt_{int(datetime.now().timestamp())}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)
        return jsonify({"success": True, "url": f"/uploads/{unique_filename}"})
    
    return jsonify({"success": False, "message": "上传失败"}), 500


@app.route("/app")
def app_main():
    """主应用界面：左侧导航菜单 + 内容区域（角色管理等）"""
    user = session.get("user")
    if not user:
        return redirect(url_for("login"))
    return render_template("app.html", user=user)


@app.get("/api/dashboard/stats")
def api_dashboard_stats():
    """获取仪表盘统计数据"""
    start_date_str = request.args.get("start_date")  # 格式: YYYY-MM
    end_date_str = request.args.get("end_date")      # 格式: YYYY-MM
    customer_type = request.args.get("customer_type")  # 客户类型

    now = datetime.now()
    if not start_date_str:
        start_date_str = now.strftime("%Y-%m")
    if not end_date_str:
        end_date_str = now.strftime("%Y-%m")

    try:
        try:
            start_parts = start_date_str.split("-")
            end_parts = end_date_str.split("-")
            start_year, start_month = int(start_parts[0]), int(start_parts[1])
            end_year, end_month = int(end_parts[0]), int(end_parts[1])
        except (ValueError, IndexError):
            # 格式错误时回退到当前月
            now = datetime.now()
            start_year, start_month = now.year, now.month
            end_year, end_month = now.year, now.month
        
        start_dt = datetime(start_year, start_month, 1)
        if end_month == 12:
            end_dt = datetime(end_year + 1, 1, 1)
        else:
            end_dt = datetime(end_year, end_month + 1, 1)

        # 根据选择的客户类型，确定对应的字段
        id_field = Waybill.unit_customer_id
        if customer_type == "头程客户":
            id_field = Waybill.first_leg_customer_id
        elif customer_type == "尾程客户":
            id_field = Waybill.last_leg_customer_id
        elif customer_type == "差价客户":
            id_field = Waybill.differential_customer_id

        # 1. 单量数据统计 (日期范围) - 合并件数和重量查询以减少数据库交互
        vol_query = db.session.query(
            func.count(Waybill.id),
            func.sum(Waybill.weight)
        ).filter(
            Waybill.order_time >= start_dt,
            Waybill.order_time < end_dt
        ).filter(id_field != None)
        
        if customer_type:
            vol_query = vol_query.join(Customer, id_field == Customer.id).filter(
                Customer.customer_types.like(f"%{customer_type}%")
            )
            
        vol_res = vol_query.first()
        total_pieces = vol_res[0] or 0
        total_weight = vol_res[1] or 0
        
        # 产品分布 (饼图)
        product_stats_query = db.session.query(
            Product.name, func.count(Waybill.id)
        ).join(Waybill).filter(
            Waybill.order_time >= start_dt,
            Waybill.order_time < end_dt
        ).filter(id_field != None)
        
        if customer_type:
            product_stats_query = product_stats_query.join(Customer, id_field == Customer.id).filter(
                Customer.customer_types.like(f"%{customer_type}%")
            )
        product_stats = product_stats_query.group_by(Product.name).all()
        
        # 客户分布 (单量和重量)
        customer_stats_query = db.session.query(
            Customer.short_name, 
            func.count(Waybill.id).label('pieces'),
            func.sum(Waybill.weight).label('weight')
        ).join(Waybill, id_field == Customer.id).filter(
            Waybill.order_time >= start_dt,
            Waybill.order_time < end_dt
        )
        if customer_type:
            customer_stats_query = customer_stats_query.filter(
                Customer.customer_types.like(f"%{customer_type}%")
            )
        customer_stats = customer_stats_query.group_by(Customer.short_name).order_by(func.count(Waybill.id).desc()).all()

        # 2. 财务数据统计 (全局累计)
        total_receipts = db.session.query(func.sum(Payment.amount)).filter(Payment.payment_type == '收款').scalar() or 0
        total_payments = db.session.query(func.sum(Payment.amount)).filter(Payment.payment_type == '付款').scalar() or 0
        cash_balance = float(total_receipts) - float(total_payments)
        total_receivable = db.session.query(func.sum(Invoice.amount)).filter(Invoice.is_paid == False).scalar() or 0
        total_payable = db.session.query(func.sum(SupplierInvoice.amount)).filter(SupplierInvoice.is_paid == False).scalar() or 0

        return jsonify({
            "success": True,
            "volume": {
                "pieces": total_pieces,
                "weight": round(float(total_weight), 2),
                "product_distribution": [{"name": r[0], "value": r[1]} for r in product_stats],
                "customer_distribution": {
                    "names": [r[0] for r in customer_stats],
                    "pieces": [r[1] for r in customer_stats],
                    "weights": [round(float(r[2]), 2) if r[2] else 0 for r in customer_stats]
                }
            },
            "finance": {
                "cash_balance": round(float(cash_balance), 2),
                "receivable": round(float(total_receivable), 2),
                "payable": round(float(total_payable), 2)
            }
        })
    except Exception as e:
        print(f"Dashboard stats error: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.get("/api/dashboard/handling-fee")
def api_dashboard_handling_fee():
    """获取物流处理费统计（专线处理费）"""
    start_date_str = request.args.get("start_date")  # 格式: YYYY-MM
    end_date_str = request.args.get("end_date")      # 格式: YYYY-MM
    customer_id = request.args.get("customer_id")
    
    try:
        query = db.session.query(func.sum(Waybill.dedicated_line_fee)).filter(Waybill.differential_customer_id != None)
        
        # 时间过滤
        if start_date_str and start_date_str != "all":
            sy, sm = map(int, start_date_str.split("-"))
            start_dt = datetime(sy, sm, 1)
            query = query.filter(Waybill.order_time >= start_dt)
            
        if end_date_str and end_date_str != "all":
            ey, em = map(int, end_date_str.split("-"))
            if em == 12:
                end_dt = datetime(ey + 1, 1, 1)
            else:
                end_dt = datetime(ey, em + 1, 1)
            query = query.filter(Waybill.order_time < end_dt)
            
        # 客户过滤
        if customer_id and customer_id != "all":
            query = query.filter(Waybill.differential_customer_id == int(customer_id))
            
        total_fee = query.scalar() or 0
        return jsonify({"success": True, "total": round(float(total_fee), 2)})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.get("/api/dashboard/trend")
def api_dashboard_trend():
    """获取仪表盘趋势图数据"""
    # 同时支持 customer_ids[] 和 customer_ids
    customer_ids = request.args.getlist("customer_ids[]", type=int)
    if not customer_ids:
        customer_ids = request.args.getlist("customer_ids", type=int)
        
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")
    customer_type = request.args.get("customer_type")

    print(f"Trend Request: start={start_date_str}, end={end_date_str}, type={customer_type}, ids={customer_ids}")

    if not start_date_str or not end_date_str:
        return jsonify({"success": False, "message": "缺少日期范围"}), 400

    try:
        # 处理可能的无效日期字符串
        try:
            start_parts = start_date_str.split("-")
            end_parts = end_date_str.split("-")
            start_year, start_month = int(start_parts[0]), int(start_parts[1])
            end_year, end_month = int(end_parts[0]), int(end_parts[1])
        except (ValueError, IndexError):
            return jsonify({"success": False, "message": "日期格式不正确"}), 400

        start_dt = datetime(start_year, start_month, 1)
        if end_month == 12:
            end_dt = datetime(end_year + 1, 1, 1)
        else:
            end_dt = datetime(end_year, end_month + 1, 1)

        # 根据选择的客户类型，确定对应的字段
        id_field = Waybill.unit_customer_id
        if customer_type == "头程客户":
            id_field = Waybill.first_leg_customer_id
        elif customer_type == "尾程客户":
            id_field = Waybill.last_leg_customer_id
        elif customer_type == "差价客户":
            id_field = Waybill.differential_customer_id

        # 如果没有指定客户，获取该类别下趋势单量最多的一个
        if not customer_ids:
            # 这里的 join(Customer, id_field == Customer.id) 是关键，它排除了 NULL id_field
            top_customer = db.session.query(Customer.id).join(Waybill, id_field == Customer.id).filter(
                Waybill.order_time >= start_dt,
                Waybill.order_time < end_dt
            ).group_by(Customer.id).order_by(func.count(Waybill.id).desc()).first()
            
            if top_customer:
                customer_ids = [top_customer[0]]
                print(f"Auto-selected top customer: {customer_ids}")
            else:
                print("No top customer found for this range/type")
                # 生成日期序列即便没有数据也返回
                days = []
                curr = start_dt
                while curr < end_dt:
                    days.append(curr.strftime("%Y-%m-%d"))
                    curr += timedelta(days=1)
                return jsonify({"success": True, "series": [], "dates": days})

        # 获取涉及的客户名称
        customers = Customer.query.filter(Customer.id.in_(customer_ids)).all()
        customer_name_map = {c.id: c.short_name for c in customers}
        
        # 补全可能由于 type=int 转换失败导致的缺失名称（虽然理论上不会）
        for cid in customer_ids:
            if cid not in customer_name_map:
                c_obj = Customer.query.get(cid)
                if c_obj: customer_name_map[cid] = c_obj.short_name

        # 生成日期序列
        days = []
        curr = start_dt
        while curr < end_dt:
            days.append(curr.strftime("%Y-%m-%d"))
            curr += timedelta(days=1)

        # 批量查询数据
        trend_raw = db.session.query(
            id_field,
            func.date(Waybill.order_time).label('date'),
            func.count(Waybill.id)
        ).filter(
            id_field.in_(customer_ids),
            Waybill.order_time >= start_dt,
            Waybill.order_time < end_dt
        ).group_by(id_field, func.date(Waybill.order_time)).all()
        print(f"Trend Raw Results: {len(trend_raw)}")

        # 组织数据结构
        # { customer_id: { date_str: count } }
        data_map = {}
        for cid, dt, count in trend_raw:
            if cid is None: continue
            
            # 强制转为 int 避免类型不匹配问题
            try:
                cid_key = int(cid)
            except:
                cid_key = cid
                
            if cid_key not in data_map:
                data_map[cid_key] = {}
            
            # 处理日期格式
            if hasattr(dt, 'strftime'):
                d_str = dt.strftime("%Y-%m-%d")
            else:
                d_str = str(dt)[:10]
            data_map[cid_key][d_str] = count

        result_series = []
        for cid in customer_ids:
            try:
                cid_int = int(cid)
            except:
                cid_int = cid
                
            if cid_int not in customer_name_map:
                continue
                
            counts = [data_map.get(cid_int, {}).get(d, 0) for d in days]
            result_series.append({
                "name": customer_name_map[cid_int],
                "data": counts,
                "type": 'line',
                "smooth": True,
                "symbol": 'circle',
                "symbolSize": 6
            })

        return jsonify({
            "success": True,
            "dates": days,
            "series": result_series,
            "selected_customer_ids": [int(x) for x in customer_ids]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@app.get("/api/dashboard/unpaid-details")
def api_dashboard_unpaid_details():
    """获取未核销账单详情"""
    target_type = request.args.get("type")  # receivable or payable
    try:
        details = []
        if target_type == 'receivable':
            # 获取未付客户账单
            invoices = Invoice.query.filter_by(is_paid=False).order_by(Invoice.year.desc(), Invoice.month.desc()).all()
            for inv in invoices:
                details.append({
                    "id": inv.id,
                    "name": f"{inv.year}年{inv.month}月 - {inv.customer.short_name} - {inv.fee_type}",
                    "amount": float(inv.amount),
                    "period": f"{inv.year}年{inv.month}月"
                })
        elif target_type == 'payable':
            # 获取未付供应商账单
            invoices = SupplierInvoice.query.filter_by(is_paid=False).order_by(SupplierInvoice.year.desc(), SupplierInvoice.month.desc()).all()
            for inv in invoices:
                details.append({
                    "id": inv.id,
                    "name": f"{inv.year}年{inv.month}月 - {inv.supplier.short_name} 对账单",
                    "amount": float(inv.amount),
                    "period": f"{inv.year}年{inv.month}月"
                })
        
        return jsonify({"success": True, "details": details})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ==================== GitHub 自动部署 Webhook ====================
@app.post('/hooks/github')
def github_webhook():
    """处理 GitHub 的 Webhook 请求，自动部署代码更新"""
    # 验证请求来源
    signature = request.headers.get('X-Hub-Signature')
    if not signature:
        return jsonify({'success': False, 'message': 'Missing signature'}), 400
    
    # 获取密钥（建议从环境变量中读取）
    webhook_secret = os.environ.get('GITHUB_WEBHOOK_SECRET', 'your-webhook-secret-here')
    
    # 验证签名
    payload_body = request.data
    mac = hmac.new(
        webhook_secret.encode('utf-8'),
        msg=payload_body,
        digestmod=hashlib.sha1
    )
    expected_signature = 'sha1=' + mac.hexdigest()
    
    if not hmac.compare_digest(signature, expected_signature):
        return jsonify({'success': False, 'message': 'Invalid signature'}), 401
    
    # 解析请求体
    payload = request.json
    
    # 检查是否是 push 事件且推送到主分支
    if request.headers.get('X-GitHub-Event') == 'push':
        ref = payload.get('ref')
        if ref in ['refs/heads/main', 'refs/heads/master']:  # 支持常见的主分支名称
            try:
                # 执行 git pull 拉取最新代码
                result = subprocess.run(['git', 'pull'], cwd=os.getcwd(), capture_output=True, text=True)
                if result.returncode != 0:
                    return jsonify({'success': False, 'message': f'Git pull failed: {result.stderr}'})
                
                # 重新安装依赖（如果有变动）
                pip_result = subprocess.run(['pip', 'install', '-r', 'requirements.txt'], 
                                           cwd=os.getcwd(), capture_output=True, text=True)
                
                return jsonify({
                    'success': True, 
                    'message': 'Code updated successfully',
                    'git_output': result.stdout,
                    'pip_output': pip_result.stdout if 'pip_result' in locals() else ''
                })
            except Exception as e:
                return jsonify({'success': False, 'message': f'Deployment failed: {str(e)}'})
    
    return jsonify({'success': True, 'message': 'Event received but not processed'})


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True)
